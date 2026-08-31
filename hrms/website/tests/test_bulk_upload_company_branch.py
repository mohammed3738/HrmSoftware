"""
Regression tests for bulk Excel upload of Companies and Branches.
Synchronous (not chunked like the attendance upload) -- these lists are
realistically tens of rows, not thousands.

Run with: python manage.py test website.tests.test_bulk_upload_company_branch
"""
import io

import openpyxl
from django.contrib.auth.models import User, Group
from django.test import TestCase, Client
from django.urls import reverse

from website.models import Branch, Company


def _build_excel(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "upload.xlsx"
    return buf


class CompanyBulkUploadTest(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(username="cbu_admin", password="pass12345")
        self.admin.groups.add(Group.objects.get(name="Admin"))
        self.client = Client()
        self.client.login(username="cbu_admin", password="pass12345")

    def test_valid_rows_are_created(self):
        excel_file = _build_excel(
            ["Short Name", "Company Name", "Phone Number", "Email", "Company Address",
             "TAN Number", "PAN Number", "Employer PF Number", "PTRC Number", "PTEC Number", "ESIC Number"],
            [
                ["CBU1", "Company One", "1111111111", "one@test.com", "Addr 1", "", "", "", "", "", ""],
                ["CBU2", "Company Two", "2222222222", "two@test.com", "Addr 2", "TAN2", "PAN2", "PF2", "", "", ""],
            ],
        )
        resp = self.client.post(reverse("upload-company-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertTrue(data["success"], data)
        self.assertEqual(data["created"], 2)
        self.assertEqual(data["skipped"], 0)
        self.assertTrue(Company.objects.filter(short_name="CBU1", status="active").exists())
        c2 = Company.objects.get(short_name="CBU2")
        self.assertEqual(c2.tan_number, "TAN2")

    def test_row_missing_required_field_is_skipped_and_reported(self):
        excel_file = _build_excel(
            ["Short Name", "Company Name", "Phone Number", "Email", "Company Address"],
            [["CBU3", "", "3333333333", "three@test.com", "Addr 3"]],  # missing Company Name
        )
        resp = self.client.post(reverse("upload-company-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertTrue(data["success"], data)
        self.assertEqual(data["created"], 0)
        self.assertEqual(data["skipped"], 1)
        self.assertIn("Row 2", data["errors"][0])
        self.assertFalse(Company.objects.filter(short_name="CBU3").exists())

    def test_duplicate_short_name_against_existing_db_row_is_skipped(self):
        Company.objects.create(short_name="CBU4", name="Existing", phone="1", email="e@test.com", address="Addr")
        excel_file = _build_excel(
            ["Short Name", "Company Name", "Phone Number", "Email", "Company Address"],
            [["CBU4", "Duplicate", "4444444444", "dup@test.com", "Addr 4"]],
        )
        resp = self.client.post(reverse("upload-company-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertEqual(data["created"], 0)
        self.assertEqual(data["skipped"], 1)
        self.assertIn("already exists", data["errors"][0])
        self.assertEqual(Company.objects.filter(short_name__iexact="CBU4").count(), 1)

    def test_duplicate_short_name_within_same_file_is_skipped(self):
        excel_file = _build_excel(
            ["Short Name", "Company Name", "Phone Number", "Email", "Company Address"],
            [
                ["CBU5", "First", "5555555555", "first@test.com", "Addr"],
                ["CBU5", "Second", "6666666666", "second@test.com", "Addr"],
            ],
        )
        resp = self.client.post(reverse("upload-company-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertEqual(data["created"], 1)
        self.assertEqual(data["skipped"], 1)
        self.assertEqual(Company.objects.filter(short_name__iexact="CBU5").count(), 1)

    def test_manager_cannot_upload_companies(self):
        manager = User.objects.create_user(username="cbu_mgr", password="pass12345")
        manager.groups.add(Group.objects.get(name="Manager"))
        client = Client()
        client.login(username="cbu_mgr", password="pass12345")
        excel_file = _build_excel(
            ["Short Name", "Company Name", "Phone Number", "Email", "Company Address"],
            [["CBU6", "Blocked", "1", "b@test.com", "Addr"]],
        )
        resp = client.post(reverse("upload-company-excel"), {"excel_file": excel_file})
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(Company.objects.filter(short_name="CBU6").exists())

    def test_template_download_has_expected_headers(self):
        resp = self.client.get(reverse("download-company-upload-template"))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in next(wb.active.iter_rows(max_row=1))]
        self.assertEqual(headers, [
            "Short Name", "Company Name", "Phone Number", "Email", "Company Address",
            "TAN Number", "PAN Number", "Employer PF Number", "PTRC Number", "PTEC Number", "ESIC Number",
        ])


class BranchBulkUploadTest(TestCase):
    def setUp(self):
        self.hr = User.objects.create_user(username="bbu_hr", password="pass12345")
        self.hr.groups.add(Group.objects.get(name="HR"))
        self.client = Client()
        self.client.login(username="bbu_hr", password="pass12345")

    def test_valid_rows_are_created(self):
        excel_file = _build_excel(
            ["Branch Name", "Branch Address"],
            [["Branch One", "Addr 1"], ["Branch Two", ""]],
        )
        resp = self.client.post(reverse("upload-branch-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertTrue(data["success"], data)
        self.assertEqual(data["created"], 2)
        self.assertEqual(data["skipped"], 0)
        self.assertTrue(Branch.objects.filter(branch_name="Branch One", is_active=True).exists())

    def test_blank_branch_name_row_is_skipped(self):
        excel_file = _build_excel(
            ["Branch Name", "Branch Address"],
            [["", "No name here"]],
        )
        resp = self.client.post(reverse("upload-branch-excel"), {"excel_file": excel_file})
        data = resp.json()
        self.assertEqual(data["created"], 0)
        self.assertEqual(data["skipped"], 1)
        self.assertIn("Row 2", data["errors"][0])

    def test_hr_can_upload_branches(self):
        # branch_management:edit grants both Admin and HR (unlike
        # company_management:edit, which is Admin-only) -- confirm HR
        # specifically (not just Admin) can use this.
        excel_file = _build_excel(["Branch Name", "Branch Address"], [["HR Branch", "Addr"]])
        resp = self.client.post(reverse("upload-branch-excel"), {"excel_file": excel_file})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["success"])

    def test_employee_cannot_upload_branches(self):
        employee = User.objects.create_user(username="bbu_emp", password="pass12345")
        employee.groups.add(Group.objects.get(name="Employee"))
        client = Client()
        client.login(username="bbu_emp", password="pass12345")
        excel_file = _build_excel(["Branch Name", "Branch Address"], [["Blocked Branch", "Addr"]])
        resp = client.post(reverse("upload-branch-excel"), {"excel_file": excel_file})
        self.assertEqual(resp.status_code, 403)

    def test_template_download_has_expected_headers(self):
        resp = self.client.get(reverse("download-branch-upload-template"))
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        headers = [c.value for c in next(wb.active.iter_rows(max_row=1))]
        self.assertEqual(headers, ["Branch Name", "Branch Address"])
