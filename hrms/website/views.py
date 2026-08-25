from urllib import request

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .forms import *
from django.http import HttpResponse, Http404
# from .tasks import *
# Create your views here.
from django.utils.timezone import make_aware
import datetime
from django.contrib import messages
import pandas as pd
from django.http import HttpResponse
from .models import *  # Import your Employee model
from datetime import datetime,timedelta,time
from django.http import JsonResponse, HttpResponseBadRequest
from django.utils.timezone import now
import json
from django.views.decorators.http import require_http_methods
from django.contrib.auth import authenticate, login, logout
from datetime import date, timedelta
from decimal import Decimal
from django.core.paginator import Paginator
from dateutil.relativedelta import relativedelta
from django.template.loader import render_to_string
from django.utils import timezone    
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
# from .signals import recalculate_leave_balance_for_employee
from django.db.models import Q
import io
import os
from django.http import FileResponse, Http404
import openpyxl
from .services import *  # from previous services.py
from .utils.payroll_lock import (
    get_locking_run,
    get_locking_run_for_period,
    lock_response,
    build_date_locked_cache,
    date_in_cache,
    _lock_message,
)
from django.forms.models import model_to_dict
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from .utils.decorators import group_required

# def parse_time(time_value):
#     """Convert time string or float to a proper datetime.time object."""
#     if pd.isna(time_value) or time_value is None:
#         return None  # Return None if NaN or None
    
#     # If the value is a float (Excel's internal time format)
#     if isinstance(time_value, float):
#         hours = int(time_value * 24)  # Convert float to hours
#         minutes = int((time_value * 24 * 60) % 60)  # Convert to minutes
#         seconds = int((time_value * 24 * 3600) % 60)  # Convert to seconds
#         return timedelta(hours=hours, minutes=minutes, seconds=seconds).time()

#     # If it's a string, use strptime to convert it
#     try:
#         return datetime.strptime(time_value, "%H:%M:%S").time()
#     except ValueError:
#         print(f"⚠️ Invalid time format: {time_value}")
#         return None  # Return None if parsing fails

# atul
@login_required
@group_required("Admin", "HR")

def company_details_api(request, pk):
    company = get_object_or_404(Company, pk=pk)

    data = {
        "short_name": company.short_name,
        "name": company.name,
        "phone": company.phone,
        "email": company.email,
        "address": company.address,
        "tan_number": company.tan_number,
        "pan_number": company.pan_number,
        "employer_pf": company.employer_pf,
        "ptrc_number": company.ptrc_number,
        "ptec_number": company.ptec_number,
        "esic_number": company.esic_number,
        "status": company.status,
    }

    return JsonResponse(data)


@login_required
@group_required("Admin", "HR")
def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == "POST":
        employee.delete()
        messages.success(request, "Employee deleted successfully.")
        return redirect("create-employee")  

    messages.error(request, "Invalid request.")
    return redirect("create-employee")


# atul

# Vaishu
@login_required
@group_required("Admin", "HR")
def branch_details_api(request, pk):
    branch = get_object_or_404(Branch, pk=pk)

    data = {
        "branch_name": branch.branch_name,
        "branch_address": branch.branch_address,
        # "tan_number": company.tan_number,
        # "pan_number": company.pan_number,
        # "employer_pf": company.employer_pf,
        # "ptrc_number": company.ptrc_number,
        # "ptec_number": company.ptec_number,
        # "esic_number": company.esic_number,
        # "status": company.status,
    }

    return JsonResponse(data)
# Vaishu



def parse_time(time_value):
    """
    Convert time value to a proper datetime.time object.
    Handles multiple formats:
    - Excel time format (decimal 0-1)
    - String format "HH:MM:SS" or "HH:MM"
    - datetime.time objects
    - Pandas Timestamp objects
    """
    if pd.isna(time_value) or time_value is None:
        return None
    
    # If it's already a time object, return it
    if isinstance(time_value, time):
        return time_value
    
    # If it's a pandas Timestamp, extract the time
    if isinstance(time_value, pd.Timestamp):
        return time_value.time()
    
    # Convert to string and strip whitespace
    time_str = str(time_value).strip()
    
    if not time_str or time_str.lower() == 'nan':
        return None
    
    # Try to parse as string format (HH:MM:SS or HH:MM)
    for fmt in ["%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"]:
        try:
            return datetime.strptime(time_str, fmt).time()
        except ValueError:
            continue
    
    # If all string formats fail, try to handle Excel decimal format
    try:
        # Excel stores time as decimal where 0.5 = 12:00 PM
        time_float = float(time_value)
        if 0 <= time_float <= 1:
            seconds = int(time_float * 86400)  # 86400 seconds in a day
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            secs = seconds % 60
            return time(hours, minutes, secs)
    except (ValueError, TypeError):
        pass
    
    print(f"⚠️ Could not parse time value: {time_value} (type: {type(time_value).__name__})")
    return None





def fill_holiday_attendance_for_month(year, month):
    from calendar import monthrange
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    
    # Build per-company branch_specific_holidays flag
    company_branch_specific = {
        ps.company_id: getattr(ps, 'branch_specific_holidays', True)
        for ps in PayrollSettings.objects.all()
    }

    # Get all holidays in month
    all_holidays = Holiday.objects.filter(
        holiday_date__gte=first_day,
        holiday_date__lte=last_day
    ).select_related('holiday_calendar__branch')

    # Get half-day scenarios with branch info
    half_day_scenarios = HalfDayScenario.objects.filter(
        scenario_date__gte=first_day,
        scenario_date__lte=last_day,
        is_approved=True
    ).select_related('branch')

    # Build half-day dict: {date: set(branch_ids)}; None means "all branches"
    half_day_by_date = {}
    for s in half_day_scenarios:
        half_day_by_date.setdefault(s.scenario_date, set()).add(s.branch_id)  # branch_id=None → all

    employees = Employee.objects.filter(status='Active').select_related('branch', 'company')
    created_count = 0
    
    for employee in employees:
        dates_to_create = set()
        branch_specific = company_branch_specific.get(employee.company_id, True)

        # Determine which holiday dates apply to this employee
        for holiday in all_holidays:
            if holiday.is_national:
                # National holidays apply to everyone always
                dates_to_create.add(holiday.holiday_date)
            elif not branch_specific:
                # Branch-specific setting is OFF — all holidays apply to all
                dates_to_create.add(holiday.holiday_date)
            else:
                # Branch-specific ON — only apply if branch matches
                # holiday_calendar is None means the holiday has no branch restriction
                cal = holiday.holiday_calendar
                if cal is None or (employee.branch_id and cal.branch_id == employee.branch_id):
                    dates_to_create.add(holiday.holiday_date)
        
        # Half-day: None in branch_ids means "all branches" scenario
        for hd_date, branch_ids in half_day_by_date.items():
            if None in branch_ids or (employee.branch_id and employee.branch_id in branch_ids):
                dates_to_create.add(hd_date)
        
        for special_date in dates_to_create:
            if Attendance.objects.filter(
                employee=employee, date=special_date
            ).exists():
                continue
            
            attendance = Attendance(
                employee=employee,
                date=special_date,
                in_time=None,
                out_time=None,
            )
            attendance.save()
            
            if attendance.status in ('Holiday', 'Present (Half-Day)'):
                created_count += 1
            else:
                attendance.delete()
    
    return created_count

# @login_required
# @group_required("Admin", "HR")    
# def upload_attendance_excel(request):
#     if request.method != "POST" or not request.FILES.get("attendance_file"):
#         return JsonResponse(
#             {"success": False, "message": "No file uploaded!"},
#             status=400
#         )

#     file = request.FILES["attendance_file"]

#     try:
#         # Read Excel with proper time parsing
#         df = pd.read_excel(file)

#         # Clean column names
#         df.columns = df.columns.str.strip()

#         print(f"📄 Columns found: {df.columns.tolist()}")
#         print(f"📊 Data types:\n{df.dtypes}\n")

#         for index, row in df.iterrows():
#             employee_code = row.get("Employee Code")
#             attendance_date = row.get("Date")
#             in_time_raw = row.get("In Time")
#             out_time_raw = row.get("Out Time")

#             # Validate mandatory fields
#             if not employee_code or pd.isna(attendance_date):
#                 print(f"⚠️ Invalid row {index}, skipping")
#                 continue

#             # Parse date & time
#             attendance_date = pd.to_datetime(attendance_date).date()
#             in_time = parse_time(in_time_raw)
#             out_time = parse_time(out_time_raw)

#             print(f"🔍 Row {index}: Employee={employee_code}, Date={attendance_date}, "
#                   f"In={in_time}, Out={out_time}")

#             # Fetch employee
#             employee = Employee.objects.filter(employee_code=employee_code).first()
#             if not employee:
#                 print(f"⚠️ Employee not found: {employee_code}")
#                 continue

#             # 🔐 Prevent overwrite (ANTI-CHEAT)
#             if Attendance.objects.filter(employee=employee, date=attendance_date).exists():
#                 print(
#                     f"🔒 Attendance already exists for {employee_code} on {attendance_date}, skipped"
#                 )
#                 continue

#             # ✅ Create attendance (NO calculation here)
#             attendance = Attendance(
#                 employee=employee,
#                 date=attendance_date,
#                 in_time=in_time if in_time else None,
#                 out_time=out_time if out_time else None,
#             )

#             try:
#                 attendance.save()  # 🔥 shift-based calculation happens here
#                 print(
#                     f"✅ Saved: {employee_code} | {attendance_date} | "
#                     f"In: {attendance.in_time} | Out: {attendance.out_time} | "
#                     f"Status: {attendance.status} | Count: {attendance.count}"
#                 )
#             except Exception as e:
#                 print(
#                     f"❌ Error saving attendance for {employee_code} "
#                     f"on {attendance_date}: {str(e)}"
#                 )
#                 continue

#         uploaded_months = set()
#         for index, row in df.iterrows():
#             attendance_date_raw = row.get("Date")
#             if attendance_date_raw and not pd.isna(attendance_date_raw):
#                 d = pd.to_datetime(attendance_date_raw).date()
#                 uploaded_months.add((d.year, d.month))

#         for year, month in uploaded_months:
#             count = fill_holiday_attendance_for_month(year, month)
#             print(f"✅ Auto-created {count} holiday/half-day attendance records for {month}/{year}")


#         messages.success(request, "Attendance uploaded successfully!")
#         return JsonResponse(
#             {"success": True, "message": "Attendance uploaded successfully!"}
#         )

#     except Exception as e:
#         print(f"❌ File processing error: {str(e)}")
#         messages.error(request, f"Error processing file: {str(e)}")
#         return JsonResponse(
#             {"success": False, "message": str(e)}
#         )

ATTENDANCE_UPLOAD_CHUNK_SIZE = 200


@login_required
@group_required("Admin", "HR")
def upload_attendance_excel(request):
    if request.method == "GET":
        return render(request, "attendance/attendance_upload.html")

    # Legacy single-shot POST — kept for backward compatibility with any
    # external callers, but the UI now uses the chunked init/chunk flow below.
    try:
        return _do_import_attendance(request)
    except Exception as e:
        import traceback
        print(f"[ATTENDANCE IMPORT FATAL]\n{traceback.format_exc()}")
        return JsonResponse({"success": False, "error": str(e)})


def _parse_excel_time(val):
    """Parse a cell value into a datetime.time. Returns None if blank/invalid."""
    import datetime as dt
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    # Already a time object
    if isinstance(val, dt.time):
        return val
    # datetime / Timestamp → extract .time()
    if hasattr(val, "time") and callable(val.time):
        try:
            return val.time()
        except Exception:
            pass
    # Excel stores times as fraction of a day (float 0–1)
    if isinstance(val, float) and 0 <= val < 1:
        total_sec = int(round(val * 86400))
        return dt.time(total_sec // 3600, (total_sec % 3600) // 60, total_sec % 60)
    # String parsing
    s = str(val).strip()
    if s.lower() in ("", "nan", "none", "nat"):
        return None
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p"):
        try:
            return dt.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    try:
        t = pd.to_datetime(s, errors="coerce")
        if t is not None and not pd.isna(t):
            return t.time()
    except Exception:
        pass
    return None


def _excel_engine_for(uploaded_file):
    """Pick the right pandas engine from the filename extension: legacy
    .xls needs xlrd, modern .xlsx/.xlsm needs openpyxl."""
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    return "xlrd" if filename.endswith(".xls") else "openpyxl"


def _read_attendance_table(uploaded_file, header):
    """Read the uploaded attendance file into a DataFrame, dispatching to
    pandas' CSV or Excel reader based on the filename extension. `header` is
    the row index to use as column names, or None to read every row as data."""
    filename = (getattr(uploaded_file, "name", "") or "").lower()
    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file, header=header)
    engine = _excel_engine_for(uploaded_file)
    return pd.read_excel(uploaded_file, engine=engine, header=header)


def _normalize_attendance_excel(uploaded_file):
    """Parse an uploaded attendance file (.xlsx, .xls, or .csv) into a
    cleaned DataFrame with detected/normalized headers. Raises ValueError
    with a user-facing message on failure."""
    try:
        # Read without assuming a header row — check if row 0 is the header
        df_raw = _read_attendance_table(uploaded_file, header=None)
    except Exception as e:
        raise ValueError(f"Cannot read file: {e}")

    # ── Auto-detect header row ────────────────────────────────────────────────
    # Some exports put the header in row 0 as data (columns become Unnamed).
    # Check if the first row contains known header values.
    header_row = None
    for i in range(min(5, len(df_raw))):  # check first 5 rows
        row_vals = [str(v).strip().lower() for v in df_raw.iloc[i].tolist()]
        if "emp code" in row_vals or "employee code" in row_vals or "employee" in row_vals:
            header_row = i
            break

    if header_row is not None:
        # Re-read using the detected header row
        uploaded_file.seek(0)  # rewind file pointer
        df = _read_attendance_table(uploaded_file, header=header_row)
    else:
        df = df_raw

    # Coerce to str before stripping — if header detection above found
    # nothing, df's columns fall back to a plain 0..N RangeIndex (integers),
    # and pandas' .str accessor raises on non-string column dtypes.
    df.columns = [str(c).strip() for c in df.columns]

    # ── Normalize column names ────────────────────────────────────────────────
    # Handle variants: "Att. Date", "Att.Date", "Date", "Attendance Date", "AttendanceDate"
    col_aliases = {
        "emp code":        "Emp Code",
        "employee code":   "Emp Code",
        "employee":        "Emp Code",
        "att. date":       "Att.Date",
        "att.date":        "Att.Date",
        "attendance date": "Att.Date",
        "attendancedate":  "Att.Date",
        "date":            "Att.Date",
        "in time":         "In Time",
        "intime":          "In Time",
        "out time":        "Out Time",
        "outtime":         "Out Time",
    }
    df.rename(columns={c: col_aliases.get(c.strip().lower(), c) for c in df.columns}, inplace=True)

    # Verify required columns exist after normalization
    required = ["Emp Code", "Att.Date"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Missing column(s): {', '.join(missing)}. "
            f"Columns found after normalizing: {', '.join(df.columns.tolist())}"
        )

    return df


def _process_attendance_rows(df_chunk):
    """Process a slice of the normalized attendance DataFrame.
    Returns (created_count, updated_count, skipped_count, errors)."""
    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors = []

    # Build the lock cache lazily — we'll add company ids as we discover them
    _lock_cache = {}

    for row_idx, row in df_chunk.iterrows():
        row_num = row_idx + 2
        try:
            emp_code_raw = row.get("Emp Code")

            # Skip blank or header-repeat rows
            if emp_code_raw is None or str(emp_code_raw).strip().lower() in ("nan", "none", "", "emp code"):
                skipped_count += 1
                continue

            emp_code = str(emp_code_raw).strip()
            # Remove trailing ".0" if pandas read it as float (e.g. 2.0 → "2")
            if emp_code.endswith(".0"):
                emp_code = emp_code[:-2]

            employee = Employee.objects.filter(employee_code=emp_code).first()
            if not employee:
                errors.append(f"Row {row_num}: Employee code '{emp_code}' not found")
                skipped_count += 1
                continue

            att_date_raw = row.get("Att.Date")
            att_date = parse_excel_date(att_date_raw)
            if not att_date:
                errors.append(f"Row {row_num}: Invalid or missing date '{att_date_raw}'")
                skipped_count += 1
                continue

            if employee.company_id and employee.company_id not in _lock_cache:
                _lock_cache.update(build_date_locked_cache([employee.company_id]))
            if date_in_cache(_lock_cache, employee.company_id, att_date):
                errors.append(
                    f"Row {row_num}: {att_date:%d %b %Y} is part of a finalized payroll run — attendance cannot be modified."
                )
                skipped_count += 1
                continue

            in_time  = _parse_excel_time(row.get("In Time"))
            out_time = _parse_excel_time(row.get("Out Time"))

            # A re-upload of the same employee/date must overwrite the punch
            # times with the latest source data — get_or_create only applies
            # its defaults when creating a NEW row, so a previously-blank
            # row (e.g. left that way by a parsing bug) would never get
            # fixed by a corrected re-upload. Deliberately NOT using
            # Django's update_or_create() here: on the update path it calls
            # save(update_fields=<only the defaults keys>), which silently
            # drops status/count/late — calculate_status() computes them
            # in-memory but update_fields excludes them from the SQL UPDATE,
            # so the stale status would never actually persist. A plain
            # full .save() (no update_fields) writes every field. A manual
            # status override (status_overridden) is still preserved either
            # way — calculate_status() skips recomputation for those rows
            # regardless of what in/out time gets written here.
            try:
                attendance = Attendance.objects.get(employee=employee, date=att_date)
                created = False
            except Attendance.DoesNotExist:
                attendance = Attendance(employee=employee, date=att_date)
                created = True
            attendance.in_time = in_time
            attendance.out_time = out_time
            attendance.save()

            if created:
                created_count += 1
            else:
                updated_count += 1

        except Exception as e:
            import traceback
            print(f"[ATTENDANCE] Row {row_num} error: {traceback.format_exc()}")
            errors.append(f"Row {row_num}: {e}")
            skipped_count += 1

    return created_count, updated_count, skipped_count, errors


def _do_import_attendance(request):
    uploaded_file = request.FILES.get("file") or request.FILES.get("attendance_file")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "No file uploaded."})

    try:
        df = _normalize_attendance_excel(uploaded_file)
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)})

    created_count, updated_count, skipped_count, errors = _process_attendance_rows(df)

    return JsonResponse({
        "success": True,
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "total_rows": len(df),
        "errors": errors[:50],  # cap at 50 so response doesn't get huge
    })


@login_required
@group_required("Admin", "HR")
@require_http_methods(["POST"])
def upload_attendance_init(request):
    """
    Step 1 of the chunked upload flow: save the file and report how many
    rows it has, without processing any of them yet. The frontend then
    calls upload_attendance_chunk repeatedly to process it in small batches,
    so a single large file never has to be handled inside one long-running
    request (which is what caused uploads of 2000+ rows to fail/time out).
    """
    uploaded_file = request.FILES.get("file") or request.FILES.get("attendance_file")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "No file uploaded."})

    try:
        df = _normalize_attendance_excel(uploaded_file)
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)})

    uploaded_file.seek(0)
    upload = AttendanceUpload.objects.create(
        file=uploaded_file,
        total_rows=len(df),
        processed_rows=0,
        status="processing" if len(df) else "completed",
    )

    return JsonResponse({
        "success": True,
        "upload_id": upload.id,
        "total_rows": upload.total_rows,
        "chunk_size": ATTENDANCE_UPLOAD_CHUNK_SIZE,
    })


@login_required
@group_required("Admin", "HR")
@require_http_methods(["POST"])
def upload_attendance_chunk(request, upload_id):
    """Step 2 of the chunked upload flow: process the next batch of rows
    for a previously-initialized upload and report cumulative progress."""
    upload = get_object_or_404(AttendanceUpload, id=upload_id)

    if upload.status == "completed":
        return JsonResponse({
            "success": True, "done": True,
            "upload_id": upload.id,
            "processed_rows": upload.processed_rows,
            "total_rows": upload.total_rows,
            "created": upload.created_count,
            "updated": upload.updated_count,
            "skipped": upload.skipped_count,
            "errors": upload.errors,
        })

    try:
        upload.file.open("rb")
        try:
            df = _normalize_attendance_excel(upload.file)
        finally:
            upload.file.close()
    except ValueError as e:
        upload.status = "failed"
        upload.save(update_fields=["status"])
        return JsonResponse({"success": False, "error": str(e)})

    offset = upload.processed_rows
    chunk = df.iloc[offset: offset + ATTENDANCE_UPLOAD_CHUNK_SIZE]
    created_count, updated_count, skipped_count, errors = _process_attendance_rows(chunk)

    upload.processed_rows = min(offset + len(chunk), upload.total_rows)
    upload.created_count += created_count
    upload.updated_count += updated_count
    upload.skipped_count += skipped_count
    if errors:
        upload.errors = ((upload.errors or []) + errors)[:50]  # cap so it never grows unbounded
    done = upload.processed_rows >= upload.total_rows
    upload.status = "completed" if done else "processing"
    upload.save(update_fields=["processed_rows", "created_count", "updated_count", "skipped_count", "errors", "status"])

    return JsonResponse({
        "success": True,
        "done": done,
        "upload_id": upload.id,
        "processed_rows": upload.processed_rows,
        "total_rows": upload.total_rows,
        "created": upload.created_count,
        "updated": upload.updated_count,
        "skipped": upload.skipped_count,
        "errors": upload.errors,
    })


ATTENDANCE_RECALC_CHUNK_SIZE = 200


def _recalculate_attendance_queryset(request):
    """Build the (optionally date-filtered, company-scoped) Attendance
    queryset shared by the recalculate init/chunk endpoints."""
    qs = Attendance.objects.all()
    company = get_company_filter(request.user)
    if company:
        qs = qs.filter(employee__company=company)

    date_from = request.POST.get("date_from")
    date_to = request.POST.get("date_to")
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    return qs.order_by("id")


@login_required
@group_required("Admin", "HR")
@require_http_methods(["POST"])
def recalculate_attendance_init(request):
    """Step 1: report how many existing Attendance rows match the (optional)
    date range, without touching any of them yet."""
    total = _recalculate_attendance_queryset(request).count()
    return JsonResponse({
        "success": True,
        "total": total,
        "chunk_size": ATTENDANCE_RECALC_CHUNK_SIZE,
    })


@login_required
@group_required("Admin", "HR")
@require_http_methods(["POST"])
def recalculate_attendance_chunk(request):
    """Step 2: re-run calculate_status()/calculate_lateness() on the next
    batch of existing Attendance rows (via their normal .save(), so any
    calculation-logic fix takes effect immediately without deleting and
    re-uploading the Excel file). Offset-driven and stateless — the caller
    just keeps advancing `offset` by however many rows this returned until
    processed_in_chunk comes back as 0."""
    try:
        offset = int(request.POST.get("offset", 0))
    except (TypeError, ValueError):
        offset = 0

    qs = _recalculate_attendance_queryset(request)
    chunk = list(qs[offset: offset + ATTENDANCE_RECALC_CHUNK_SIZE])
    for attendance in chunk:
        attendance.save()  # triggers calculate_status() via Attendance.save()

    return JsonResponse({"success": True, "processed_in_chunk": len(chunk)})


LATE_REVIEW_SORT_FIELDS = {
    "employee": "employee__first_name",
    "code": "employee__employee_code",
    "date": "date",
    "in_time": "in_time",
    "out_time": "out_time",
    # No stored "hours worked" field, but for Late Present rows the shortfall
    # (late minutes) is a direct 1:1 inverse of hours worked, so it's sorted
    # via the existing `late` column with the direction flipped (see below).
    "worked": "late",
}


@login_required
@group_required("Admin", "HR", "Manager")
def late_attendance_review(request):
    """
    Payroll-time review of every 'Late Present' day for the selected month,
    with a per-row dropdown letting a manager convert a late day into
    Present / Half Day / Holiday (i.e. forgive it) or leave it as Late
    Present. Also reports each employee's total late-day count for the
    selected period, so a manager can decide e.g. "forgive 5 of this
    person's 10 late days this month."
    """
    company_filter = get_company_filter(request.user)

    base_qs = Attendance.objects.filter(status="Late Present").select_related("employee")
    if company_filter:
        base_qs = base_qs.filter(employee__company=company_filter)

    years = base_qs.dates("date", "year", order="DESC")
    raw_months_all = base_qs.dates("date", "month", order="DESC")
    seen_months = set()
    months = []
    for m in raw_months_all:
        if m.month not in seen_months:
            seen_months.add(m.month)
            months.append(m)

    # Show everything by default (no year/month filter) — a silent
    # "current month" default previously hid all data whenever the current
    # month had no late records, while the dropdown still rendered as if
    # "All Months" was selected, making the page look broken.
    selected_year = request.GET.get("year", "").strip()
    selected_month = request.GET.get("month", "").strip()

    qs = base_qs
    if selected_year:
        try:
            qs = qs.filter(date__year=int(selected_year))
        except ValueError:
            pass
    if selected_month:
        try:
            qs = qs.filter(date__month=int(selected_month))
        except ValueError:
            pass

    employee_search = request.GET.get("employee", "").strip()
    if employee_search:
        from django.db.models import Value
        from django.db.models.functions import Concat
        # Match first name, last name, "First Last" combined (autocomplete
        # fills the full name), or employee code.
        qs = qs.annotate(
            full_name=Concat("employee__first_name", Value(" "), "employee__last_name")
        ).filter(
            Q(full_name__icontains=employee_search) |
            Q(employee__employee_code__icontains=employee_search)
        )

    # Late-day count per employee within the current filters. Computed from
    # an unordered copy — order_by() fields leak into GROUP BY otherwise,
    # which would fragment the count down to one row per (employee, date).
    late_counts = dict(
        qs.order_by().values("employee_id")
        .annotate(late_count=Count("id"))
        .values_list("employee_id", "late_count")
    )

    sort_key = request.GET.get("sort", "")
    sort_dir = request.GET.get("dir", "asc")
    if sort_key in LATE_REVIEW_SORT_FIELDS:
        field = LATE_REVIEW_SORT_FIELDS[sort_key]
        effective_dir = sort_dir
        if sort_key == "worked":
            # "Hours Worked" ascending should mean fewest hours worked, which
            # is the MOST late-minutes — so the DB sort direction is inverted.
            effective_dir = "desc" if sort_dir == "asc" else "asc"
        order_field = field if effective_dir == "asc" else f"-{field}"
        qs = qs.order_by(order_field, "employee__first_name", "date")
    else:
        sort_key, sort_dir = "", "asc"
        qs = qs.order_by("employee__first_name", "employee__last_name", "date")

    paginator = Paginator(qs, 50)
    records = paginator.get_page(request.GET.get("page"))

    return render(request, "attendance/late_attendance_review.html", {
        "records": records,
        "late_counts": late_counts,
        "years": years,
        "months": months,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "employee_search": employee_search,
        "sort_key": sort_key,
        "sort_dir": sort_dir,
    })


ATTENDANCE_OVERRIDE_STATUS_MAP = {
    "Present":  Decimal("1.00"),
    "Half Day": Decimal("0.50"),
    "Holiday":  Decimal("1.00"),
}


ATTENDANCE_OVERRIDE_VALID_STATUSES = {"Present", "Half Day", "Holiday", "Late Present"}


def _override_attendance_status_item(attendance, new_status, company_filter=None):
    """Apply a single status conversion to one Attendance record. Returns an
    error message string on failure (permission/payroll lock), or None on
    success. Shared by the single-record and bulk override endpoints."""
    if company_filter and attendance.employee.company_id != company_filter.id:
        return "Permission denied."

    locking_run = get_locking_run(attendance.employee.company, attendance.date)
    if locking_run:
        return _lock_message(locking_run, action="change this attendance status")

    if new_status == "Late Present":
        attendance.status_overridden = False
        attendance.save()
    else:
        attendance.status_overridden = True
        attendance.status = new_status
        attendance.count = ATTENDANCE_OVERRIDE_STATUS_MAP[new_status]
        attendance.is_holiday = (new_status == "Holiday")
        attendance.is_half_day = (new_status == "Half Day")
        attendance.save()
    return None


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def override_attendance_status(request):
    """Manually convert a single Attendance record's status — used on the
    late-attendance review page to forgive/reclassify late days at payroll
    time. Choosing 'Late Present' clears any prior override and lets the
    normal calculate_status() logic run again."""
    attendance_id = request.POST.get("attendance_id")
    new_status = request.POST.get("new_status", "").strip()

    if new_status not in ATTENDANCE_OVERRIDE_VALID_STATUSES:
        return JsonResponse({"success": False, "error": "Invalid status choice."}, status=400)

    attendance = get_object_or_404(
        Attendance.objects.select_related("employee__company"), id=attendance_id
    )

    company_filter = get_company_filter(request.user)
    error = _override_attendance_status_item(attendance, new_status, company_filter)
    if error:
        status_code = 403 if error == "Permission denied." else 400
        return JsonResponse({"success": False, "error": error}, status=status_code)

    return JsonResponse({
        "success": True,
        "attendance_id": attendance.id,
        "status": attendance.status,
        "count": str(attendance.count),
    })


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def bulk_override_attendance_status(request):
    """Convert a batch of Attendance records to the same status in one
    action — the 'select multiple, apply one status' option on the
    late-attendance review page."""
    new_status = request.POST.get("new_status", "").strip()
    if request.content_type == "application/json":
        try:
            new_status = new_status or json.loads(request.body or "{}").get("new_status", "").strip()
        except ValueError:
            pass

    if new_status not in ATTENDANCE_OVERRIDE_VALID_STATUSES:
        return JsonResponse({"success": False, "error": "Invalid status choice."}, status=400)

    ids = _extract_bulk_ids(request)
    company_filter = get_company_filter(request.user)

    updated, failed = [], []
    attendances = Attendance.objects.select_related("employee__company").filter(id__in=ids)
    attendances_by_id = {a.id: a for a in attendances}

    for att_id in ids:
        attendance = attendances_by_id.get(att_id)
        if not attendance:
            failed.append({"id": att_id, "error": "Record not found."})
            continue
        error = _override_attendance_status_item(attendance, new_status, company_filter)
        if error:
            failed.append({"id": att_id, "error": error})
        else:
            updated.append(att_id)

    return JsonResponse({"success": True, "status": new_status, "updated": updated, "failed": failed})


def _current_payroll_period(company):
    """Return (period_from, period_to) for the payroll cycle containing
    today, for the given company — falls back to the calendar month if no
    custom cycle is configured, or (None, None) if there's no company/settings."""
    if not company:
        return None, None
    payroll_settings = PayrollSettings.objects.filter(company=company).first()
    if not payroll_settings:
        return None, None
    return get_payroll_period_for_date(payroll_settings, now().date())


def _validate_shift_dates_within_current_period(company, start_date_str, end_date_str):
    """Shift assignments may only be scheduled from today through the end of
    the current payroll period — not in the past, and not beyond the period
    that's currently running. Returns an error message, or None if OK (or if
    the company has no payroll cycle configured, in which case it doesn't block)."""
    period_from, period_to = _current_payroll_period(company)
    if not period_to:
        return None

    today = now().date()
    try:
        start = date.fromisoformat(start_date_str)
        end = date.fromisoformat(end_date_str)
    except (TypeError, ValueError):
        return "Invalid date."

    if start < today:
        return f"Start date cannot be before today ({today.strftime('%d %b %Y')})."
    if end > period_to:
        return f"End date cannot be after the current payroll period ends ({period_to.strftime('%d %b %Y')})."
    return None


@login_required
@group_required("Admin", "HR", "Manager")
def shift_roster_list(request):
    """List/filter shift roster assignments — 'who's on which shift, and
    when' at a glance, for planning day/night/rotational shifts. Purely a
    scheduling record; does not feed attendance calculation."""
    company_filter = get_company_filter(request.user)

    base_qs = ShiftAssignment.objects.select_related("employee", "created_by")
    if company_filter:
        base_qs = base_qs.filter(employee__company=company_filter)

    qs = base_qs

    employee_search = request.GET.get("employee", "").strip()
    if employee_search:
        from django.db.models import Value
        from django.db.models.functions import Concat
        qs = qs.annotate(
            full_name=Concat("employee__first_name", Value(" "), "employee__last_name")
        ).filter(
            Q(full_name__icontains=employee_search) |
            Q(employee__employee_code__icontains=employee_search)
        )

    shift_name = request.GET.get("shift_name", "").strip()
    if shift_name:
        qs = qs.filter(shift_name=shift_name)

    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    if date_from:
        qs = qs.filter(end_date__gte=date_from)
    if date_to:
        qs = qs.filter(start_date__lte=date_to)

    shift_names = list(
        base_qs.order_by("shift_name").values_list("shift_name", flat=True).distinct()
    )

    qs = qs.order_by("-start_date", "employee__first_name")
    paginator = Paginator(qs, 50)
    assignments = paginator.get_page(request.GET.get("page"))

    today = now().date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    # Bounds for the Assign/Edit date pickers: today through the end of the
    # current payroll period. Based on the logged-in user's own company when
    # scoped; for global-access users this is just a UI hint — the actual
    # enforcement in add/edit happens per employee's own company.
    _, period_to = _current_payroll_period(company_filter or get_user_company(request.user))

    return render(request, "attendance/shift_roster.html", {
        "assignments": assignments,
        "employee_search": employee_search,
        "shift_name": shift_name,
        "date_from": date_from,
        "date_to": date_to,
        "shift_names": shift_names,
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "roster_min_date": today.isoformat(),
        "roster_max_date": period_to.isoformat() if period_to else "",
    })


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def add_shift_assignment(request):
    """Assign one shift (name + date range) to one or more employees at
    once — HR typically rosters a whole team's rotation in one action."""
    from django.core.exceptions import ValidationError

    employee_ids = request.POST.getlist("employee_ids")
    shift_name = request.POST.get("shift_name", "").strip()
    start_date = request.POST.get("start_date", "").strip()
    end_date = request.POST.get("end_date", "").strip()
    shift_start_time = request.POST.get("shift_start_time", "").strip() or None
    shift_end_time = request.POST.get("shift_end_time", "").strip() or None
    notes = request.POST.get("notes", "").strip()

    if not employee_ids:
        return JsonResponse({"success": False, "error": "Select at least one employee."}, status=400)
    if not shift_name or not start_date or not end_date:
        return JsonResponse({"success": False, "error": "Shift name, start date, and end date are required."}, status=400)

    company_filter = get_company_filter(request.user)
    created, failed = [], []

    for emp_id in employee_ids:
        try:
            employee = Employee.objects.get(id=emp_id)
        except (Employee.DoesNotExist, ValueError):
            failed.append({"employee_id": emp_id, "error": "Employee not found."})
            continue
        if company_filter and employee.company_id != company_filter.id:
            failed.append({"employee_id": emp_id, "employee": str(employee), "error": "Permission denied."})
            continue

        window_error = _validate_shift_dates_within_current_period(employee.company, start_date, end_date)
        if window_error:
            failed.append({"employee_id": emp_id, "employee": str(employee), "error": window_error})
            continue

        assignment = ShiftAssignment(
            employee=employee, shift_name=shift_name,
            start_date=start_date, end_date=end_date,
            shift_start_time=shift_start_time, shift_end_time=shift_end_time,
            notes=notes, created_by=request.user,
        )
        try:
            assignment.full_clean()
            assignment.save()
            created.append(assignment.id)
        except ValidationError as e:
            failed.append({
                "employee_id": emp_id, "employee": str(employee),
                "error": "; ".join(e.messages),
            })

    return JsonResponse({"success": True, "created": created, "failed": failed})


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def edit_shift_assignment(request, pk):
    from django.core.exceptions import ValidationError

    assignment = get_object_or_404(ShiftAssignment.objects.select_related("employee"), id=pk)

    company_filter = get_company_filter(request.user)
    if company_filter and assignment.employee.company_id != company_filter.id:
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)

    new_start_date = request.POST.get("start_date", "").strip()
    new_end_date = request.POST.get("end_date", "").strip()
    window_error = _validate_shift_dates_within_current_period(
        assignment.employee.company, new_start_date, new_end_date
    )
    if window_error:
        return JsonResponse({"success": False, "error": window_error}, status=400)

    assignment.shift_name = request.POST.get("shift_name", "").strip()
    assignment.start_date = new_start_date
    assignment.end_date = new_end_date
    assignment.shift_start_time = request.POST.get("shift_start_time", "").strip() or None
    assignment.shift_end_time = request.POST.get("shift_end_time", "").strip() or None
    assignment.notes = request.POST.get("notes", "").strip()

    try:
        assignment.full_clean()
        assignment.save()
    except ValidationError as e:
        return JsonResponse({"success": False, "error": "; ".join(e.messages)}, status=400)

    return JsonResponse({"success": True})


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def delete_shift_assignment(request, pk):
    assignment = get_object_or_404(ShiftAssignment.objects.select_related("employee"), id=pk)
    company_filter = get_company_filter(request.user)
    if company_filter and assignment.employee.company_id != company_filter.id:
        return JsonResponse({"success": False, "error": "Permission denied."}, status=403)
    assignment.delete()
    return JsonResponse({"success": True})


@login_required
def api_get_shift_assignment(request, pk):
    assignment = get_object_or_404(
        ShiftAssignment.objects.select_related("employee", "created_by"), id=pk
    )
    return JsonResponse({
        "id": assignment.id,
        "employee_id": assignment.employee_id,
        "employee_name": f"{assignment.employee.first_name} {assignment.employee.last_name}",
        "employee_code": assignment.employee.employee_code,
        "shift_name": assignment.shift_name,
        "start_date": assignment.start_date.isoformat(),
        "end_date": assignment.end_date.isoformat(),
        "shift_start_time": assignment.shift_start_time.strftime("%H:%M") if assignment.shift_start_time else "",
        "shift_end_time": assignment.shift_end_time.strftime("%H:%M") if assignment.shift_end_time else "",
        "notes": assignment.notes,
        "created_by": assignment.created_by.get_full_name() or assignment.created_by.username if assignment.created_by else "—",
        "created_at": assignment.created_at.strftime("%d %b %Y, %I:%M %p"),
    })


@login_required
def attendance_upload_progress(request, upload_id):
    try:
        upload = AttendanceUpload.objects.get(id=upload_id)
        progress = 0 if upload.total_rows == 0 else int((upload.processed_rows / upload.total_rows) * 100)
        return JsonResponse({"progress": progress, "status": upload.status})
    except AttendanceUpload.DoesNotExist:
        return JsonResponse({"progress": 0, "status": "not_found"})

    
@login_required
@group_required("Admin", "HR", "Manager")
def attendance_list(request):
    user = request.user

    # 🔁 If employee → redirect to own attendance page
    if user.groups.filter(name="Employee").exists():
        try:
            employee = user.employee_profile  # OneToOne relation
            return redirect(
                "employee_attendance_detail",
                employee_id=employee.id
            )
        except Exception:
            # Safety fallback if employee profile missing
            return redirect("dashboard")


    """View attendance by selected date (default: today)"""
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = date.fromisoformat(date_str)
        except ValueError:
            selected_date = now().date()
    else:
        selected_date = now().date()

    attendance_records = Attendance.objects.filter(date=selected_date)

    # For navigation buttons
    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)

    context = {
        "attendance_records": attendance_records,
        "selected_date": selected_date.isoformat(),
        "prev_date": prev_date.isoformat(),
        "next_date": next_date.isoformat(),
    }
    return render(request, "attendance/today.html", context)


@login_required
def employee_attendance_detail(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    attendance_records = Attendance.objects.filter(employee=employee)

    # If the company has a custom payroll cycle configured (e.g. 27th to
    # 26th), filter by that actual payroll period instead of the naive
    # calendar month — "June" should mean 27 May - 26 Jun, not 1-30 Jun.
    payroll_settings = PayrollSettings.objects.filter(company=employee.company).first()
    uses_custom_period = bool(payroll_settings and payroll_settings.from_date and payroll_settings.to_date)

    selected_year = request.GET.get("year")
    selected_month = request.GET.get("month")
    selected_period = request.GET.get("period", "")
    years = months = []
    payroll_periods = []

    # Bounds of the currently selected range, if any — used both to filter
    # the attendance table and to scope the leave-taken summary below to
    # the same window (a payroll period, a calendar month/year, or "all time").
    range_start = range_end = None

    if uses_custom_period:
        payroll_periods = get_all_payroll_periods_from_attendance(employee.company, payroll_settings)
        if selected_period:
            for p in payroll_periods:
                if p["to_date"].isoformat() == selected_period:
                    range_start, range_end = p["from_date"], p["to_date"]
                    attendance_records = attendance_records.filter(
                        date__gte=range_start, date__lte=range_end
                    )
                    break
    else:
        dates_qs = Attendance.objects.filter(employee=employee)
        years = dates_qs.dates("date", "year", order="DESC")
        months = dates_qs.dates("date", "month", order="DESC")
        if selected_year:
            attendance_records = attendance_records.filter(date__year=selected_year)
        if selected_month:
            attendance_records = attendance_records.filter(date__month=selected_month)
        if selected_year and selected_month:
            range_start = date(int(selected_year), int(selected_month), 1)
            range_end = (range_start + relativedelta(months=1)) - timedelta(days=1)
        elif selected_year:
            range_start = date(int(selected_year), 1, 1)
            range_end = date(int(selected_year), 12, 31)

    # ── Summary counters (Present / Late / Half Day / Leaves Taken) ────────
    days_present_count = attendance_records.filter(status__in=["Present", "Late Present"]).count()
    late_remarks_count = attendance_records.filter(status="Late Present").count()
    half_days_count = attendance_records.filter(status__in=["Half Day", "Present (Half-Day)"]).count()

    leaves_qs = LeaveApplication.objects.filter(employee=employee, status="Approved")
    if range_start and range_end:
        leaves_qs = leaves_qs.filter(start_date__lte=range_end, end_date__gte=range_start)
        leaves_taken = sum(
            (min(leave.end_date, range_end) - max(leave.start_date, range_start)).days + 1
            for leave in leaves_qs
        )
    else:
        leaves_taken = sum(leave.total_days() for leave in leaves_qs)

    attendance_records = attendance_records.only(
        "date", "in_time", "out_time", "status"
    ).order_by("-date")

    paginator = Paginator(attendance_records, 30)
    page = request.GET.get("page")
    attendance_records = paginator.get_page(page)

    context = {
        "employee": employee,
        "attendance_records": attendance_records,
        "years": years,
        "months": months,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "uses_custom_period": uses_custom_period,
        "payroll_periods": payroll_periods,
        "selected_period": selected_period,
        "payroll_settings": payroll_settings,
        "days_present_count": days_present_count,
        "late_remarks_count": late_remarks_count,
        "half_days_count": half_days_count,
        "leaves_taken": leaves_taken,
    }
    return render(request, "attendance/employee_attendance_detail.html", context)



@login_required
def employee_search(request):
    q = request.GET.get("q", "").strip()

    employees = Employee.objects.filter(
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(employee_code__icontains=q)
    )[:10]

    data = [
        {
            "id": emp.id,
            "name": f"{emp.first_name} {emp.last_name}",
            "code": emp.employee_code
        }
        for emp in employees
    ]

    return JsonResponse(data, safe=False)


@login_required
def submit_correction_request(request):
    if request.method == "POST":
        attendance_id = request.POST.get("attendance_id")
        new_in_time = request.POST.get("new_in_time")
        new_out_time = request.POST.get("new_out_time")
        reason = request.POST.get("reason")

        attendance = get_object_or_404(Attendance, id=attendance_id)

        # Block raising a correction request for a finalized payroll period
        locking_run = get_locking_run(attendance.employee.company, attendance.date)
        if locking_run:
            return lock_response(locking_run, action="raise a correction request")

        # Create a correction request
        correction_request = AttendanceCorrectionRequest.objects.create(
            attendance=attendance,
            # requested_by=request.user,
            old_in_time=attendance.in_time,
            old_out_time=attendance.out_time,
            new_in_time=new_in_time,
            new_out_time=new_out_time,
            reason=reason,
        )
        
        return JsonResponse({"message": "Correction request submitted successfully!"})


# approval
def _approve_correction_item(correction_request):
    """Approve a single AttendanceCorrectionRequest and write the corrected
    times onto its Attendance record. Returns an error message string on
    failure (e.g. payroll lock), or None on success."""
    attendance = correction_request.attendance
    locking_run = get_locking_run(attendance.employee.company, attendance.date)
    if locking_run:
        return _lock_message(locking_run, action="approve this correction")

    attendance.in_time = correction_request.new_in_time
    attendance.out_time = correction_request.new_out_time
    attendance.save()

    correction_request.status = "Approved"
    correction_request.reviewed_at = now()
    correction_request.save()
    return None


@login_required
def approve_correction_request(request, request_id):
    correction_request = get_object_or_404(AttendanceCorrectionRequest, id=request_id)

    error = _approve_correction_item(correction_request)
    if error:
        return JsonResponse({"success": False, "error": error}, status=400)
    return JsonResponse({"message": "Correction Approved!"})


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def bulk_approve_correction(request):
    ids = _extract_bulk_ids(request)
    approved, failed = [], []
    for req_id in ids:
        try:
            correction_request = AttendanceCorrectionRequest.objects.get(id=req_id)
        except AttendanceCorrectionRequest.DoesNotExist:
            failed.append({"id": req_id, "error": "Request not found."})
            continue
        error = _approve_correction_item(correction_request)
        if error:
            failed.append({"id": req_id, "error": error})
        else:
            approved.append(req_id)
    return JsonResponse({"success": True, "approved": approved, "failed": failed})


@login_required
def reject_correction_request(request, request_id):
    correction_request = get_object_or_404(AttendanceCorrectionRequest, id=request_id)

    if request.method == "POST":
        data = json.loads(request.body)
        rejection_reason = data.get("reason", "")

        if not rejection_reason:
            return JsonResponse({"error": "Rejection reason is required."}, status=400)

        # Update request as rejected
        correction_request.status = "Rejected"
        correction_request.rejection_reason = rejection_reason  # Save reason for employee reference
        correction_request.reviewed_at = now()
        correction_request.save()

        return JsonResponse({"message": "Correction Request Rejected!"})

    return JsonResponse({"error": "Invalid request"}, status=400)


# --- Date / Time helpers (platform-safe) ---
def _format_date(d):
    """Return 'Dec. 5, 2025' or None."""
    if not d:
        return None
    month_abbr = d.strftime("%b")  # e.g. "Dec"
    return f"{month_abbr}. {d.day}, {d.year}"

def _format_time(t):
    """Return '9 a.m.' or '1:30 p.m.' or '—' for None."""
    if not t:
        return "—"
    hour_24 = t.hour
    minute = t.minute
    period = "a.m." if hour_24 < 12 else "p.m."
    hour_12 = hour_24 % 12
    if hour_12 == 0:
        hour_12 = 12
    if minute == 0:
        return f"{hour_12} {period}"
    return f"{hour_12}:{minute:02d} {period}"

def _format_date_iso(d):
    """Return ISO string or None."""
    return d.isoformat() if d else None

@login_required
@group_required("Admin", "HR", "Manager")
def attendance_correction_requests_list(request):
    from .models import AttendanceCorrectionRequest, Attendance

    # Base queryset
    base_qs = AttendanceCorrectionRequest.objects.select_related("attendance__employee")

    # Dropdown values from Attendance table
    years = Attendance.objects.dates("date", "year", order="DESC")

    raw_months_all = Attendance.objects.dates("date", "month", order="DESC")

    # --- Deduplicate months (2024-Nov & 2025-Nov appear only once) ---
    seen_months = set()
    months_all = []
    for m in raw_months_all:
        if m.month not in seen_months:
            seen_months.add(m.month)
            months_all.append(m)

    # Read filters
    selected_year = request.GET.get("year", "").strip()
    selected_month = request.GET.get("month", "").strip()

    # Start with all correction requests
    qs = base_qs

    # ----------------
    # YEAR FILTER
    # ----------------
    if selected_year:
        try:
            year_int = int(selected_year)
            qs = qs.filter(attendance__date__year=year_int)

            # Build month dropdown only for that year (also deduped)
            raw_months_year = Attendance.objects.filter(date__year=year_int).dates(
                "date", "month", order="DESC"
            )

            seen = set()
            months_qs = []
            for m in raw_months_year:
                if m.month not in seen:
                    seen.add(m.month)
                    months_qs.append(m)

        except ValueError:
            months_qs = months_all
    else:
        months_qs = months_all

    if selected_month:
        try:
            month_int = int(selected_month)
            qs = qs.filter(attendance__date__month=month_int)
        except ValueError:
            pass

    # Final ordering
    qs = qs.order_by("-created_at")

    # Build rows (your same structure)
    rows = []
    for req in qs:
        att = req.attendance
        emp = att.employee

        rows.append({
            "id": req.id,
            "emp_name": f"{emp.first_name} {emp.last_name}".strip(),
            "emp_code": emp.employee_code,
            "shift_date": att.date,
            "shift_date_str": _format_date(att.date),
            "old_in_time": _format_time(req.old_in_time),
            "old_out_time": _format_time(req.old_out_time),
            "new_in_time": _format_time(req.new_in_time),
            "new_out_time": _format_time(req.new_out_time),
            "status": req.status,
            "created_at": _format_date_iso(req.created_at),
        })

    # Render
    return render(request, "attendance/attendance_request_status.html", {
        "years": years,
        "months": months_qs,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "attendance_requests": rows,
    })

@login_required
@group_required("Admin", "HR", "Manager")
def attendance_correction_detail(request, pk):
    obj = get_object_or_404(AttendanceCorrectionRequest, pk=pk)

    att = obj.attendance  # related Attendance object
    employee = att.employee  # related Employee object

    data = {
        "attendance": {
            "id": att.id,
            "date": att.date.isoformat() if att.date else None,
            "in_time": att.in_time.isoformat() if att.in_time else None,
            "out_time": att.out_time.isoformat() if att.out_time else None,
            "status": att.status,
        },
        "old_in_time": obj.old_in_time.isoformat() if obj.old_in_time else None,
        "old_out_time": obj.old_out_time.isoformat() if obj.old_out_time else None,
        "new_in_time": obj.new_in_time.isoformat() if obj.new_in_time else None,
        "new_out_time": obj.new_out_time.isoformat() if obj.new_out_time else None,
        "reason": obj.reason,
        "rejection_reason": obj.rejection_reason,
        "created_at": obj.created_at.strftime("%Y-%m-%d") if obj.created_at else None,
        "status": obj.status,
        "employee": {
            "id": employee.id,
            "first_name": employee.first_name,
            "last_name": employee.last_name,
            "employee_code": employee.employee_code,
        },
    }

    return JsonResponse(data)


# def comp_off_requests_list(request): 
#     # Pull attendance correction requests and avoid N+1 queries 
#     attendance_requests = CompOffRequest.objects.select_related( 'employee' ) 
#     rows = [] 
#     for r in attendance_requests: 
#         employee = r.employee # Employee object 
#         # employee = attendance.employee # Employee object 
#         # Build clean row for frontend table 
#         rows.append({ 
#             "id": r.id, 
#             "emp_name": f"{employee.first_name} {employee.last_name}".strip(), 
#             "emp_code": getattr(employee, "employee_code", "—"), 
#             "from_date": r.from_date,
#             "to_date": r.to_date,
#             "reason": r.reason,
#             "status": r.status,
#             }) 
#             # Debug print (optional) 
#             # for row in rows: 
#             # print(row) 
#         return render(request, "attendance/comp_off_request.html", { "attendance_requests": rows }) 
from django.db.models import Count
from django.db.models.functions import ExtractMonth, ExtractYear
import calendar        

# def comp_off_requests_list(request):

#     # read optional filters from GET
#     selected_year = request.GET.get("year")   # e.g. "2025" or empty
#     selected_month = request.GET.get("month") # e.g. "12" or empty

#     # base queryset: only requests that have from_date (safe)
#     qs = CompOffRequest.objects.filter(from_date__isnull=False).select_related("employee")

#     # annotate year/month using from_date
#     qs = qs.annotate(year=ExtractYear("from_date"), month=ExtractMonth("from_date"))

#     # apply filters if provided
#     if selected_year:
#         try:
#             qs = qs.filter(year=int(selected_year))
#         except ValueError:
#             pass
#     if selected_month:
#         try:
#             qs = qs.filter(month=int(selected_month))
#         except ValueError:
#             pass

#     # group by employee + year + month and count requests
#     grouped = (
#         qs.values(
#             "employee__id",
#             "employee__first_name",
#             "employee__last_name",
#             "employee__employee_code",
#             "year",
#             "month",
#         )
#         .annotate(requests_count=Count("id"))
#         .order_by("employee__employee_code", "year", "month")
#     )

#     # build rows for frontend
#     rows = []
#     for g in grouped:
#         month_num = g["month"]
#         month_name = calendar.month_abbr[month_num] if month_num else "—"
#         emp_name = f"{g['employee__first_name'] or ''} {g['employee__last_name'] or ''}".strip()

#         rows.append(
#             {
#                 "emp_id": g["employee__id"],
#                 "emp_code": g.get("employee__employee_code") or "—",
#                 "emp_name": emp_name or "—",
#                 "year": g.get("year"),
#                 "month": month_name,
#                 "month_number": month_num,
#                 "count": g["requests_count"],  # number of comp-off requests in that month
#             }
#         )

#     return render(
#         request,
#         "attendance/comp_off_request.html",
#         {
#             "comp_off_requests": rows,
#             "selected_year": selected_year,
#             "selected_month": selected_month,
#         },
#     )


import calendar
from django.db.models import Count
from django.db.models.functions import ExtractYear, ExtractMonth
from django.shortcuts import render

@login_required
@group_required("Admin", "HR", "Manager")
def comp_off_requests_list(request):
    # read optional filters from GET
    selected_year = request.GET.get("year")   # e.g. "2025" or ""
    selected_month = request.GET.get("month") # e.g. "12" or ""

    # base queryset: only requests that have from_date (safe)
    qs = CompOffRequest.objects.filter(from_date__isnull=False).select_related("employee")

    # annotate year/month using from_date
    qs = qs.annotate(year=ExtractYear("from_date"), month=ExtractMonth("from_date"))

    # Build distinct year/month lists from the annotated queryset for the dropdowns
    years_qs = qs.values("year").distinct().order_by("-year")
    years = [{"year": y["year"]} for y in years_qs if y.get("year") is not None]

    months_qs = qs.values("month").distinct().order_by("month")
    # convert numeric month to name for template convenience
    months = []
    for m in months_qs:
        mn = m.get("month")
        if mn:
            months.append({"month": mn, "name": calendar.month_name[mn]})

    # apply filters if provided
    if selected_year:
        try:
            qs = qs.filter(year=int(selected_year))
        except ValueError:
            pass
    if selected_month:
        try:
            qs = qs.filter(month=int(selected_month))
        except ValueError:
            pass

    # group by employee + year + month and count requests
    grouped = (
        qs.values(
            "employee__id",
            "employee__first_name",
            "employee__last_name",
            "employee__employee_code",
            "year",
            "month",
        )
        .annotate(requests_count=Count("id"))
        .order_by("employee__employee_code", "year", "month")
    )

    # build rows for frontend
    rows = []
    for g in grouped:
        month_num = g["month"]
        month_name = calendar.month_abbr[month_num] if month_num else "—"
        emp_name = f"{g['employee__first_name'] or ''} {g['employee__last_name'] or ''}".strip()

        rows.append(
            {
                "emp_id": g["employee__id"],
                "emp_code": g.get("employee__employee_code") or "—",
                "emp_name": emp_name or "—",
                "year": g.get("year"),
                "month": month_name,
                "month_number": month_num,
                "count": g["requests_count"],
            }
        )

    return render(
        request,
        "attendance/comp_off_request.html",
        {
            "comp_off_requests": rows,
            "selected_year": selected_year,
            "selected_month": selected_month,
            "years": years,
            "months": months,
        },
    )

from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from datetime import datetime
import datetime as dt


@login_required
@group_required("Admin", "HR", "Manager")
def comp_off_requests(request, pk):
    """
    Return all comp-off requests for the employee with id=pk for the requested month/year.
    Expects optional GET params: ?year=2025&month=12
    If year/month are missing, falls back to the current month/year.
    Renders a template fragment that can be loaded into a modal.
    """
    # ensure employee exists (optional, helpful for clearer 404)
    employee = get_object_or_404(Employee, pk=pk)

    # read optional filters from GET; fallback to current month/year
    selected_year = request.GET.get("year")
    selected_month = request.GET.get("month")

    now = timezone.now().date()
    try:
        year = int(selected_year) if selected_year else now.year
    except (TypeError, ValueError):
        year = now.year

    try:
        month = int(selected_month) if selected_month else now.month
    except (TypeError, ValueError):
        month = now.month

    # filter requests for the employee and month/year
    qs = (
        CompOffRequest.objects.filter(employee__id=pk, from_date__isnull=False)
        .filter(from_date__year=year, from_date__month=month)
        .order_by("from_date")
    )

    # build rows for the modal table (individual requests)
    rows = []
    for r in qs:
        rows.append(
            {
                "id": r.id,
                "from_date": r.from_date,
                "to_date": r.to_date,
                "days_count": r.count,    # the per-request day span from your model
                "reason": r.reason,
                "status": r.status,
                "rejection_reason": r.rejection_reason
            }
        )

    context = {
        "employee": employee,
        "year": year,
        "month": month,
        # "month_name": datetime.date(year, month, 1).strftime("%b"),
        "month_name": dt.date(year, month, 1).strftime("%b"),
        "comp_off_entries": rows,
    }

    # Render a fragment suitable for modal body
    return render(request, "attendance/comp_off_request_table.html", context)

@login_required(login_url="login")
@group_required("Admin", "HR")
def admin_dashboard(request):
    today = date.today()
    user = request.user
    user_company = get_user_company(user)
    company_filter = get_company_filter(user)

    base_qs = Employee.objects.filter(company=company_filter) if company_filter else Employee.objects.all()
    active_qs = base_qs.filter(status='Active')

    # ── KPI numbers ──────────────────────────────────────────────
    total_active = active_qs.count()
    total_left   = base_qs.filter(status='Left').count()

    new_joiners_month = active_qs.filter(
        date_of_joining__year=today.year,
        date_of_joining__month=today.month,
    ).count()

    pending_leaves      = LeaveApplication.objects.filter(employee__in=active_qs, status='Pending').count()
    pending_compoffs    = CompOffRequest.objects.filter(employee__in=active_qs, status='Pending').count()
    pending_corrections = AttendanceCorrectionRequest.objects.filter(
        attendance__employee__in=active_qs, status='Pending'
    ).count()
    total_pending       = pending_leaves + pending_compoffs + pending_corrections

    # ── Today's attendance breakdown ─────────────────────────────
    today_attn   = Attendance.objects.filter(employee__in=active_qs, date=today)
    present_today   = today_attn.filter(status__in=['Present', 'Late Present', 'Half Day']).count()
    absent_today    = today_attn.filter(status='Absent').count()
    late_today      = today_attn.filter(status='Late Present').count()
    on_leave_today  = today_attn.filter(status='On Leave').count()
    not_marked      = total_active - today_attn.count()

    # ── Department distribution ───────────────────────────────────
    dept_data = list(
        active_qs.exclude(department__isnull=True).exclude(department='')
        .values('department').annotate(count=Count('id')).order_by('-count')[:8]
    )

    # ── Detail lists ──────────────────────────────────────────────
    recent_joiners = active_qs.filter(date_of_joining__isnull=False).order_by('-date_of_joining')[:6]

    anniversary_employees = active_qs.filter(
        date_of_joining__month=today.month,
        date_of_joining__isnull=False,
    ).exclude(date_of_joining__year=today.year).order_by('date_of_joining__day')[:5]

    # Capped at 50 (rather than showing every pending row) so the dashboard
    # stays fast, while still being generous enough that "select all" in the
    # bulk-approve UI covers realistic pending queues, not just a handful.
    leave_requests = (
        LeaveApplication.objects.select_related('employee')
        .filter(employee__in=active_qs, status='Pending').order_by('-id')[:50]
    )
    compoff_requests = (
        CompOffRequest.objects.select_related('employee')
        .filter(employee__in=active_qs, status='Pending').order_by('-id')[:50]
    )
    correction_requests = (
        AttendanceCorrectionRequest.objects.select_related('attendance__employee')
        .filter(attendance__employee__in=active_qs, status='Pending').order_by('-id')[:50]
    )
    upcoming_offboarding = (
        Offboarding.objects.select_related('employee')
        .filter(date_of_relieving__gte=today, date_of_relieving__lte=today + timedelta(days=30))
        .order_by('date_of_relieving')[:5]
    )

    return render(request, 'd/f.html', {
        'user': user,
        'today': today,
        'user_company': user_company,
        # KPI
        'total_active': total_active,
        'total_left': total_left,
        'new_joiners_month': new_joiners_month,
        'total_pending': total_pending,
        'present_today': present_today,
        'absent_today': absent_today,
        'late_today': late_today,
        'on_leave_today': on_leave_today,
        'not_marked': not_marked,
        # Pending counts
        'pending_leaves': pending_leaves,
        'pending_compoffs': pending_compoffs,
        'pending_corrections': pending_corrections,
        # Lists
        'recent_joiners': recent_joiners,
        'anniversary_employees': anniversary_employees,
        'leave_requests': leave_requests,
        'compoff_requests': compoff_requests,
        'correction_requests': correction_requests,
        'upcoming_offboarding': upcoming_offboarding,
        'dept_data': dept_data,
    })


def _approve_compoff_item(compoff):
    """Approve a single CompOffRequest. Returns an error message string on
    failure (e.g. payroll lock), or None on success."""
    locking_run = get_locking_run_for_period(
        compoff.employee.company, compoff.from_date, compoff.to_date
    )
    if locking_run:
        return _lock_message(locking_run, action="approve this comp-off")

    compoff.status = "Approved"
    compoff.save()
    return None


def _extract_bulk_ids(request):
    """Read a list of ids from a bulk-action POST — accepts a JSON body
    ({"ids": [...]}) or a regular form POST with repeated 'ids' fields."""
    if request.content_type == "application/json":
        try:
            data = json.loads(request.body or "{}")
        except ValueError:
            data = {}
        raw_ids = data.get("ids", [])
    else:
        raw_ids = request.POST.getlist("ids")
    ids = []
    for raw in raw_ids:
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    return ids


@login_required
@group_required("Admin", "HR", "Manager")
def approve_compoff(request, compoff_id):
    try:
        compoff = CompOffRequest.objects.get(id=compoff_id)
    except CompOffRequest.DoesNotExist:
        return JsonResponse({"message": "Request not found!"}, status=404)

    error = _approve_compoff_item(compoff)
    if error:
        return JsonResponse({"success": False, "error": error}, status=400)
    return JsonResponse({"message": "CompOff request approved successfully!"})


@login_required
@group_required("Admin", "HR", "Manager")
@require_http_methods(["POST"])
def bulk_approve_compoff(request):
    ids = _extract_bulk_ids(request)
    approved, failed = [], []
    for compoff_id in ids:
        try:
            compoff = CompOffRequest.objects.get(id=compoff_id)
        except CompOffRequest.DoesNotExist:
            failed.append({"id": compoff_id, "error": "Request not found."})
            continue
        error = _approve_compoff_item(compoff)
        if error:
            failed.append({"id": compoff_id, "error": error})
        else:
            approved.append(compoff_id)
    return JsonResponse({"success": True, "approved": approved, "failed": failed})

# @csrf_exempt
# def reject_compoff(request, compoff_id):
#     if request.method == "POST":
#         try:
#             compoff = CompOffRequest.objects.get(id=compoff_id)
#             data = json.loads(request.body)
#             compoff.status = "Rejected"
#             compoff.rejection_reason = data.get("reason", "No reason provided")
#             compoff.save()
#             return JsonResponse({"message": "CompOff request rejected successfully!"})
#         except CompOffRequest.DoesNotExist:
#             return JsonResponse({"message": "Request not found!"}, status=404)

@login_required
@group_required("Admin", "HR", "Manager")
def reject_compoff(request, compoff_id):
    correction_request = get_object_or_404(CompOffRequest, id=compoff_id)

    if request.method == "POST":
        data = json.loads(request.body)
        rejection_reason = data.get("reason", "")

        if not rejection_reason:
            return JsonResponse({"error": "Rejection reason is required."}, status=400)

        # Update request as rejected
        correction_request.status = "Rejected"
        correction_request.rejection_reason = rejection_reason  # Save reason for employee reference
        correction_request.reviewed_at = now()
        correction_request.save()

        return JsonResponse({"message": "Correction Request Rejected!"})

    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
@group_required("Admin", "HR")
def download_employees_excel(request):
    # Fetch active employees
    employees = Employee.objects.filter(status="Active").values(
        "employee_code", "first_name", "middle_name", "last_name", "father_name",
        "gender", "blood_group", "date_of_birth", "place_of_birth",
        "personal_email", "personal_mobile", "present_address", "permanent_address",
        "date_of_marriage", "designation", "department", "date_of_joining",
        "date_of_confirmation", "location", "shift_start_time", "shift_end_time",
        "pan_no", "aadhar_no", "voter_id", "passport", "uan_no", "pf_no", "esic_no",
        "name_as_per_bank", "salary_account_number", "ifsc_code",
        "emergency_contact_name1", "emergency_contact_relation1", "emergency_contact_mobile1",
        "emergency_contact_name2", "emergency_contact_relation2", "emergency_contact_mobile2"
    )

    # Convert QuerySet to Pandas DataFrame
    df = pd.DataFrame(list(employees))

    # Rename columns for better readability
    df.rename(columns={
        "employee_code": "Employee Code",
        "first_name": "First Name",      
        "middle_name": "Middle Name",
        "last_name": "Last Name",
        "father_name": "Father's Name",
        "gender": "Gender",
        "blood_group": "Blood Group",
        "date_of_birth": "Date of Birth",
        "place_of_birth": "Place of Birth",
        "personal_email": "Personal Email",
        "personal_mobile": "Personal Mobile",
        "present_address": "Present Address",
        "permanent_address": "Permanent Address",
        "date_of_marriage": "Date of Marriage",
        "designation": "Designation",
        "department": "Department",
        "date_of_joining": "Date of Joining",
        "date_of_confirmation": "Date of Confirmation",
        "location": "Location",
        # "shift": "Shift",
        "shift_start_time": "Shift Start Time",
        "shift_end_time": "Shift End Time",
        "pan_no": "PAN Number",
        "aadhar_no": "Aadhar Number",
        "voter_id": "Voter ID",
        "passport": "Passport",
        "uan_no": "UAN Number",
        "pf_no": "PF Number",
        "esic_no": "ESIC Number",
        "name_as_per_bank": "Bank Name",
        "salary_account_number": "Salary Account Number",
        "ifsc_code": "IFSC Code",
        "emergency_contact_name1": "Emergency Contact Name 1",
        "emergency_contact_relation1": "Emergency Contact Relation 1",
        "emergency_contact_mobile1": "Emergency Contact Mobile 1",
        "emergency_contact_name2": "Emergency Contact Name 2",
        "emergency_contact_relation2": "Emergency Contact Relation 2",
        "emergency_contact_mobile2": "Emergency Contact Mobile 2",
    }, inplace=True)

    # Create HTTP response with Excel content type
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="active_employees.xlsx"'

    # Save DataFrame to Excel in response
    with pd.ExcelWriter(response, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Active Employees", index=False)

    return response


@login_required
@group_required("Admin", "HR")
def download_leave_excel(request):
    # Fetch active leave balances along with related employee info
    leave_balances = LeaveBalance.objects.select_related("employee").filter(employee__status="Active")

    data = []
    for lb in leave_balances:
        emp = lb.employee
        data.append({
            "Employee Code": emp.employee_code,
            "First Name": emp.first_name,
            "Middle Name": emp.middle_name,
            "Last Name": emp.last_name,
            "Father's Name": emp.father_name,
            "Gender": emp.gender,
            "Blood Group": emp.blood_group,
            "Date of Birth": emp.date_of_birth,
            "Place of Birth": emp.place_of_birth,
            "Personal Email": emp.personal_email,
            "Personal Mobile": emp.personal_mobile,
            "Present Address": emp.present_address,
            "Permanent Address": emp.permanent_address,
            "Date of Marriage": emp.date_of_marriage,
            "Designation": emp.designation,
            "Department": emp.department,
            "Date of Joining": emp.date_of_joining,
            "Date of Confirmation": emp.date_of_confirmation,
            "Location": emp.location,
            "Shift": emp.shift,
            "PAN Number": emp.pan_no,
            "Aadhar Number": emp.aadhar_no,
            "Voter ID": emp.voter_id,
            "Passport": emp.passport,
            "UAN Number": emp.uan_no,
            "PF Number": emp.pf_no,
            "ESIC Number": emp.esic_no,
            "Bank Name": emp.name_as_per_bank,
            "Salary Account Number": emp.salary_account_number,
            "IFSC Code": emp.ifsc_code,
            "Emergency Contact Name 1": emp.emergency_contact_name1,
            "Emergency Contact Relation 1": emp.emergency_contact_relation1,
            "Emergency Contact Mobile 1": emp.emergency_contact_mobile1,
            "Emergency Contact Name 2": emp.emergency_contact_name2,
            "Emergency Contact Relation 2": emp.emergency_contact_relation2,
            "Emergency Contact Mobile 2": emp.emergency_contact_mobile2,
            # Leave-specific fields
            "Opening Balance": lb.opening_balance,
            "Leave Taken": lb.leave_taken,
            "Number of Days Present": lb.number_of_days_present,
            "Total Number of Days": lb.total_number_of_days,
            "Late": lb.late,
            "Comp Off": lb.compoff,
            "Leave Without Pay": lb.leave_without_pay,
            "Closing Balance": lb.closing_balance,
            "Leave Balance": lb.leave_balance,
        })

    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Create response with Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="leave_balance.xlsx"'

    with pd.ExcelWriter(response, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Active Leave Balance", index=False)

    return response


def test(request):
    test_func.delay()

    employees = Employee.objects.filter(status='Active')
    return HttpResponse("Done")


@login_required
def home(request):
    return render(request, 'base2.html')


# working code 
# def create_employee(request):
#     if request.method == 'POST':
#         form = EmployeeForm(request.POST, request.FILES)
#         if form.is_valid():
#             # Get unsaved employee instance
#             employee = form.save(commit=False)
#             # Get branch id from POST and assign the branch if valid
#             branch_id = request.POST.get('branch')
#             if branch_id:
#                 try:
#                     branch = Branch.objects.get(id=branch_id)
#                     employee.branch = branch
#                 except Branch.DoesNotExist:
#                     pass
#             # Instantiate formset with the employee instance
#             formset = PreviousEmploymentFormSet(request.POST, instance=employee)
#             if formset.is_valid():
#                 employee.save()       # Save the employee first
#                 formset.save()        # Then save all previous employment records
#                 messages.success(request, "Employee created successfully!")
#                 return redirect('home')
#             else:
#                 messages.error(request, "There were errors in the previous employment details.")
#                 print("Formset errors:", formset.errors)
#                 print("Formset non-form errors:", formset.non_form_errors())
#         else:
#             messages.error(request, "There were errors in the employee details.")
#             print("Form errors:", form.errors)
#             # Instantiate formset with POST data to re-render errors
#             formset = PreviousEmploymentFormSet(request.POST)
#     else:
#         form = EmployeeForm()
#         formset = PreviousEmploymentFormSet(instance=Employee())

#     # For context, also retrieve all employees and branches if needed in the template.
#     employees = Employee.objects.all()
#     branches = Branch.objects.all()
#     active_employees = Employee.objects.filter(status="Active")
#     employee_count = active_employees.count()
#     inactive_employees = Employee.objects.filter(status="Left ")
#     employee_count_inactive = inactive_employees.count()
#     companies = Company.objects.all()
#     context = {
#         'form': form,
#         'formset': formset,
#         'employees': employees,
#         'branches': branches,
#         'employee_count': employee_count,
#         'employee_count_inactive': employee_count_inactive,
#         'companies': companies,
#     }
#     return render(request, 'employee/create_employee2.html', context)

# working
# def create_employee(request):
#     if request.method == 'POST':
#         form = EmployeeForm(request.POST, request.FILES)
        
#         if form.is_valid():
#             employee = form.save(commit=False)

#             # Assign branch if selected
#             branch_id = request.POST.get("branch")
#             if branch_id:
#                 try:
#                     employee.branch = Branch.objects.get(id=branch_id)
#                 except Branch.DoesNotExist:
#                     pass

#             # Formset for previous employments
#             prev_formset = PreviousEmploymentFormSet(request.POST, request.FILES, instance=employee)

#             # ---- NEW: Prepare Attachment Formsets for each previous employment row ----
#             attachment_sets = []  # store child formsets

#             if prev_formset.is_valid():
#                 # Save employee first, then previous rows
#                 employee.save()
#                 prev_instances = prev_formset.save(commit=False)

#                 for index, prev_obj in enumerate(prev_instances):
#                     prev_obj.employee = employee  # attach FK
#                     prev_obj.save()

#                     # Detect nested attachment forms by prefix
#                     prefix = f"attach-{index}"

#                     attach_formset = AttachmentFormSet(
#                         request.POST,
#                         request.FILES,
#                         prefix=prefix,
#                         instance=prev_obj
#                     )

#                     attachment_sets.append(attach_formset)

#                 # Validate all nested attachment formsets
#                 if all([fs.is_valid() for fs in attachment_sets]):
#                     # Save attachments
#                     for fs in attachment_sets:
#                         fs.save()

#                     messages.success(request, "Employee created successfully with nested attachments!")
#                     return redirect('home')
#                 else:
#                     messages.error(request, "Error in nested attachments")
#                     print("Attachment errors:", [fs.errors for fs in attachment_sets])

#             else:
#                 messages.error(request, "There were validation errors in previous employment.")
#                 print("Previous Employment Errors:", prev_formset.errors)

#         else:
#             messages.error(request, "Employee form has errors.")
#             print("Employee Form Errors:", form.errors)

#         # If failing validation → rebuild formsets for re-render
#         prev_formset = PreviousEmploymentFormSet(request.POST, request.FILES)

#     else:
#         # GET request → empty forms
#         form = EmployeeForm()
#         prev_formset = PreviousEmploymentFormSet(instance=Employee())

#     # Context
#     context = {
#         'form': form,
#         'formset': prev_formset,
#         'employees': Employee.objects.all(),
#         'branches': Branch.objects.all(),
#         'employee_count': Employee.objects.filter(status="Active").count(),
#         'employee_count_inactive': Employee.objects.filter(status="Left ").count(),
#         'companies': Company.objects.all(),
#     }

#     return render(request, 'employee/create_employee2.html', context)

from django.contrib.auth.models import Group

# Add to context in both views:

from django.contrib.auth.models import Group, User

@login_required
@group_required("Admin", "HR")
def create_or_edit_employee(request, employee_id=None):
    employee = None
    is_edit = False

    if employee_id:
        employee = get_object_or_404(Employee, id=employee_id)
        is_edit = True

    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        formset = PreviousEmploymentFormSet(request.POST, instance=employee)
        attachment_formset = AttachmentFormSet(request.POST, request.FILES, instance=employee)

        if form.is_valid() and formset.is_valid() and attachment_formset.is_valid():
            with transaction.atomic():
                emp_obj = form.save(commit=False)
                emp_obj.save()

                # ✅ HANDLE GROUP ASSIGNMENT
                selected_group_id = request.POST.get("employee_group")
                if selected_group_id:
                    # Create or get the linked User for this employee
                    if not emp_obj.user:
                        # Auto-create a user only if employee_code exists
                        if emp_obj.employee_code:
                            username = emp_obj.employee_code.lower()
                            user, created = User.objects.get_or_create(username=username)
                            if created:
                                user.set_unusable_password()
                                user.save()
                            emp_obj.user = user
                            emp_obj.save()

                    # Clear old groups and assign new one (only if user exists)
                    if emp_obj.user:
                        emp_obj.user.groups.clear()
                        try:
                            group = Group.objects.get(id=selected_group_id)
                            emp_obj.user.groups.add(group)
                        except Group.DoesNotExist:
                            pass

                # ... rest of your existing formset saving logic
                formset.instance = emp_obj
                attachment_formset.instance = emp_obj

                for f in formset.forms:
                    if not f.cleaned_data or f.cleaned_data.get("DELETE"):
                        continue
                    record = f.instance
                    record.employee = emp_obj
                    record.save()

                    form_index = f.prefix.split("-")[-1]
                    prefix = f"attach-{form_index}-"
                    file_keys = [k for k in request.FILES.keys() if k.startswith(prefix)]

                    for key in file_keys:
                        uploaded_file = request.FILES.get(key)
                        if not uploaded_file:
                            continue
                        file_index = key.split("-")[2]
                        doc_name = request.POST.get(f"attach-{form_index}-{file_index}-document_name", "")
                        PreviousEmploymentAttachment.objects.create(
                            previous_employment=record,
                            file=uploaded_file,
                            document_name=doc_name
                        )

                attachments = attachment_formset.save(commit=False)
                has_attachment_error = False
                for attachment in attachments:
                    attachment.employee = emp_obj
                    if attachment.file_name == "other" and not attachment.other_file_name:
                        messages.error(request, "Please enter a document name for the 'Other' file type in KYC tab.")
                        has_attachment_error = True
                        break
                    attachment.save()

                if not has_attachment_error:
                    for obj in attachment_formset.deleted_objects:
                        obj.delete()

                    messages.success(
                        request,
                        "Employee updated successfully!" if is_edit else "Employee created successfully!"
                    )
                    return redirect("employee_create")

        else:
            messages.error(request, "Please correct the errors below.")

    else:
        form = EmployeeForm(instance=employee)
        formset = PreviousEmploymentFormSet(instance=employee)
        attachment_formset = AttachmentFormSet(instance=employee)

    # ✅ Get current group of employee's user (for pre-selecting in edit mode)
    current_group_id = None
    if employee and employee.user:
        first_group = employee.user.groups.first()
        current_group_id = first_group.id if first_group else None

    return render(request, "employee/create_employee2.html", {
        "form": form,
        "formset": formset,
        "attachment_formset": attachment_formset,
        "employees": Employee.objects.all(),
        "is_edit": is_edit,
        "employee": employee,
        "groups": Group.objects.all(),
        "current_group_id": current_group_id,
        "all_groups": Group.objects.all(),
        "open_modal": request.method == "POST",
    })

from django.contrib.auth.models import Group

# Add to context in both views:

# @permission_required("website.view_employee", raise_exception=True)
@login_required
@group_required("Admin", "HR")
def employee_list(request):
    return render(request, "employee/create_employee2.html", {
        "form": EmployeeForm(),
        "formset": PreviousEmploymentFormSet(),
        "attachment_formset": AttachmentFormSet(),
        "employees": Employee.objects.all(),
        "is_edit": False,
        "groups": Group.objects.all(),
        "current_group_id": None,
        "all_groups": Group.objects.all(),
    })




# def employee_detail(request, pk):
#     employee = get_object_or_404(Employee, pk=pk)
#     employeement= PreviousEmployment.objects.filter(employee=employee)
#     first_name = getattr(employee, "first_name", None) or (getattr(user, "first_name", None) if user else None)
#     # last_name  = getattr(employee, "last_name", None)  or (getattr(user, "last_name", None) if user else None)
#     context = {
#         'employee': employee,
#         'employeement': employeement,
#         'display': {
#             'first_name': first_name,
#         }
#     }
#     print('Employee Details:', employee)
#     print('Previous Employeement Details:', employeement)
#     return render(request, 'employee/employee_detail.html', context)


@login_required
@group_required("Admin", "HR")
@require_http_methods(["POST"])
def bulk_employee_action(request):
    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

    employee_ids = data.get("employee_ids", [])
    new_status   = data.get("status", "").strip()
    new_role     = data.get("role", "").strip()

    if not employee_ids:
        return JsonResponse({"success": False, "error": "No employees selected."}, status=400)
    if not new_status and not new_role:
        return JsonResponse({"success": False, "error": "No action specified."}, status=400)

    valid_statuses = {"Active", "Pending", "Left"}
    if new_status and new_status not in valid_statuses:
        return JsonResponse({"success": False, "error": f"Invalid status: {new_status}"}, status=400)

    employees = Employee.objects.filter(id__in=employee_ids)
    updated = 0

    for emp in employees:
        changed = False
        if new_status:
            emp.status = new_status
            changed = True
        if new_role and emp.user:
            emp.user.groups.clear()
            try:
                group = Group.objects.get(name=new_role)
                emp.user.groups.add(group)
                changed = True
            except Group.DoesNotExist:
                pass
        if changed:
            emp.save(update_fields=["status"] if new_status else [])
            updated += 1

    return JsonResponse({"success": True, "updated": updated})

from django.shortcuts import get_object_or_404, render

@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    # employeement = PreviousEmployment.objects.filter(employee=employee)
    # previous_employments = employee.previous_employments.all().order_by('-to_date')
    previous_employments = PreviousEmployment.objects.filter(employee=employee).order_by('-to_date')
    previous_employments_attachment = PreviousEmploymentAttachment.objects.filter(previous_employment__employee=employee)
    # attachments = employee.attachments.all().order_by('-uploaded_at')
    attachments = EmployeeAttachment.objects.filter(employee=employee).order_by('-uploaded_at')
    salary = SalaryMaster.objects.filter(employee=employee, is_active=True).first()
    # safe related user (if Employee has a user FK/OneToOne)
    user = getattr(employee, "user", None)

    # safe lookups with fallbacks
    first_name = getattr(employee, "first_name", None) or (getattr(user, "first_name", None) if user else None)
    last_name  = getattr(employee, "last_name", None)  or (getattr(user, "last_name", None) if user else None)
    email      = getattr(employee, "email", None)      or (getattr(user, "email", None) if user else None)
    phone      = getattr(employee, "phone", None)      or getattr(employee, "contact", None) or ""

    full_name = " ".join(filter(None, [first_name, last_name])) or getattr(employee, "full_name", None) or str(employee)

    # photo url safe
    photo_url = None
    photo_field = getattr(employee, "photo", None) or getattr(employee, "avatar", None)
    if photo_field:
        try:
            photo_url = photo_field.url
        except Exception:
            photo_url = None

    context = {
        'employee': employee,
        'previous_employments': previous_employments,
        'attachments': attachments,
        'previous_employments_attachment': previous_employments_attachment,
        'salary': salary,
        'display': {
            'first_name': first_name,
            'last_name': last_name,
            'full_name': full_name,
            'email': email,
            'phone': phone,
            'photo_url': photo_url,
        }
    }

    # debug prints (ok to remove later)
    print('Employee Details:', employee)
    print('Previous Employment Details:', previous_employments)
    print('Display dict:', context['display'])

    return render(request, 'employee/employee_detail.html', context)




@login_required
def my_profile(request):
    if not hasattr(request.user, 'employee_profile'):
        return redirect('dashboard')  # or raise PermissionDenied

    employee = request.user.employee_profile
    return redirect('employee_detail', pk=employee.pk)

@login_required
def download_attachment(request, pk):
    attachment = get_object_or_404(EmployeeAttachment, pk=pk)

    # OPTIONAL: extra permission checks
    # if not request.user.has_perm('yourapp.view_attachment') or attachment.owner != request.user:
    #     raise Http404()

    if not attachment.file:
        raise Http404("File not found")

    # open file for streaming
    try:
        file_handle = attachment.file.open('rb')
    except Exception:
        raise Http404("Unable to open file")

    filename = os.path.basename(attachment.file.name)
    response = FileResponse(file_handle, as_attachment=True, filename=filename)
    return response


# def employee_edit(request, pk):
#     employee = get_object_or_404(Employee, pk=pk)
#     form = EmployeeForm(instance=employee)
#     prev_formset = PreviousEmploymentFormSet(instance=employee)
#     attachment_formset = AttachmentFormSet(instance=employee)

#     if request.method == "POST":
#         form = EmployeeForm(request.POST, request.FILES, instance=employee)
#         prev_formset = PreviousEmploymentFormSet(request.POST, request.FILES, instance=employee)
#         attachment_formset = AttachmentFormSet(request.POST, request.FILES, instance=employee)

#         if form.is_valid() and prev_formset.is_valid() and attachment_formset.is_valid():
#             form.save()
#             prev_formset.save()
#             attachment_formset.save()
#             return redirect("employee_list")

#     return render(request, "employee/create_employee2.html", {
#         "form": form,
#         "formset": prev_formset,
#         "attachment_formset": attachment_formset
#     })

# def create_offboarding(request):
#     if request.method == 'POST':
#         print('POST data:', request.POST)  # Debugging line
#         form = OffboardingForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             return redirect('home')  # Redirect after successful save
#         else:
#             print(form.errors)  # Debugging validation errors
#     else:
#         form = OffboardingForm()
#     employees = Employee.objects.all() 
#     offboarding = Offboarding.objects.all() 
#     context = {
#         'form': form,
#         'employees': employees,
#         'offboarding': offboarding,
#     }
#     return render(request, 'employee/offboarding2.html', context)



# Formset definition


# Formset definition

# Formset definition - Remove custom prefix, use Django default
# Create the formset
AssetHandoverFormSet = inlineformset_factory(
    Offboarding,
    AssetHandover,
    form=AssetHandoverForm,
    extra=0,
    can_delete=True
)

@login_required
@group_required("Admin", "HR")
def offboarding_list(request):
    """Main page - List, Create, Edit, View, Delete all in one"""
    if request.method == 'POST':
        off_id = request.POST.get("offboarding_id")

        if off_id:
            offboarding_instance = get_object_or_404(Offboarding, id=off_id)
            form = OffboardingForm(request.POST, request.FILES, instance=offboarding_instance)
            formset = AssetHandoverFormSet(request.POST, request.FILES, instance=offboarding_instance)
        else:
            offboarding_instance = None
            form = OffboardingForm(request.POST, request.FILES)
            formset = AssetHandoverFormSet(request.POST, request.FILES, instance=Offboarding())

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                offboarding = form.save()
                formset.instance = offboarding
                formset.save()

                # Auto-mark employee as Left when creating a new offboarding
                if not off_id:
                    employee = offboarding.employee
                    if employee.status != 'Left':
                        employee.status = 'Left'
                        employee.save(update_fields=['status'])

            if off_id:
                messages.success(request, "Offboarding updated successfully!")
            else:
                messages.success(request, f"Offboarding created. {offboarding.employee} has been marked as Left.")

            return redirect('offboarding-list')
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = OffboardingForm()
        formset = AssetHandoverFormSet(instance=Offboarding())

    offboardings = Offboarding.objects.all().select_related('employee')
    employees = Employee.objects.filter(status__in=['Active', 'Pending'])

    return render(request, 'employee/offboarding2.html', {
        'form': form,
        'formset': formset,
        'offboardings': offboardings,
        'employees': employees,
    })

@login_required
def offboarding_detail(request, off_id):
    """Return JSON data for view modal"""
    off = get_object_or_404(Offboarding, id=off_id)
    emp = off.employee
    assets = off.asset_handovers.all()

    asset_data = []
    for a in assets:
        asset_data.append({
            "asset_type": a.asset_type,
            "quantity": a.quantity,
            "condition": a.condition_on_return,
            "remarks": a.remarks or "-",
            "returned": "Yes" if a.returned else "No",
            "asset_photo": a.asset_photo.url if a.asset_photo else None,
            "receipt": a.receipt.url if a.receipt else None,
        })

    notice_days = (off.date_of_relieving - off.date_of_resignation).days if off.date_of_resignation and off.date_of_relieving else None

    data = {
        # Employee
        "employee_code": emp.employee_code or "-",
        "employee_name": f"{emp.first_name or ''} {emp.last_name or ''}".strip() or str(emp),
        "designation": emp.designation or "-",
        "department": emp.department or "-",
        "email": emp.personal_email or "-",
        "phone": emp.personal_mobile or "-",

        # Offboarding
        "resignation_date": off.date_of_resignation.strftime("%b %d, %Y") if off.date_of_resignation else "-",
        "relieving_date": off.date_of_relieving.strftime("%b %d, %Y") if off.date_of_relieving else "-",
        "notice_period_days": notice_days,

        # Documents
        "experience_certificate": off.experience_certificate.url if off.experience_certificate else None,
        "relieving_letter": off.relieving_letter.url if off.relieving_letter else None,
        "other_documents": off.other_documents.url if off.other_documents else None,
        "fnf_documents": off.fnf_documents.url if off.fnf_documents else None,

        # Assets
        "assets": asset_data
    }

    return JsonResponse(data)

@login_required
def offboarding_edit_data(request, id):
    """Return JSON data for edit modal"""
    off = get_object_or_404(Offboarding, id=id)
    assets = AssetHandover.objects.filter(offboarding=off)

    return JsonResponse({
        "offboarding": {
            "employee": off.employee_id,
            "date_of_resignation": off.date_of_resignation.strftime("%Y-%m-%d") if off.date_of_resignation else "",
            "date_of_relieving": off.date_of_relieving.strftime("%Y-%m-%d") if off.date_of_relieving else "",

            # File info (name + url)
            "experience_certificate": {
                "name": off.experience_certificate.name.split("/")[-1] if off.experience_certificate else "",
                "url": off.experience_certificate.url if off.experience_certificate else ""
            },
            "relieving_letter": {
                "name": off.relieving_letter.name.split("/")[-1] if off.relieving_letter else "",
                "url": off.relieving_letter.url if off.relieving_letter else ""
            },
            "other_documents": {
                "name": off.other_documents.name.split("/")[-1] if off.other_documents else "",
                "url": off.other_documents.url if off.other_documents else ""
            },
            "fnf_documents": {
                "name": off.fnf_documents.name.split("/")[-1] if off.fnf_documents else "",
                "url": off.fnf_documents.url if off.fnf_documents else ""
            },
        },

        "assets": [
            {
                "id": a.id,
                "asset_type": a.asset_type,
                "quantity": a.quantity,
                "condition": a.condition_on_return,
                "remarks": a.remarks or "",
                "returned": a.returned,

                # Asset files
                "asset_photo": {
                    "name": a.asset_photo.name.split("/")[-1] if a.asset_photo else "",
                    "url": a.asset_photo.url if a.asset_photo else ""
                },
                "receipt": {
                    "name": a.receipt.name.split("/")[-1] if a.receipt else "",
                    "url": a.receipt.url if a.receipt else ""
                },
            }
            for a in assets
        ]
    })

@login_required
def offboarding_delete(request, pk):
    """Delete offboarding via AJAX"""
    if request.method == 'POST':
        offboarding = get_object_or_404(Offboarding, pk=pk)
        offboarding.delete()
        return JsonResponse({'success': True, 'message': 'Offboarding deleted successfully!'})
    return JsonResponse({'success': False, 'message': 'Invalid request.'})

# def create_branch(request):
#     if request.method == 'POST':
#         form = BranchForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('admin-dashboard')  # Redirect after successful save
#         else:
#             print(form.errors)  # Debugging validation errors
#     else:
#         form = BranchForm()

#     context = {
#         'form': form,
#     }
#     return render(request, 'branch/create-branch.html', context)

@login_required
@group_required("Admin", "HR")
def create_branchs(request):
    if request.method == "POST":
        branch_name = request.POST.get("branch_name")
        branch_address = request.POST.get("branch_address")
       
        # Save company to database
        branch = Branch.objects.create(
            branch_name=branch_name,
            branch_address=branch_address,          
        )
        messages.success(request, "Branch added successfully!")
        return redirect("create-branch")  # Redirect to company list page
    branches = Branch.objects.all()
    return render(request, "branch/create-branch.html",{"branches": branches})


@login_required
def edit_branch(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == "POST":
        form = BranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch Updated successfully!")
            return redirect("create-branch")
    else:
        form = BranchForm(instance=branch)

    return render(request, "branch/_edit_branch_form.html", {"form": form, "branch": branch})


@login_required
def get_branch(request, branch_id):
    branch = Branch.objects.get(id=branch_id)
    return JsonResponse({
        "branch_name": branch.branch_name,
        "branch_address": branch.branch_address,
        # "tan_number": company.tan_number,
        # "pan_number": company.pan_number,
        # "employer_pf": company.employer_pf,
        # "ptrc_number": company.ptrc_number,
        # "ptec_number": company.ptec_number,
        # "esic_number": company.esic_number,
        # "status": company.status,
    })

@login_required
def delete_branch(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == "POST":
        branch.delete()
        messages.success(request, "Branch deleted successfully.")
        return redirect("create-branch")

    messages.error(request, "Invalid request.")
    return redirect("create-branch")

# def create_company(request):
#     if request.method == 'POST':
#         form = CompanyForm(request.POST)
#         if form.is_valid():
#             form.save()
#             return redirect('admin-dashboard')  # Redirect after successful save
#         else:
#             print(form.errors)  # Debugging validation errors
#     else:
#         form = CompanyForm()

#     context = {
#         'form': form,
#     }
#     return render(request, 'company/home2.html', context)
@login_required
@group_required("Admin")

def create_company(request):
    if request.method == "POST":
        short_name = request.POST.get("short_name")
        name = request.POST.get("name")
        phone = request.POST.get("phone")
        email = request.POST.get("email")
        address = request.POST.get("address")
        tan_number = request.POST.get("tan_number")
        pan_number = request.POST.get("pan_number")
        employer_pf = request.POST.get("employer_pf")
        ptrc_number = request.POST.get("ptrc_number")
        ptec_number = request.POST.get("ptec_number")
        esic_number = request.POST.get("esic_number")
        status = request.POST.get("status")

        if not short_name or not name or not address:
            messages.error(request, "Short Name, Company Name, and Address are required.")
            return redirect("create-company")  # Redirect back if validation fails

        # Save company to database
        company = Company.objects.create(
            short_name=short_name,
            name=name,
            phone=phone,
            email=email,
            address=address,
            tan_number=tan_number,
            pan_number=pan_number,
            employer_pf=employer_pf,
            ptrc_number=ptrc_number,
            ptec_number=ptec_number,
            esic_number=esic_number,
            status=status,
        )
        messages.success(request, "Company added successfully!")
        return redirect("create-company")  # Redirect to company list page
    companies = Company.objects.all()
    return render(request, "company/home2.html",{"companies": companies})




@login_required
@group_required("Admin")
def edit_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    if request.method == "POST":
        form = CompanyForm(request.POST, instance=company)   # ← IMPORTANT
        if form.is_valid():
            form.save()
            messages.success(request, "Company updated successfully.")
            return redirect("create-company")   # your listing page
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = CompanyForm(instance=company)   # ← IMPORTANT

    return render(request, "company/_edit_company_form.html", {"form": form, "company": company})


@login_required
@group_required("Admin", "HR")
def get_company(request, company_id):
    company = Company.objects.get(id=company_id)
    return JsonResponse({
        "short_name": company.short_name,
        "name": company.name,
        "address": company.address,
        "tan_number": company.tan_number,
        "pan_number": company.pan_number,
        "employer_pf": company.employer_pf,
        "ptrc_number": company.ptrc_number,
        "ptec_number": company.ptec_number,
        "esic_number": company.esic_number,
        "status": company.status,
    })



@login_required
@group_required("Admin")
def delete_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    if request.method == "POST":
        company.delete()
        messages.success(request, "Company deleted successfully.")
        return redirect("create-company")

    messages.error(request, "Invalid request.")
    return redirect("create-company")


# def create_salary(request):
#     form = SalaryMasterForm()
#     return render(request,'employee/create_employee.html',{'form':form})




# def leave_balance_view(request):
#     employees = LeaveBalance.objects.all()  # Get all leave balances

#     if request.method == "POST":
#         employee_id = request.POST.get("employee_id")
#         from_date = request.POST.get("from_date")
#         to_date = request.POST.get("to_date")

#         # Convert input dates to datetime objects
#         from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
#         to_date = datetime.strptime(to_date, "%Y-%m-%d").date()

#         # Get employee's leave balance record
#         leave_balance = get_object_or_404(LeaveBalance, employee_id=employee_id)

#         # Calculate leave details
#         leave_balance.calculate_leave_data(from_date, to_date)

#         return JsonResponse({
#             "status": "success",
#             "leave_balance": leave_balance.leave_balance,
#             "leave_taken": leave_balance.leave_taken,
#             "lwp": leave_balance.leave_without_pay,
#             "closing_balance": leave_balance.closing_balance
#         })

#     return render(request, "leave_balance/leave_balance.html", {"employees": employees})



# def leave_balance_view(request):
#     """
#     Display leave balance with month filter
#     Shows current month by default, with option to view history
#     """
#     # Get selected month from query params
#     month_filter = request.GET.get("month")
    
#     # Get all unique months from history for the dropdown
#     available_months = (
#         LeaveBalanceHistory.objects
#         .values_list("month", flat=True)
#         .distinct()
#         .order_by("-month")
#     )
    
#     # Filter by company if user has company association
#     company_filter = {}
#     if hasattr(request.user, 'employee_profile'):
#         company_filter['employee__company'] = request.user.employee_profile.company
    
#     if month_filter:
#         # Show historical data for selected month
#         try:
#             selected_month = date.fromisoformat(month_filter)
#             leave_balances = LeaveBalanceHistory.objects.filter(
#                 month=selected_month,
#                 **company_filter
#             ).select_related('employee', 'employee__company', 'employee__branch')
            
#             display_month = selected_month.strftime("%B %Y")
#             is_current = False
            
#         except (ValueError, TypeError):
#             # Invalid date format, fallback to current
#             leave_balances = LeaveBalance.objects.filter(
#                 **company_filter
#             ).select_related('employee', 'employee__company', 'employee__branch')
            
#             display_month = "Current Month"
#             is_current = True
#     else:
#         # Show current leave balances
#         leave_balances = LeaveBalance.objects.filter(
#             **company_filter
#         ).select_related('employee', 'employee__company', 'employee__branch')
        
#         display_month = "Current Month"
#         is_current = True
    
#     # Order by employee code
#     leave_balances = leave_balances.order_by('employee__employee_code')
    
#     context = {
#         "leave_balances": leave_balances,
#         "available_months": available_months,
#         "selected_month": month_filter,
#         "display_month": display_month,
#         "is_current": is_current,
#     }
    
#     return render(request, "leave_balance/leave_balance.html", context)




@login_required
def employee_leave_history(request, employee_id):
    """Return HTML for the employee’s month-by-month leave history."""
    history = LeaveBalanceHistory.objects.filter(employee_id=employee_id).order_by("-recorded_on")
    html = render_to_string("leave_balance/_leave_history_table.html", {"history": history})
    return JsonResponse({"html": html})

# from website.signals import update_leave_balance_on_attendance
from django.views.decorators.http import require_POST

# @require_POST
# def recalculate_leave_balances(request):
#     """
#     Manually triggers recalculation of all leave balances.
#     Uses the SAME attendance-based logic as the signal.
#     """
#     from website.models import LeaveBalance, Attendance

#     records = LeaveBalance.objects.select_related("employee")

#     for record in records:
#         employee = record.employee
#         last_att = Attendance.objects.filter(employee=employee).order_by("-date").first()

#         if last_att:
#             # call your existing signal method
#             update_leave_balance_on_attendance(Attendance, last_att, created=False)

#     return JsonResponse({"status": "ok", "message": "Leave balances recalculated successfully!"})


# def leave_credit_policy_view(request):
#     """Display the current company's leave credit policy."""
#     # Assuming HR is logged in and linked to a company
#     company = getattr(request.user, "company", None)
#     if not company:
#         return render(request, "leave_policy/leave_credit_policy.html", {"error": "No company found for user."})

#     policy, _ = LeaveCreditPolicy.objects.get_or_create(company=company)

#     return render(request, "leave_policy/leave_credit_policy.html", {
#         "policy": policy
#     })



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Min, Max
from datetime import date, timedelta
from decimal import Decimal
from website.models import LeaveBalance, Employee, PayrollSettings, Attendance, CompOffRequest, Company, MonthlyEarnedLeaves
from django.db import transaction


def get_user_company(user):
    """Get the company linked to this user via their employee profile.
    Returns None for global users (superuser/staff) without an employee profile
    — callers must handle None and resolve company via URL param or selection."""
    try:
        employee = Employee.objects.get(user=user)
        return employee.company
    except Employee.DoesNotExist:
        return None


def user_has_global_access(user):
    """
    Admin, HR, and Manager groups can view and manage data across ALL companies.
    Only Employee group is restricted to their own company.
    """
    if user.is_superuser or user.is_staff:
        return True
    return user.groups.filter(name__in=['Admin', 'HR', 'Manager']).exists()


def get_company_filter(user):
    """
    Returns the company to scope data queries to.
    Returns None for Admin/HR/Manager (global access — no company filter).
    Returns the user's own company for Employee group.
    """
    if user_has_global_access(user):
        return None
    return get_user_company(user)


def get_payroll_period_for_date(payroll_settings, target_date):
    """
    Determine which payroll period a date falls into.
    If from_date and to_date (day numbers) are explicitly configured, they always
    take priority over the is_auto flag.  is_auto is only used as a fallback when
    neither day number is set.
    """
    from_day = payroll_settings.from_date
    to_day = payroll_settings.to_date

    # Use custom period when day numbers are configured
    if from_day and to_day:
        if from_day <= to_day:
            # Same-month period (e.g. 1 → 31)
            if target_date.day >= from_day:
                from_d = date(target_date.year, target_date.month, from_day)
                to_d = date(target_date.year, target_date.month, to_day)
            else:
                prev_month = target_date.month - 1
                prev_year = target_date.year
                if prev_month == 0:
                    prev_month = 12
                    prev_year -= 1
                from_d = date(prev_year, prev_month, from_day)
                to_d = date(prev_year, prev_month, to_day)
        else:
            # Cross-month period (e.g. 27 → 26)
            if target_date.day >= from_day:
                from_d = date(target_date.year, target_date.month, from_day)
                next_month = target_date.month + 1
                next_year = target_date.year
                if next_month > 12:
                    next_month = 1
                    next_year += 1
                to_d = date(next_year, next_month, to_day)
            else:
                prev_month = target_date.month - 1
                prev_year = target_date.year
                if prev_month == 0:
                    prev_month = 12
                    prev_year -= 1
                from_d = date(prev_year, prev_month, from_day)
                to_d = date(target_date.year, target_date.month, to_day)
        return from_d, to_d

    # Fallback: calendar month (is_auto behaviour)
    first_day = target_date.replace(day=1)
    if target_date.month == 12:
        last_day = date(target_date.year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(target_date.year, target_date.month + 1, 1) - timedelta(days=1)
    return first_day, last_day

def get_all_payroll_periods_from_attendance(company, payroll_settings):
    """
    Return all payroll periods (derived from PayrollSettings.from_date / to_date)
    that contain at least one attendance record for the company.

    Steps period-by-period rather than day-by-day, so it is O(months) not O(days).
    """
    attendance_stats = Attendance.objects.filter(
        employee__company=company
    ).aggregate(min_date=Min('date'), max_date=Max('date'))

    min_date = attendance_stats.get('min_date')
    max_date = attendance_stats.get('max_date')

    if not min_date or not max_date:
        return []

    periods = []

    # Start at the payroll period that contains the earliest attendance date
    from_d, to_d = get_payroll_period_for_date(payroll_settings, min_date)

    while from_d <= max_date:
        has_attendance = Attendance.objects.filter(
            employee__company=company,
            date__gte=from_d,
            date__lte=to_d,
        ).exists()

        if has_attendance:
            label = f"{from_d.strftime('%d %b %Y')} - {to_d.strftime('%d %b %Y')}"
            periods.append({
                'label': label,
                'from_date': from_d,
                'to_date': to_d,
                'display_date': to_d,
            })

        # Advance one full period forward (day after current period ends)
        from_d, to_d = get_payroll_period_for_date(payroll_settings, to_d + timedelta(days=1))

    periods.sort(key=lambda x: x['to_date'], reverse=True)
    return periods


def get_monthly_earned_leaves(payroll_settings, month, year):
    """Get earned leaves for a specific month, falling back to earned_leaves_per_year / 12."""
    try:
        monthly_leave = MonthlyEarnedLeaves.objects.get(
            payroll_settings=payroll_settings,
            month=month,
            year=year
        )
        return Decimal(str(monthly_leave.earned_leaves))
    except MonthlyEarnedLeaves.DoesNotExist:
        return (Decimal(str(payroll_settings.earned_leaves_per_year)) / Decimal('12')).quantize(Decimal('0.01'))

def calculate_leave_balance_for_period(employee, payroll_settings, from_date, to_date):
    """
    ✅ UPDATED: Calculate leave balance for a specific payroll period
    Now reads monthly credit from MonthlyEarnedLeaves table
    """
    
    # ============================================
    # STEP 1: Opening Balance
    # Get PREVIOUS PERIOD's final balance (by period dates!)
    # ============================================
    prev_period_end = from_date - timedelta(days=1)
    prev_from, prev_to = get_payroll_period_for_date(payroll_settings, prev_period_end)
    
    previous_record = (
        LeaveBalance.objects
        .filter(
            employee=employee,
            period_from_date=prev_from,
            period_to_date=prev_to
        )
        .first()
    )
    
    opening_balance = (
        previous_record.final_leave_balance 
        if previous_record 
        else Decimal("0.00")
    )
    
    # Handle reset logic
    if not getattr(payroll_settings, 'carry_forward', True):
        reset_month = getattr(payroll_settings, 'reset_month', None)
        if reset_month and reset_month == to_date.month:
            opening_balance = Decimal("0.00")
    
    # ============================================
    # STEP 2: Attendance for THIS period
    # Exclude holidays (is_holiday=False)
    # ============================================
    # Determine weekend exclusion based on PayrollSettings (Django week_day: 1=Sunday, 7=Saturday)
    if getattr(payroll_settings, 'weekend_days', 'sat_sun') == 'sun':
        weekend_exclude = [1]       # Sunday only
    else:
        weekend_exclude = [1, 7]    # Saturday & Sunday

    attendance_records = Attendance.objects.filter(
        employee=employee,
        date__gte=from_date,
        date__lte=to_date,
        is_holiday=False,
    ).exclude(date__week_day__in=weekend_exclude)

    # Total days = actual calendar days in the payroll period (from_date to to_date inclusive)
    total_days = (to_date - from_date).days + 1

    # Working days (excluding weekends/holidays) used for leave taken calculation
    working_days = attendance_records.count()

    # ============================================
    # STEP 3: Paid Days
    # ============================================
    paid_days_sum = attendance_records.aggregate(
        total=Sum("count")
    )["total"]
    paid_days = paid_days_sum if paid_days_sum else Decimal("0.00")

    # Count weekend days in the period — always treated as present
    sunday_only = getattr(payroll_settings, 'weekend_days', 'sat_sun') == 'sun'
    weekend_day_count = 0
    d = from_date
    while d <= to_date:
        is_weekend = (d.weekday() == 6) if sunday_only else (d.weekday() >= 5)
        if is_weekend:
            weekend_day_count += 1
        d += timedelta(days=1)

    # Days present = actual paid working days + all weekends in period
    days_present = paid_days + Decimal(str(weekend_day_count))

    # ============================================
    # STEP 4: Leave Taken
    # Based on working days only — weekends are never counted as leave taken
    # ============================================
    leave_taken = Decimal(str(working_days)) - paid_days
    if leave_taken < 0:
        leave_taken = Decimal("0.00")
    
    # ============================================
    # STEP 5: Late
    # Count "Late Present" records directly rather than summing
    # Attendance.late (minutes short of the flexible 9-hour duty).
    # ============================================
    late_count = attendance_records.filter(status="Late Present").count()
    if getattr(payroll_settings, 'late_marks_affect_lwp', True):
        # Grace: first 5 late marks are free. Every 3 marks after that = 1 day deducted.
        late_days = Decimal(str((late_count - 5) // 3)) if late_count > 5 else Decimal("0")
    else:
        late_days = Decimal("0")

    # ============================================
    # STEP 6: Comp-Off
    # ============================================
    compoff_total = (
        CompOffRequest.objects
        .filter(
            employee=employee,
            status="Approved",
            from_date__gte=from_date,
            to_date__lte=to_date
        )
        .aggregate(total=Sum("count"))["total"]
        or Decimal("0.00")
    )

    # ============================================
    # STEP 7: LWP
    # Check if this record has a manual LWP override — preserve it if so.
    # ============================================
    existing_lb = LeaveBalance.objects.filter(
        employee=employee,
        period_from_date=from_date,
        period_to_date=to_date,
        lwp_overridden=True,
    ).first()

    if existing_lb:
        leave_without_pay = existing_lb.leave_without_pay
        leave_balance = max(Decimal("0.00"), opening_balance + compoff_total - leave_taken - late_days - leave_without_pay)
        if leave_balance < 0:
            leave_balance = Decimal("0.00")
    else:
        balance_before_credit = opening_balance + compoff_total - leave_taken - late_days
        if balance_before_credit < 0:
            leave_without_pay = abs(balance_before_credit)
            leave_balance = Decimal("0.00")
        else:
            leave_without_pay = Decimal("0.00")
            leave_balance = balance_before_credit
    
    # ============================================
    # STEP 8: Monthly Leave Credit via LeaveCreditPolicy
    # Credit depends on how many days the employee was present this period.
    # ============================================
    try:
        policy = employee.company.leave_credit_policy
        present_for_credit = int(days_present)
        if present_for_credit <= policy.credit_1_limit:
            monthly_credit = Decimal(str(policy.credit_low))
        elif present_for_credit <= policy.credit_2_limit:
            monthly_credit = Decimal(str(policy.credit_mid))
        else:
            monthly_credit = Decimal(str(policy.credit_high))
        monthly_cap = Decimal(str(payroll_settings.earned_leaves_per_year)) / Decimal('12')
        monthly_credit = min(monthly_credit, monthly_cap)
    except Exception:
        monthly_credit = Decimal("1.00")
    
    # ============================================
    # STEP 9: Closing Balance
    # ============================================
    closing_balance = leave_balance + monthly_credit
    
    final_leave_balance = min(
        closing_balance,
        Decimal(str(payroll_settings.max_leave_balance))
    )
    
    # ============================================
    # STEP 10: Save Record (with period dates)
    # Never overwrite lwp_overridden — it is managed only by override_lwp_view.
    # ============================================
    LeaveBalance.objects.update_or_create(
        employee=employee,
        period_from_date=from_date,
        period_to_date=to_date,
        defaults={
            'opening_balance': opening_balance,
            'leave_taken': leave_taken,
            'number_of_days_present': days_present,
            'total_number_of_days': total_days,
            'late': late_count,
            'compoff': compoff_total,
            'leave_without_pay': leave_without_pay,
            'leave_balance': leave_balance,
            'closing_balance': closing_balance,
            'final_leave_balance': final_leave_balance,
        }
    )

    return final_leave_balance



def generate_leave_balances_for_all_periods(company, payroll_settings):
    """Generate leave balance records for ALL periods"""

    periods = get_all_payroll_periods_from_attendance(company, payroll_settings)

    if not periods:
        return 0

    employees = Employee.objects.filter(company=company, status='Active')
    employee_ids = list(employees.values_list('id', flat=True))

    # Delete stale records whose period_to_date does NOT match any valid payroll
    # period — these are old calendar-month records (e.g. Apr 1-30) that were
    # created before the payroll cycle was configured, and would otherwise shadow
    # the correct payroll-cycle records in the report.
    valid_to_dates = [p['to_date'] for p in periods]
    if employee_ids:
        LeaveBalance.objects.filter(
            employee_id__in=employee_ids,
            period_to_date__isnull=False,
        ).exclude(
            period_to_date__in=valid_to_dates
        ).delete()

    total_calculated = 0

    for period_info in periods:
        from_date = period_info['from_date']
        to_date = period_info['to_date']

        for employee in employees:
            try:
                with transaction.atomic():
                    calculate_leave_balance_for_period(
                        employee,
                        payroll_settings,
                        from_date,
                        to_date
                    )
                    total_calculated += 1
            except Exception as e:
                print(f"Error: {employee.employee_code} {from_date}-{to_date}: {e}")
                continue

    return total_calculated


@login_required
def leave_balance_view(request):
    """
    Leave balance report.
    - Global users (Admin/HR/Manager): see ALL companies by default.
      ?company_id=X filters to a single company and enables period selection.
    - Employee group: sees only their own company with period selection.
    """
    from website.models import Company
    user = request.user
    user_own_company = get_user_company(user)
    is_global = user_has_global_access(user)
    all_companies = Company.objects.all().order_by('name') if is_global else None

    # Resolve which company we're scoping to (None = all)
    company_id_param = request.GET.get('company_id', '').strip()
    if is_global and company_id_param:
        try:
            scoped_company = Company.objects.get(id=company_id_param)
        except Company.DoesNotExist:
            scoped_company = None
    elif is_global:
        scoped_company = None          # all companies
    else:
        scoped_company = user_own_company

    # ── Base context ──────────────────────────────────────────────────────────
    context = {
        'is_global': is_global,
        'all_companies': all_companies,
        'selected_company_id': scoped_company.id if scoped_company else None,
        'user_company_id': scoped_company.id if scoped_company else (user_own_company.id if user_own_company else None),
        # Global users can always recalculate (all companies or one specific)
        'can_recalculate': is_global or bool(scoped_company),
        'show_company_col': is_global and scoped_company is None,
    }

    # ── Non-global with no company — error ────────────────────────────────────
    if not is_global and not scoped_company:
        context.update({'leave_balances': [], 'display_month': date.today().strftime("%B %Y"), 'available_periods': [], 'available_months': [], 'period_month_str': ''})
        return render(request, 'leave_balance/leave_balance_report.html', context)

    search_query = request.GET.get('q', '').strip()
    company_filter_q = request.GET.get('company_filter', '').strip()  # text filter for all-company view

    # ── ALL-COMPANIES MODE (global user, no specific company selected) ─────────
    if scoped_company is None:
        from django.db.models import OuterRef, Subquery, Max

        # Build available periods from actual attendance + payroll settings.
        # Only use the dominant explicit cycle (most companies sharing the same
        # from_day / to_day). Companies still on calendar-month defaults (no
        # explicit from_date / to_date) are excluded so they don't pollute the
        # dropdown with Apr 1-30 / Mar 1-31 entries.
        from collections import Counter as _Counter

        _ps_map = {}  # company_pk -> (company, payroll_settings)
        for _company in Company.objects.filter(status='active'):
            try:
                _ps = PayrollSettings.objects.get(company=_company)
                _ps_map[_company.pk] = (_company, _ps)
            except PayrollSettings.DoesNotExist:
                pass

        # Count explicitly configured cycles only
        _cycle_counts = _Counter()
        for _, (_c, _ps) in _ps_map.items():
            if _ps.from_date and _ps.to_date:
                _cycle_counts[(_ps.from_date, _ps.to_date)] += 1

        dominant_cycle = _cycle_counts.most_common(1)[0][0] if _cycle_counts else None

        period_map = {}  # str(to_date) -> dict
        for _company_pk, (_company, _ps) in _ps_map.items():
            ps_cycle = (_ps.from_date, _ps.to_date) if (_ps.from_date and _ps.to_date) else None
            # Skip companies not on the dominant cycle (includes calendar-month defaults)
            if dominant_cycle and ps_cycle != dominant_cycle:
                continue
            for _p in get_all_payroll_periods_from_attendance(_company, _ps):
                key = str(_p['to_date'])
                if key not in period_map:
                    period_map[key] = {
                        'value': key,
                        'label': f"{_p['from_date'].strftime('%d %b %Y')} - {_p['to_date'].strftime('%d %b %Y')}",
                        'to_date': _p['to_date'],
                    }

        available_months = sorted(period_map.values(), key=lambda x: x['to_date'], reverse=True)

        # Parse ?period_month=YYYY-MM-DD (exact to_date of the selected period)
        period_month_str = request.GET.get('period_month', '').strip()
        sel_to_date = None
        if period_month_str:
            try:
                sel_to_date = date.fromisoformat(period_month_str)
            except ValueError:
                period_month_str = ''

        # Subquery: get the relevant LB id per employee.
        # Restrict to records whose period_to_date is a valid payroll-cycle period
        # so that old calendar-month records (e.g. Apr 1-30) don't shadow newer
        # payroll-cycle records (e.g. Mar 27 - Apr 26) just because Apr 30 > Apr 26.
        valid_to_dates = [m['to_date'] for m in available_months]

        if sel_to_date:
            latest_lb_id = (
                LeaveBalance.objects
                .filter(
                    employee=OuterRef('pk'),
                    period_to_date=sel_to_date,
                )
                .order_by('-id')
                .values('id')[:1]
            )
        elif valid_to_dates:
            latest_lb_id = (
                LeaveBalance.objects
                .filter(
                    employee=OuterRef('pk'),
                    period_to_date__in=valid_to_dates,
                )
                .order_by('-period_to_date', '-id')
                .values('id')[:1]
            )
        else:
            latest_lb_id = (
                LeaveBalance.objects
                .filter(employee=OuterRef('pk'))
                .order_by('-period_to_date', '-id')
                .values('id')[:1]
            )

        employees_qs = (
            Employee.objects
            .filter(status='Active')
            .select_related('company')
            .annotate(latest_lb_id=Subquery(latest_lb_id))
            .order_by('company__name', 'first_name', 'last_name')
        )

        # Text search
        if search_query:
            sq = search_query.lower()
            employees_qs = employees_qs.filter(
                first_name__icontains=sq
            ) | employees_qs.filter(
                last_name__icontains=sq
            ) | employees_qs.filter(
                employee_code__icontains=sq
            )
            employees_qs = employees_qs.order_by('company__name', 'first_name', 'last_name')

        # Company text filter
        if company_filter_q:
            employees_qs = employees_qs.filter(company__name__icontains=company_filter_q)

        employees_list = list(employees_qs)

        # Build leave balance lookup by id
        lb_ids = [emp.latest_lb_id for emp in employees_list if emp.latest_lb_id]
        lb_map = {lb.id: lb for lb in LeaveBalance.objects.filter(id__in=lb_ids).select_related('employee')}

        rows = [{'employee': emp, 'lb': lb_map.get(emp.latest_lb_id)} for emp in employees_list]

        total_employees = len(rows)
        lb_with_data = [r['lb'] for r in rows if r['lb']]
        total_lwp = sum(lb.leave_without_pay for lb in lb_with_data) if lb_with_data else Decimal("0.00")
        total_leaves_taken = sum(lb.leave_taken for lb in lb_with_data) if lb_with_data else Decimal("0.00")
        avg_balance = (sum(lb.final_leave_balance for lb in lb_with_data) / len(lb_with_data)) if lb_with_data else Decimal("0.00")

        if sel_to_date and period_month_str in period_map:
            display_month = period_map[period_month_str]['label'] + ' (All Companies)'
        else:
            display_month = 'Latest Period (All Companies)'

        paginator = Paginator(rows, 25)
        page_obj = paginator.get_page(request.GET.get('page', 1))

        context.update({
            'leave_balances': page_obj,
            'page_obj': page_obj,
            'paginator': paginator,
            'search_query': search_query,
            'company_filter_q': company_filter_q,
            'display_month': display_month,
            'available_periods': [],
            'available_months': available_months,
            'selected_period': None,
            'period_month_str': period_month_str,
            'total_employees': total_employees,
            'total_lwp': float(total_lwp),
            'total_leaves_taken': float(total_leaves_taken),
            'avg_balance': float(avg_balance),
        })
        return render(request, 'leave_balance/leave_balance_report.html', context)

    # ── SINGLE-COMPANY MODE ───────────────────────────────────────────────────
    user_company = scoped_company
    context['user_company_id'] = user_company.id
    context['user_company_name'] = getattr(user_company, 'name', 'Your Company')

    try:
        payroll_settings = PayrollSettings.objects.get(company=user_company)
    except PayrollSettings.DoesNotExist:
        context.update({'leave_balances': [], 'error': 'Payroll settings not configured.', 'available_periods': [], 'available_months': [], 'period_month_str': ''})
        return render(request, 'leave_balance/leave_balance_report.html', context)

    available_periods = get_all_payroll_periods_from_attendance(user_company, payroll_settings)

    selected_period_str = request.GET.get('period')
    selected_period = None
    if selected_period_str and available_periods:
        for period in available_periods:
            if str(period['to_date']) == selected_period_str:
                selected_period = period
                break
    if not selected_period and available_periods:
        selected_period = available_periods[0]

    all_employees = list(
        Employee.objects.filter(company=user_company, status='Active')
        .order_by('first_name', 'last_name')
    )

    lb_lookup = {}
    if selected_period:
        for lb in LeaveBalance.objects.filter(
            employee__company=user_company,
            period_from_date=selected_period['from_date'],
            period_to_date=selected_period['to_date'],
        ).select_related('employee'):
            lb_lookup[lb.employee_id] = lb

    if search_query:
        sq = search_query.lower()
        all_employees = [
            emp for emp in all_employees
            if sq in (emp.first_name or '').lower()
            or sq in (emp.last_name or '').lower()
            or sq in (emp.employee_code or '').lower()
        ]

    rows = [{'employee': emp, 'lb': lb_lookup.get(emp.id)} for emp in all_employees]

    display_month = selected_period['label'] if selected_period else date.today().strftime("%b %d - %b %d, %Y")

    total_employees = len(rows)
    lb_with_data = [r['lb'] for r in rows if r['lb']]
    total_lwp = sum(lb.leave_without_pay for lb in lb_with_data) if lb_with_data else Decimal("0.00")
    total_leaves_taken = sum(lb.leave_taken for lb in lb_with_data) if lb_with_data else Decimal("0.00")
    avg_balance = (sum(lb.final_leave_balance for lb in lb_with_data) / len(lb_with_data)) if lb_with_data else Decimal("0.00")

    paginator = Paginator(rows, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context.update({
        'leave_balances': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'search_query': search_query,
        'display_month': display_month,
        'available_periods': available_periods,
        'available_months': [],
        'period_month_str': '',
        'selected_period': str(selected_period['to_date']) if selected_period else None,
        'can_recalculate': True,
        'total_employees': total_employees,
        'total_lwp': float(total_lwp),
        'total_leaves_taken': float(total_leaves_taken),
        'avg_balance': float(avg_balance),
    })
    return render(request, 'leave_balance/leave_balance_report.html', context)

@login_required
def recalculate_leave_balances_view(request):
    """Recalculate leave balances for all periods and employees"""

    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method.'
        }, status=405)

    from website.models import Company
    company_id_param = request.POST.get('company_id') or request.GET.get('company_id')

    # Global user with no company selected → recalculate ALL companies
    if user_has_global_access(request.user) and not company_id_param:
        companies = Company.objects.all()
        total_count = 0
        errors = []
        for company in companies:
            try:
                ps = PayrollSettings.objects.get(company=company)
                total_count += generate_leave_balances_for_all_periods(company, ps)
            except PayrollSettings.DoesNotExist:
                errors.append(f"{company.name}: no payroll settings")
            except Exception as e:
                errors.append(f"{company.name}: {e}")

        msg = f'✓ Calculated {total_count} leave balance(s) across {companies.count()} company(ies).'
        if errors:
            msg += f' Skipped ({len(errors)} company(ies) with no settings — use Settings Hub to broadcast): {"; ".join(errors)}'
        return JsonResponse({'success': True, 'message': msg, 'recalculated_count': total_count, 'skipped': errors})

    # Specific company
    if company_id_param and user_has_global_access(request.user):
        try:
            target_company = Company.objects.get(id=company_id_param)
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Company not found.'}, status=400)
    else:
        target_company = get_user_company(request.user)

    if not target_company:
        return JsonResponse({'success': False, 'message': 'No company associated.'}, status=400)

    try:
        payroll_settings = PayrollSettings.objects.get(company=target_company)
        calculated_count = generate_leave_balances_for_all_periods(target_company, payroll_settings)
        return JsonResponse({
            'success': True,
            'message': f'✓ Successfully calculated {calculated_count} leave balance(s) for {target_company.name}.',
            'recalculated_count': calculated_count
        })
    except PayrollSettings.DoesNotExist:
        return JsonResponse({'success': False, 'message': f'Payroll settings not configured for {target_company.name}.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


@login_required
def recalc_employee_leave_balance(request, employee_id):
    """Recalculate for single employee"""
    if not (request.user.is_staff or user_has_global_access(request.user)):
        messages.error(request, "Permission denied.")
        return redirect("leave-balance")

    try:
        employee = Employee.objects.select_related('company').get(id=employee_id)

        company_filter = get_company_filter(request.user)
        if company_filter and company_filter.id != employee.company_id:
            messages.error(request, "Company mismatch.")
            return redirect("leave-balance")
        
        payroll_settings = PayrollSettings.objects.get(company=employee.company)
        
        periods = get_all_payroll_periods_from_attendance(employee.company, payroll_settings)
        count = 0
        
        for period in periods:
            calculate_leave_balance_for_period(
                employee,
                payroll_settings,
                period['from_date'],
                period['to_date']
            )
            count += 1
        
        messages.success(request, f"✓ Recalculated {count} periods for {employee.first_name}")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect("leave-balance")



@login_required
@require_http_methods(["POST"])
def override_lwp_view(request):
    """Manually override the LWP value for a single LeaveBalance record."""
    lb_id = request.POST.get('lb_id')
    lwp_raw = request.POST.get('lwp_value', '0')

    try:
        lb = LeaveBalance.objects.select_related('employee__company').get(id=lb_id)

        company_filter = get_company_filter(request.user)
        if company_filter and lb.employee.company_id != company_filter.id:
            return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

        # Block override if the leave balance period overlaps a finalized payroll run
        locking_run = get_locking_run_for_period(
            lb.employee.company, lb.period_from_date, lb.period_to_date
        )
        if locking_run:
            return lock_response(locking_run, action="override LWP")

        new_lwp = Decimal(str(lwp_raw))
        if new_lwp < 0:
            return JsonResponse({'success': False, 'message': 'LWP cannot be negative.'}, status=400)

        # Derive the monthly credit that was applied (credit = closing - leave_balance)
        monthly_credit = lb.closing_balance - lb.leave_balance

        try:
            payroll = PayrollSettings.objects.get(company=lb.employee.company)
            max_cap = Decimal(str(payroll.max_leave_balance))
            affect_lwp = getattr(payroll, 'late_marks_affect_lwp', True)
        except PayrollSettings.DoesNotExist:
            max_cap = Decimal('30')
            affect_lwp = True

        # Recalculate leave_balance under the new LWP
        if affect_lwp:
            late_days = Decimal(str((lb.late - 5) // 3)) if lb.late > 5 else Decimal('0')
        else:
            late_days = Decimal('0')

        balance_before_credit = lb.opening_balance + lb.compoff - lb.leave_taken - late_days - new_lwp
        leave_balance = max(Decimal('0.00'), balance_before_credit)

        closing_balance = leave_balance + monthly_credit
        final_leave_balance = min(closing_balance, max_cap)

        lb.leave_without_pay = new_lwp
        lb.leave_balance = leave_balance
        lb.closing_balance = closing_balance
        lb.final_leave_balance = final_leave_balance
        lb.lwp_overridden = True
        lb.save(update_fields=['leave_without_pay', 'leave_balance', 'closing_balance', 'final_leave_balance', 'lwp_overridden'])

        return JsonResponse({
            'success': True,
            'lwp': f'{new_lwp:.2f}',
            'leave_balance': f'{leave_balance:.2f}',
            'closing_balance': f'{closing_balance:.2f}',
            'final_leave_balance': f'{final_leave_balance:.2f}',
        })

    except LeaveBalance.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Record not found.'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
def recalc_all_employees(request):
    """Recalculate for all employees"""
    if not (request.user.is_staff or user_has_global_access(request.user)):
        messages.error(request, "Permission denied.")
        return redirect("leave-balance")

    user_company = get_user_company(request.user)

    if not user_company:
        messages.error(request, "No company.")
        return redirect("leave-balance")
    
    try:
        payroll_settings = PayrollSettings.objects.get(company=user_company)
        count = generate_leave_balances_for_all_periods(user_company, payroll_settings)
        messages.success(request, f"✓ Generated leave balances for {count} combinations.")
    except PayrollSettings.DoesNotExist:
        messages.error(request, "Payroll settings not found.")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect("leave-balance")


@login_required
def employee_leave_detail(request, employee_id):
    """Show detailed leave history for specific employee"""
    try:
        employee = Employee.objects.select_related('company').get(id=employee_id)
        
        if not request.user.is_staff:
            user_employee = Employee.objects.filter(user=request.user).first()
            if not user_employee or user_employee.id != employee_id:
                messages.error(request, "Permission denied.")
                return redirect("leave-balance")
        
        history = LeaveBalance.objects.filter(
            employee=employee
        ).order_by('-period_to_date')
        
        current_balance = history.first()
        
        context = {
            'employee': employee,
            'history': history,
            'current_balance': current_balance,
        }
        
        return render(request, 'leave_balance/employee_detail.html', context)
    
    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect("leave-balance")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        return redirect("leave-balance")

        

@login_required
@group_required("Admin", "HR")
def update_leave_credit_policy(request):
    """Handle policy updates from the HR UI (AJAX or form submit)."""
    if request.method == "POST":
        company = getattr(request.user, "company", None)
        if not company:
            return JsonResponse({"status": "error", "message": "No company found for user."})

        policy, _ = LeaveCreditPolicy.objects.get_or_create(company=company)

        policy.credit_1_limit = int(request.POST.get("credit_1_limit", 15))
        policy.credit_2_limit = int(request.POST.get("credit_2_limit", 25))
        policy.credit_low = Decimal(request.POST.get("credit_low", 0))
        policy.credit_mid = Decimal(request.POST.get("credit_mid", 1))
        policy.credit_high = Decimal(request.POST.get("credit_high", 2))
        policy.save()

        return JsonResponse({
            "status": "success",
            "message": "Leave Credit Policy updated successfully!"
        })

    return JsonResponse({"status": "error", "message": "Invalid request."})


# def recalc_leave_balances_view(request):
#     recalculate_all_leave_balances()
#     messages.success(request, "Leave balances recalculated successfully!")
#     return redirect("leave_balance")






@login_required
def leave_apply_view(request):
    leaves = LeaveApplication.objects.select_related("employee").order_by("-id")

    # ================== Filters ==================
    status = request.GET.get("status")
    search = request.GET.get("search")

    if status and status != "All":
        leaves = leaves.filter(status=status)

    if search:
        leaves = leaves.filter(
            Q(employee__first_name__icontains=search) | 
            Q(employee__last_name__icontains=search) |
            Q(leave_type__icontains=search)
        )

    # Pagination
    paginator = Paginator(leaves, 8)  # 8 rows per page
    page = request.GET.get("page")
    leaves = paginator.get_page(page)

    # ================== POST Submit ==================
    if request.method == "POST":
        form = LeaveApplicationForm(request.POST)
        if form.is_valid():
            leave_obj = form.save(commit=False)
            # Block applying for leave that overlaps a finalized payroll period
            locking_run = get_locking_run_for_period(
                leave_obj.employee.company, leave_obj.start_date, leave_obj.end_date
            )
            if locking_run:
                messages.error(
                    request,
                    f"Cannot apply for leave: {leave_obj.start_date:%d %b %Y}–{leave_obj.end_date:%d %b %Y} "
                    f"overlaps the finalized payroll run "
                    f"({locking_run.start_date:%d %b %Y}–{locking_run.end_date:%d %b %Y}).",
                )
                return redirect("leave_apply")
            leave_obj.save()
            messages.success(request,"Leave Applied Successfully!")
            return redirect("leave_apply")
        messages.error(request,"Please fix the errors.")
    else:
        form = LeaveApplicationForm()

    return render(request,"leave_balance/leave_apply.html",{
        "form":form,
        "leaves":leaves,
        "status":status,
        "search":search,
    })


def _approve_leave_item(leave):
    """Approve a single LeaveApplication. Returns an error message string on
    failure (e.g. payroll lock), or None on success."""
    locking_run = get_locking_run_for_period(
        leave.employee.company, leave.start_date, leave.end_date
    )
    if locking_run:
        return _lock_message(locking_run, action="approve this leave")

    leave.status = "Approved"
    leave.save()
    return None


@login_required
@group_required("Admin", "HR", "Manager")
@require_POST
def approve_leave(request, leave_id):
    try:
        leave = LeaveApplication.objects.get(id=leave_id)
    except LeaveApplication.DoesNotExist:
        return JsonResponse({"message": "Leave not found"}, status=404)

    error = _approve_leave_item(leave)
    if error:
        return JsonResponse({"success": False, "error": error}, status=400)
    return JsonResponse({"message": "Leave Approved Successfully!"})


@login_required
@group_required("Admin", "HR", "Manager")
@require_POST
def bulk_approve_leave(request):
    ids = _extract_bulk_ids(request)
    approved, failed = [], []
    for leave_id in ids:
        try:
            leave = LeaveApplication.objects.get(id=leave_id)
        except LeaveApplication.DoesNotExist:
            failed.append({"id": leave_id, "error": "Leave not found."})
            continue
        error = _approve_leave_item(leave)
        if error:
            failed.append({"id": leave_id, "error": error})
        else:
            approved.append(leave_id)
    return JsonResponse({"success": True, "approved": approved, "failed": failed})


@login_required
@group_required("Admin", "HR", "Manager")
@require_POST
def reject_leave(request, leave_id):
    try:
        data = json.loads(request.body)
        reason = data.get("reason")

        if not reason:
            return JsonResponse({"message": "Reason required"}, status=400)

        leave = LeaveApplication.objects.get(id=leave_id)
        leave.status = "Rejected"
        leave.save()

        return JsonResponse({"message": "Leave Rejected Successfully!"})
    except:
        return JsonResponse({"message": "Error while rejecting"}, status=404)


@login_required
@group_required("Admin", "HR")
def leave_credit_policy_view(request):
    if user_has_global_access(request.user):
        return redirect('company-settings-hub')

    company = get_user_company(request.user)
    if not company:
        messages.error(request, "No company linked to your account.")
        return redirect('admin-dashboard')

    policy, created = LeaveCreditPolicy.objects.get_or_create(company=company)

    if request.method == "POST":
        form = LeaveCreditPolicyForm(request.POST, instance=policy)
        if form.is_valid():
            form.save()
            messages.success(request, "Leave Credit Policy updated successfully!")
            return redirect("admin-dashboard")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LeaveCreditPolicyForm(instance=policy)

    return render(request, "leave_policy/leave_credit_policy.html", {
        "form": form,
        "company": company
    })


# def leave_balance_view(request):
#     month_filter = request.GET.get("month")

#     if month_filter:
#         # If a month is selected, load history for that month
#         leave_balances = LeaveBalanceHistory.objects.select_related("employee").filter(month=month_filter)
#     else:
#         # Default: show current month live balances
#         leave_balances = LeaveBalance.objects.select_related("employee").all()

#     # List of all months available for dropdown
#     months = LeaveBalanceHistory.objects.values_list("month", flat=True).distinct()

#     return render(request, "leave_balance/leave_balance2.html", {
#         "leave_balances": leave_balances,
#         "months": months,
#         "selected_month": month_filter or "Current Month"
#     })



@login_required
def employee_compoff_details(request, employee_id):
    """Return comp-off details for a given employee (for modal)."""
    today = date.today()
    employee_obj = get_object_or_404(Employee, pk=employee_id)
    payroll_settings = PayrollSettings.objects.filter(company=employee_obj.company).first()
    if payroll_settings:
        from_date, to_date = payroll_settings.get_payroll_period()
    else:
        from_date = today.replace(day=1)
        to_date = (today.replace(day=1) + relativedelta(months=1)) - timedelta(days=1)

    compoffs = CompOffRequest.objects.filter(
        employee_id=employee_id,
        status="Approved",
        from_date__gte=from_date,
        to_date__lte=to_date
    )

    total_days = 0 # Initialize total counter

    # ✅ Compute total days in Python so template stays simple
    for c in compoffs:
        if c.from_date and c.to_date:
            # Calculate days for this individual request
            c.days = (c.to_date - c.from_date).days + 1
            # Add to the running total
            total_days += c.days 
        else:
            c.days = 0

    html = render_to_string(
        "leave_balance/_compoff_modal_table.html", 
        {
            "compoffs": compoffs, 
            "total_compoff_days": total_days # Pass the total
        }
    )
    # The JsonResponse is usually for the modal itself, not the main page number.
    # It seems your issue is with the number on the *main* page.
    # If the main page number needs to be updated after this call, you'll need to modify the JS.
    return JsonResponse({"html": html, "total_days": total_days})


# @csrf_exempt
# def submit_comp_off_request(request):
#     if request.method == "POST":
#         data = json.loads(request.body)
#         employee_id = data.get("employee_id")
#         from_date = data.get("from_date")
#         to_date = data.get("to_date")
#         reason = data.get("reason")

#         employee = Employee.objects.get(id=employee_id)
#         comp_off = CompOff.objects.create(
#             employee=employee, 
#             from_date=from_date, 
#             to_date=to_date, 
#             reason=reason
#         )
#         return JsonResponse({"message": "Comp-Off request submitted successfully!"})
#     return JsonResponse({"error": "Invalid request"}, status=400)




from datetime import datetime



@login_required
def submit_comp_off_request(request):
    if request.method == "POST":
        try:
            employee_id = request.POST.get("employee_id")
            from_date = request.POST.get("from_date")
            to_date = request.POST.get("to_date")
            reason = request.POST.get("reason")

            if not all([employee_id, from_date, to_date, reason]):
                return JsonResponse({"message": "Missing required fields"}, status=400)

            # Convert to date objects
            from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
            to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()

            # ✅ Calculate count
            count = (to_date_obj - from_date_obj).days + 1

            employee = Employee.objects.get(id=employee_id)

            # Block raising a comp-off whose date range falls in a finalized payroll period
            locking_run = get_locking_run_for_period(employee.company, from_date_obj, to_date_obj)
            if locking_run:
                return lock_response(locking_run, action="raise a comp-off request")

            comp_off = CompOffRequest.objects.create(
                employee=employee,
                from_date=from_date_obj,
                to_date=to_date_obj,
                reason=reason,
                count=count
            )

            return JsonResponse({"message": "Comp-Off request submitted successfully!"})

        except Exception as e:
            return JsonResponse({"message": f"Error: {str(e)}"}, status=400)
    return JsonResponse({"message": "Invalid request"}, status=400)






# def create_salary(request):
#     """A view to create a new SalaryMaster record with calculations in the view."""
#     if request.method == 'POST':
#         # 1. Parse form data (from request.POST)
#         employee_id = request.POST.get('employee')
#         pf_deducted = request.POST.get('pf_deducted') == 'on'  # checkbox
#         gratuity_applicable = request.POST.get('gratuity_applicable') == 'on'
#         esic_applicable = request.POST.get('esic_applicable') == 'on'

#         # Convert numeric inputs to Decimal safely
#         gross_ctc_pm = Decimal(request.POST.get('gross_ctc_pm') or '0')
#         basic_pm = Decimal(request.POST.get('basic_pm') or '0')
#         hra_pm = Decimal(request.POST.get('hra_pm') or '0')
#         guaranteed_cash_pm = Decimal(request.POST.get('guaranteed_cash_pm') or '0')

#         # 2. Perform calculations
#         # -- PF (Employer Contribution) --
#         # If PF is deducted: minimum of 1800 or 12% of basic
#         if pf_deducted:
#             pf_er_cont_pm = min(Decimal('1800'), basic_pm * Decimal('0.12'))
#         else:
#             pf_er_cont_pm = Decimal('0')

#         # -- Gratuity --
#         # 4.81% of Basic if applicable
#         gratuity_pm = Decimal('0')
#         if gratuity_applicable:
#             gratuity_pm = basic_pm * Decimal('0.0481')

#         # -- ESIC (Employer) --
#         # If guaranteed_cash <= 21000, 3.75% of guaranteed_cash
#         esic_er_cont_pm = Decimal('0')
#         if esic_applicable and guaranteed_cash_pm <= Decimal('21000'):
#             esic_er_cont_pm = guaranteed_cash_pm * Decimal('0.0375')

#         # -- ESIC (Employee) --
#         # If guaranteed_cash <= 21000, 0.25% (example from your sheet)
#         esic_ee_cont_pm = Decimal('0')
#         if esic_applicable and guaranteed_cash_pm <= Decimal('21000'):
#             esic_ee_cont_pm = guaranteed_cash_pm * Decimal('0.0025')

#         # -- Profession Tax (example logic) --
#         profession_tax_pm = Decimal('0')
#         if guaranteed_cash_pm > Decimal('10000'):
#             profession_tax_pm = Decimal('200')  # example threshold

#         # -- Net Salary PM --
#         # net_salary_pm = guaranteed_cash_pm - (PF(EE) + ESIC(EE) + profession tax)
#         # Suppose PF(EE) = same as PF(ER) or 12%? You can define your logic:
#         pf_ee_cont_pm = min(Decimal('1800'), basic_pm * Decimal('0.12')) if pf_deducted else Decimal('0')
#         net_salary_pm = guaranteed_cash_pm - (pf_ee_cont_pm + esic_ee_cont_pm + profession_tax_pm)

#         # 3. Create the SalaryMaster object
#         salary = SalaryMaster(
#             employee_id=employee_id,
#             pf_deducted=pf_deducted,
#             gratuity_applicable=gratuity_applicable,
#             esic_applicable=esic_applicable,

#             gross_ctc_pm=gross_ctc_pm,
#             gross_ctc_pa=gross_ctc_pm * Decimal('12'),
#             basic_pm=basic_pm,
#             basic_pa=basic_pm * Decimal('12'),
#             hra_pm=hra_pm,
#             hra_pa=hra_pm * Decimal('12'),
#             guaranteed_cash_pm=guaranteed_cash_pm,
#             guaranteed_cash_pa=guaranteed_cash_pm * Decimal('12'),

#             pf_er_cont_pm=pf_er_cont_pm,
#             pf_er_cont_pa=pf_er_cont_pm * Decimal('12'),
#             esic_er_cont_pm=esic_er_cont_pm,
#             esic_er_cont_pa=esic_er_cont_pm * Decimal('12'),
#             pf_ee_cont_pm=pf_ee_cont_pm,
#             pf_ee_cont_pa=pf_ee_cont_pm * Decimal('12'),
#             esic_ee_cont_pm=esic_ee_cont_pm,
#             esic_ee_cont_pa=esic_ee_cont_pm * Decimal('12'),
#             profession_tax_pm=profession_tax_pm,
#             profession_tax_pa=profession_tax_pm * Decimal('12'),

#             net_salary_pm=net_salary_pm,
#             net_salary_pa=net_salary_pm * Decimal('12'),
#         )
#         salary.save()

#         messages.success(request, f"Salary created for employee {salary.employee}!")
#         # return redirect('salary_list')  # or wherever you want

#     # If GET request, render a form
#     employees = Employee.objects.all()
#     context = {
#         'employees': employees
#     }
#     return render(request, 'salary/create_salary.html', context)

from django.conf import settings
from decimal import Decimal

# def create_salary(request):
#     if request.method == "POST":
#         employee_id = request.POST.get('employee')
#         employee = Employee.objects.get(pk=employee_id) if employee_id else None

#         # ✅ These come as 'yes'/'no' strings
#         pf_deducted = request.POST.get('pf_deducted', '').lower() == 'yes'
#         gratuity_applicable = request.POST.get('gratuity_applicable', '').lower() == 'yes'
#         esic_applicable = request.POST.get('esic_applicable', '').lower() == 'yes'

#         def get_decimal(field_name):
#             try:
#                 return Decimal(request.POST.get(field_name, '0') or '0')
#             except:
#                 return Decimal('0')

#         # ✅ Match the actual POST keys
#         gross_ctc_pm = get_decimal('gross_ctc_pm')
#         basic_pm = get_decimal('basic_pm')
#         hra_pm = get_decimal('hra_pm')
#         stat_bonus_pm = get_decimal('stat_bonus_pm')
#         allowance1_pm = get_decimal('allowance1_pm')
#         allowance2_pm = get_decimal('allowance2_pm')
#         special_allowance_pm = get_decimal('sp_allowance_pm')
#         guaranteed_cash_pm = get_decimal('guaranteed_cash_pm')
#         prof_tax_pm = get_decimal('profession_tax_pm')

#         pf_er_cont_pm = get_decimal('pf_er_cont_pm')
#         pf_ee_cont_pm = get_decimal('pf_ee_cont_pm')
#         esic_er_cont_pm = get_decimal('esic_er_cont_pm')
#         esic_ee_cont_pm = get_decimal('esic_ee_cont_pm')
#         gratuity_pm = get_decimal('gratuity_pm')
#         net_salary_pm = get_decimal('net_salary_pm')
#         cost_to_company_pm = get_decimal('ctc_pm')

#         instance = SalaryMaster(
#             employee=employee,
#             pf_deducted=pf_deducted,
#             gratuity_applicable=gratuity_applicable,
#             esic_applicable=esic_applicable,
#             gross_ctc_pm=gross_ctc_pm,
#             gross_ctc_pa=gross_ctc_pm * 12,
#             basic_pm=basic_pm,
#             basic_pa=basic_pm * 12,
#             hra_pm=hra_pm,
#             hra_pa=hra_pm * 12,
#             stat_bonus_pm=stat_bonus_pm,
#             stat_bonus_pa=stat_bonus_pm * 12,
#             sp_allowance_pm=special_allowance_pm,
#             sp_allowance_pa=special_allowance_pm * 12,
#             allowance1_pm=allowance1_pm,
#             allowance1_pa=allowance1_pm * 12,
#             allowance2_pm=allowance2_pm,
#             allowance2_pa=allowance2_pm * 12,
#             guaranteed_cash_pm=guaranteed_cash_pm,
#             guaranteed_cash_pa=guaranteed_cash_pm * 12,
#             gratuity_pm=gratuity_pm,
#             gratuity_pa=gratuity_pm * 12,
#             ctc_pm=cost_to_company_pm,
#             ctc_pa=cost_to_company_pm * 12,
#             pf_er_cont_pm=pf_er_cont_pm,
#             pf_er_cont_pa=pf_er_cont_pm * 12,
#             pf_ee_cont_pm=pf_ee_cont_pm,
#             pf_ee_cont_pa=pf_ee_cont_pm * 12,
#             esic_er_cont_pm=esic_er_cont_pm,
#             esic_er_cont_pa=esic_er_cont_pm * 12,
#             esic_ee_cont_pm=esic_ee_cont_pm,
#             esic_ee_cont_pa=esic_ee_cont_pm * 12,
#             profession_tax_pm=prof_tax_pm,
#             profession_tax_pa=prof_tax_pm * 12,
#             net_salary_pm=net_salary_pm,
#             net_salary_pa=net_salary_pm * 12,
#         )
#         instance.save()

#         return redirect('create-salary')

#     employees = Employee.objects.all()
#     salary = SalaryMaster.objects.all()
#     return render(request, 'salary/create_salary4.html', {'employees': employees, 'salary': salary})

def extract_decimal(request, key):
    """Safe decimal extraction helper."""
    try:
        return Decimal(request.POST.get(key, "0") or "0")
    except:
        return Decimal("0")


@login_required
@group_required("Admin", "HR")
def create_salary(request):
    if request.method == "POST":
        salary_id = request.POST.get("salary_id")
        employee_id = request.POST.get("employee")

        employee = Employee.objects.get(pk=employee_id)

        pf_deducted = request.POST.get("pf_deducted", "").lower() == "yes"
        gratuity_applicable = request.POST.get("gratuity_applicable", "").lower() == "yes"
        esic_applicable = request.POST.get("esic_applicable", "").lower() == "yes"

        data = {
            "gross_ctc_pm": extract_decimal(request, "gross_ctc_pm"),
            "gross_ctc_pa": extract_decimal(request, "gross_ctc_pa"),
            "basic_pm": extract_decimal(request, "basic_pm"),
            "basic_pa": extract_decimal(request, "basic_pa"),
            "hra_pm": extract_decimal(request, "hra_pm"),
            "hra_pa": extract_decimal(request, "hra_pa"),
            "stat_bonus_pm": extract_decimal(request, "stat_bonus_pm"),
            "stat_bonus_pa": extract_decimal(request, "stat_bonus_pa"),
            "allowance1_pm": extract_decimal(request, "allowance1_pm"),
            "allowance1_pa": extract_decimal(request, "allowance1_pa"),
            "allowance2_pm": extract_decimal(request, "allowance2_pm"),
            "allowance2_pa": extract_decimal(request, "allowance2_pa"),
            "sp_allowance_pm": extract_decimal(request, "sp_allowance_pm"),
            "sp_allowance_pa": extract_decimal(request, "sp_allowance_pa"),
            "guaranteed_cash_pm": extract_decimal(request, "guaranteed_cash_pm"),
            "guaranteed_cash_pa": extract_decimal(request, "guaranteed_cash_pa"),
            "profession_tax_pm": extract_decimal(request, "profession_tax_pm"),
            "profession_tax_pa": extract_decimal(request, "profession_tax_pa"),
            "pf_er_cont_pm": extract_decimal(request, "pf_er_cont_pm"),
            "pf_er_cont_pa": extract_decimal(request, "pf_er_cont_pa"),
            "pf_ee_cont_pm": extract_decimal(request, "pf_ee_cont_pm"),
            "pf_ee_cont_pa": extract_decimal(request, "pf_ee_cont_pa"),
            "esic_er_cont_pm": extract_decimal(request, "esic_er_cont_pm"),
            "esic_er_cont_pa": extract_decimal(request, "esic_er_cont_pa"),
            "esic_ee_cont_pm": extract_decimal(request, "esic_ee_cont_pm"),
            "esic_ee_cont_pa": extract_decimal(request, "esic_ee_cont_pa"),
            "gratuity_pm": extract_decimal(request, "gratuity_pm"),
            "gratuity_pa": extract_decimal(request, "gratuity_pa"),
            "net_salary_pm": extract_decimal(request, "net_salary_pm"),
            "net_salary_pa": extract_decimal(request, "net_salary_pa"),
            "ctc_pm": extract_decimal(request, "ctc_pm"),
            "ctc_pa": extract_decimal(request, "ctc_pa"),
        }

        if salary_id:
            sm = SalaryMaster.objects.get(pk=salary_id)
            message = "Salary updated successfully."
        else:
            sm = SalaryMaster(employee=employee)
            message = "Salary created successfully."

        sm.employee = employee
        sm.pf_deducted = pf_deducted
        sm.gratuity_applicable = gratuity_applicable
        sm.esic_applicable = esic_applicable

        for field, value in data.items():  # ✅ simple, no redundant logic
            setattr(sm, field, value)

        sm.save()
        messages.success(request, message)
        return redirect("create-salary")

    # GET
    employees = Employee.objects.all()
    salaries = SalaryMaster.objects.all()
    salary_id = request.GET.get("edit")
    salary_obj = SalaryMaster.objects.filter(pk=salary_id).first() if salary_id else None

    return render(request, "salary/create_salary4.html", {
        "employees": employees,
        "salary": salaries,
        "salary_obj": salary_obj,
        "is_edit": bool(salary_obj)
    })


# salary detail view

@login_required
@group_required("Admin", "HR", "Manager")
def salary_details(request, pk):
    try:
        salary = SalaryMaster.objects.select_related('employee').get(pk=pk)
        data = {
            "employee_name": str(salary.employee),
            "gross_ctc_pm": float(salary.gross_ctc_pm),
            "gross_ctc_pa": float(salary.gross_ctc_pa),
            "basic_pm": float(salary.basic_pm),
            "hra_pm": float(salary.hra_pm),
            "stat_bonus_pm": float(salary.stat_bonus_pm),
            "sp_allowance_pm": float(salary.sp_allowance_pm),
            "guaranteed_cash_pm": float(salary.guaranteed_cash_pm),
            "pf_er_cont_pm": float(salary.pf_er_cont_pm),
            "pf_ee_cont_pm": float(salary.pf_ee_cont_pm),
            "esic_er_cont_pm": float(salary.esic_er_cont_pm),
            "esic_ee_cont_pm": float(salary.esic_ee_cont_pm),
            "gratuity_pm": float(salary.gratuity_pm),
            "profession_tax_pm": float(salary.profession_tax_pm),
            "net_salary_pm": float(salary.net_salary_pm),
            "net_salary_pa": float(salary.net_salary_pa),
            "ctc_pm": float(salary.ctc_pm),
            "ctc_pa": float(salary.ctc_pa),
            "pf_deducted": salary.pf_deducted,
            "esic_applicable": salary.esic_applicable,
            "gratuity_applicable": salary.gratuity_applicable,
        }
        return JsonResponse(data)
    except SalaryMaster.DoesNotExist:
        raise Http404("Salary not found")




# employee advance





# def advances_list(request):
#     advances = AdvanceMaster.objects.all().order_by('-created_on')
#     employees = Employee.objects.all()
#     return render(request, 'advance/advance_list.html', {'advances': advances, 'employees': employees})


# @csrf_exempt
# def create_advance(request):
#     if request.method == "POST":
#         emp_id = request.POST.get("employee")
#         total_amount = request.POST.get("total_amount")
#         start_month = request.POST.get("start_month")
#         remarks = request.POST.get("remarks", "")

#         employee = get_object_or_404(Employee, id=emp_id)
#         adv = AdvanceMaster.objects.create(
#             employee=employee,
#             total_amount=total_amount,
#             start_month=start_month + "-01",  # convert YYYY-MM to full date
#             remarks=remarks
#         )
#         return JsonResponse({"status": "success", "message": "Advance created successfully!"})


# @csrf_exempt
# def add_installment(request, advance_id):
#     advance = get_object_or_404(AdvanceMaster, id=advance_id)

#     month = request.POST.get("month")
#     amount = float(request.POST.get("amount"))
#     remarks = request.POST.get("remarks", "")
#     is_paid = request.POST.get("is_paid") == "on"

#     AdvanceInstallment.objects.create(
#         advance=advance,
#         month=month + "-01",
#         amount=amount,
#         is_paid=is_paid,
#         paid_on=timezone.now().date() if is_paid else None,
#         remarks=remarks
#     )

#     if advance.remaining_amount_db <= 0:
#         advance.is_closed = True
#         advance.save()

#     return JsonResponse({"status": "success", "message": "Installment recorded successfully!"})


# def view_installments(request, advance_id):
#     advance = get_object_or_404(AdvanceMaster, id=advance_id)
#     installments = advance.installments.order_by('month')
#     html = render_to_string("partials/installments_table.html", {"installments": installments})
#     return JsonResponse({"html": html})


# @csrf_exempt
# def mark_paid(request, installment_id):
#     inst = get_object_or_404(AdvanceInstallment, id=installment_id)
#     inst.is_paid = True
#     inst.is_skipped = False
#     inst.paid_on = timezone.now().date()
#     inst.save()

#     advance = inst.advance
#     if advance.remaining_amount_db <= 0:
#         advance.is_closed = True
#         advance.save()

#     return JsonResponse({"message": "Installment marked as paid!"})


# @csrf_exempt
# def undo_paid(request, installment_id):
#     inst = get_object_or_404(AdvanceInstallment, id=installment_id)
#     inst.is_paid = False
#     inst.paid_on = None
#     inst.save()

#     inst.advance.is_closed = False
#     inst.advance.save()

#     return JsonResponse({"message": "Installment reverted to pending!"})


# @csrf_exempt
# def skip_installment(request, installment_id):
#     """✅ Mark this month's installment as skipped."""
#     inst = get_object_or_404(AdvanceInstallment, id=installment_id)
#     inst.is_skipped = True
#     inst.is_paid = False
#     inst.paid_on = None
#     inst.save()
#     return JsonResponse({"message": "Installment skipped for this month!"})



# Salary Increment****************




def D(request, key):
    try:
        return Decimal(request.POST.get(key, "0") or "0")
    except:
        return Decimal("0")

def to_float(v):
    try:
        return float(v)
    except:
        return 0.0

# Add these to your views.py
@login_required
@group_required("Admin", "HR")
def create_salary_increment(request):
    if request.method == "POST":
        try:
            employee = Employee.objects.get(id=request.POST.get("employee"))
            effective_date = datetime.strptime(
                request.POST.get("effective_date"), "%Y-%m-%d"
            ).date()

            # Flags
            pf = request.POST.get("pf_deducted") == "yes"
            esic = request.POST.get("esic_applicable") == "yes"
            gratuity = request.POST.get("gratuity_applicable") == "yes"

            # Monthly values
            m = {
                "gross_ctc": to_float(D(request, "gross_ctc_pm")),
                "basic": to_float(D(request, "basic_pm")),
                "hra": to_float(D(request, "hra_pm")),
                "stat_bonus": to_float(D(request, "stat_bonus_pm")),
                "allowance1": to_float(D(request, "allowance1_pm")),
                "allowance2": to_float(D(request, "allowance2_pm")),
                "special_allowance": to_float(D(request, "sp_allowance_pm")),
                "guaranteed_cash": to_float(D(request, "guaranteed_cash_pm")),
                "professional_tax": to_float(D(request, "profession_tax_pm")),
                "pf_er": to_float(D(request, "pf_er_cont_pm")),
                "pf_ee": to_float(D(request, "pf_ee_cont_pm")),
                "esic_er": to_float(D(request, "esic_er_cont_pm")),
                "esic_ee": to_float(D(request, "esic_ee_cont_pm")),
                "gratuity": to_float(D(request, "gratuity_pm")),
                "net_salary": to_float(D(request, "net_salary_pm")),
                "ctc": to_float(D(request, "ctc_pm")),
            }

            # Annual
            a = {k: float(v * 12) for k, v in m.items()}

            SalaryIncrement.objects.create(
                employee=employee,
                effective_date=effective_date,
                change_set={
                    "reason": request.POST.get("reason", ""),
                    "flags": {
                        "pf_deducted": pf,
                        "esic_applicable": esic,
                        "gratuity_applicable": gratuity,
                    },
                    "monthly": m,
                    "annual": a,
                }
            )

            messages.success(request, "Increment created successfully.")
            return redirect("salary_increment")  # ✅ redirect to list

        except Exception as e:
            print("CREATE ERROR:", e)
            messages.error(request, "Failed to create increment.")

    return render(request, "salary_increment/create_increment.html", {
        "employees": Employee.objects.all(),
        "increments": SalaryIncrement.objects.all(),
    })

@login_required
@group_required("Admin", "HR")
def edit_increment(request, pk):
    """Return increment data as JSON for editing"""
    try:
        inc = SalaryIncrement.objects.get(id=pk)
        
        monthly = inc.change_set.get("monthly", {})
        flags = inc.change_set.get("flags", {})
        
        data = {
            "id": inc.id,
            "employee_id": inc.employee.id,
            "effective_date": inc.effective_date.strftime("%Y-%m-%d"),
            "reason": inc.change_set.get("reason", ""),
            
            # Flags
            "pf_deducted": flags.get("pf_deducted", False),
            "esic_applicable": flags.get("esic_applicable", False),
            "gratuity_applicable": flags.get("gratuity_applicable", False),
            
            # Monthly values
            "gross_ctc": monthly.get("gross_ctc", 0),
            "basic": monthly.get("basic", 0),
            "hra": monthly.get("hra", 0),
            "stat_bonus": monthly.get("stat_bonus", 0),
            "allowance1": monthly.get("allowance1", 0),
            "allowance2": monthly.get("allowance2", 0),
            "special_allowance": monthly.get("special_allowance", 0),
            "guaranteed_cash": monthly.get("guaranteed_cash", 0),
            "professional_tax": monthly.get("professional_tax", 0),
            "pf_er": monthly.get("pf_er", 0),
            "pf_ee": monthly.get("pf_ee", 0),
            "esic_er": monthly.get("esic_er", 0),
            "esic_ee": monthly.get("esic_ee", 0),
            "gratuity": monthly.get("gratuity", 0),
            "net_salary": monthly.get("net_salary", 0),
            "ctc": monthly.get("ctc", 0),
        }
        
        return JsonResponse(data)
        
    except SalaryIncrement.DoesNotExist:
        return JsonResponse({"error": "Increment not found"}, status=404)


@login_required
@group_required("Admin", "HR")
def update_salary_increment(request, pk):
    """Update existing increment"""
    if request.method == "POST":
        try:
            inc = SalaryIncrement.objects.get(id=pk)
            
            # Update employee and date
            inc.employee = Employee.objects.get(id=request.POST.get("employee"))
            inc.effective_date = datetime.strptime(request.POST.get("effective_date"), "%Y-%m-%d").date()
            
            # Flags
            pf = request.POST.get("pf_deducted") == "yes"
            esic = request.POST.get("esic_applicable") == "yes"
            gratuity = request.POST.get("gratuity_applicable") == "yes"
            
            # Monthly Values
            m = {
                "gross_ctc": to_float(D(request, "gross_ctc_pm")),
                "basic": to_float(D(request, "basic_pm")),
                "hra": to_float(D(request, "hra_pm")),
                "stat_bonus": to_float(D(request, "stat_bonus_pm")),
                "allowance1": to_float(D(request, "allowance1_pm")),
                "allowance2": to_float(D(request, "allowance2_pm")),
                "special_allowance": to_float(D(request, "sp_allowance_pm")),
                "guaranteed_cash": to_float(D(request, "guaranteed_cash_pm")),
                "professional_tax": to_float(D(request, "profession_tax_pm")),
                "pf_er": to_float(D(request, "pf_er_cont_pm")),
                "pf_ee": to_float(D(request, "pf_ee_cont_pm")),
                "esic_er": to_float(D(request, "esic_er_cont_pm")),
                "esic_ee": to_float(D(request, "esic_ee_cont_pm")),
                "gratuity": to_float(D(request, "gratuity_pm")),
                "net_salary": to_float(D(request, "net_salary_pm")),
                "ctc": to_float(D(request, "ctc_pm")),
            }
            
            # Annual Values
            a = {k: float(v * 12) for k, v in m.items()}
            
            # Update change_set
            inc.change_set = {
                "reason": request.POST.get("reason", ""),
                "flags": {
                    "pf_deducted": pf,
                    "esic_applicable": esic,
                    "gratuity_applicable": gratuity,
                },
                "monthly": m,
                "annual": a,
            }
            
            inc.save()
            messages.success(request, "Increment updated successfully.")
            return redirect("salary_increment")
            
        except SalaryIncrement.DoesNotExist:
            return JsonResponse({"error": "Increment not found"}, status=404)
    
    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
@group_required("Admin", "HR", "Manager")
def increment_details(request, pk):
    inc = get_object_or_404(SalaryIncrement, id=pk)

    return JsonResponse({
        "employee_code": inc.employee.employee_code,
        "employee_name": str(inc.employee),
        "effective_date": inc.effective_date.strftime("%d %b, %Y"),
        "reason": inc.change_set.get("reason", "-"),

        "monthly": inc.change_set.get("monthly", {}),
        "flags": inc.change_set.get("flags", {}),

        "status": "Applied" if inc.is_processed else "Pending"
    })




@login_required
@group_required("Admin", "HR")
def delete_salary_increment(request, pk):
    """Delete increment"""
    if request.method == "POST":
        try:
            inc = SalaryIncrement.objects.get(id=pk)
            
            # Check if already processed
            if inc.is_processed:
                return JsonResponse({
                    "error": "Cannot delete processed increment"
                }, status=400)
            
            inc.delete()
            messages.success(request, "Increment deleted successfully.")
            return JsonResponse({"success": True})
            
        except SalaryIncrement.DoesNotExist:
            return JsonResponse({"error": "Increment not found"}, status=404)
    
    return JsonResponse({"error": "Invalid request"}, status=400)

@login_required
def employee_salary_ajax(request):
    employee_id = request.GET.get("employee_id")

    if not employee_id:
        return JsonResponse({"error": "Employee ID required"}, status=400)

    employee = get_object_or_404(Employee, pk=employee_id)

    salary = (
        SalaryMaster.objects
        .filter(employee=employee, is_active=True)
        .order_by("-effective_date")
        .first()
    )

    if not salary:
        return JsonResponse({"salary_exists": False})

    return JsonResponse({
        "salary_exists": True,

        # Flags
        "pf_deducted": salary.pf_deducted,
        "esic_applicable": salary.esic_applicable,
        "gratuity_applicable": salary.gratuity_applicable,

        # Monthly values
        "grossCTC": salary.gross_ctc_pm,
        "basic": salary.basic_pm,
        "hra": salary.hra_pm,
        "statBonus": salary.stat_bonus_pm,
        "specialAllowance": salary.sp_allowance_pm,
        "allowance1": salary.allowance1_pm,
        "allowance2": salary.allowance2_pm,
        "guaranteedCash": salary.guaranteed_cash_pm,
        "pfEr": salary.pf_er_cont_pm,
        "pfEe": salary.pf_ee_cont_pm,
        "esicEr": salary.esic_er_cont_pm,
        "esicEe": salary.esic_ee_cont_pm,
        "gratuity": salary.gratuity_pm,
        "professionTax": salary.profession_tax_pm,
        "ctc": salary.ctc_pm,
        "netSalary": salary.net_salary_pm,
    })



# -------------------------
# Salary History list view
# -------------------------
@login_required
@group_required("Admin", "HR", "Manager")
def salary_history(request):
    qs = SalaryHistory.objects.select_related("employee").order_by("-end_date")

    # filters
    emp = request.GET.get("employee", "")
    date_from = request.GET.get("from", "")
    date_to = request.GET.get("to", "")

    if emp:
        qs = qs.filter(employee__employee_code__icontains=emp)

    if date_from:
        qs = qs.filter(start_date__gte=date_from)

    if date_to:
        qs = qs.filter(end_date__lte=date_to)

    employees = Employee.objects.all().order_by("employee_code")

    return render(request, "salary_increment/salary_history_list.html", {
        "history": qs,
        "employees": employees,
        "get": request.GET,
    })


# -------------------------
# AJAX: SalaryHistory detail (used by modal)
# -------------------------
@login_required
@group_required("Admin", "HR", "Manager")
def salary_history_detail(request, pk):
    entry = get_object_or_404(SalaryHistory, pk=pk)
    html = render_to_string("partials/salary_history_detail.html", {"entry": entry})
    return JsonResponse({"html": html})

# -------------------------
# Export to Excel
# -------------------------
@login_required
@group_required("Admin", "HR")
def salary_history_export_excel(request):
    qs = SalaryHistory.objects.select_related("employee").order_by("-end_date")
    # apply same filters as list (to keep export consistent)
    q = request.GET.get("q", "").strip()
    emp_code = request.GET.get("employee_code", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q) |
            Q(employee__last_name__icontains=q) |
            Q(employee__employee_code__icontains=q)
        )

    if emp_code:
        qs = qs.filter(
            Q(employee__employee_code__icontains=emp_code) |
            Q(employee__first_name__icontains=emp_code) |
            Q(employee__last_name__icontains=emp_code)
        )

    if date_from:
        try:
            d1 = datetime.strptime(date_from, "%Y-%m-%d").date()
            qs = qs.filter(end_date__gte=d1)
        except ValueError:
            pass

    if date_to:
        try:
            d2 = datetime.strptime(date_to, "%Y-%m-%d").date()
            qs = qs.filter(start_date__lte=d2)
        except ValueError:
            pass

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary History"

    headers = [
        "Employee Code", "Employee Name", "Start Date", "End Date",
        "Gross CTC (PM)", "Basic (PM)", "HRA (PM)", "Net Salary (PM)",
        "PF Deducted", "ESIC Applicable", "Gratuity Applicable", "Raw JSON"
    ]
    ws.append(headers)

    for e in qs:
        emp = e.employee
        salary = (e.data or {}).get("salary", {})
        row = [
            getattr(emp, "employee_code", ""),
            f"{getattr(emp,'first_name','') or ''} {getattr(emp,'last_name','') or ''}".strip(),
            e.start_date.isoformat() if e.start_date else "",
            e.end_date.isoformat() if e.end_date else "",
            salary.get("gross_ctc_pm", ""),
            salary.get("basic_pm", ""),
            salary.get("hra_pm", ""),
            salary.get("net_salary_pm", ""),
            e.data.get("pf_deducted", "") if e.data else "",
            e.data.get("esic_applicable", "") if e.data else "",
            e.data.get("gratuity_applicable", "") if e.data else "",
            str(e.data) if e.data else "",
        ]
        ws.append(row)

    # Prepare response
    f = io.BytesIO()
    wb.save(f)
    f.seek(0)
    filename = "salary_history.xlsx"
    resp = HttpResponse(f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# -------------------------
# Export to PDF (WeasyPrint)
# -------------------------
# from xhtml2pdf import pisa

# def salary_history_export_pdf(request):
#     qs = SalaryHistory.objects.select_related("employee").order_by("-end_date")

#     # Apply same filters used in list page
#     q = request.GET.get("q", "").strip()
#     emp_code = request.GET.get("employee_code", "").strip()
#     date_from = request.GET.get("date_from", "").strip()
#     date_to = request.GET.get("date_to", "").strip()

#     if q:
#         qs = qs.filter(
#             Q(employee__first_name__icontains=q) |
#             Q(employee__last_name__icontains=q) |
#             Q(employee__employee_code__icontains=q)
#         )

#     if emp_code:
#         qs = qs.filter(
#             Q(employee__employee_code__icontains=emp_code) |
#             Q(employee__first_name__icontains=emp_code) |
#             Q(employee__last_name__icontains=emp_code)
#         )

#     if date_from:
#         try:
#             d1 = datetime.strptime(date_from, "%Y-%m-%d").date()
#             qs = qs.filter(end_date__gte=d1)
#         except ValueError:
#             pass

#     if date_to:
#         try:
#             d2 = datetime.strptime(date_to, "%Y-%m-%d").date()
#             qs = qs.filter(start_date__lte=d2)
#         except ValueError:
#             pass

#     # Render HTML
#     html_string = render(request,"website/salary_history_pdf.html", {
#         "history": qs
#     })

#     # Generate PDF
#     response = HttpResponse(content_type="application/pdf")
#     response['Content-Disposition'] = 'attachment; filename="salary_history.pdf"'

#     pisa_status = pisa.CreatePDF(
#         html_string,
#         dest=response
#     )

#     if pisa_status.err:
#         return HttpResponse("Error creating PDF", status=500)

#     return response


# -------------------------
# Compare: return JSON of compare table (modal)
# -------------------------
@login_required
@group_required("Admin", "HR", "Manager")
def salary_compare(request, history_id, employee_id):
    history = SalaryHistory.objects.get(id=history_id)
    employee = Employee.objects.get(id=employee_id)
    current = SalaryMaster.objects.filter(employee=employee).first()

    # Convert SalaryMaster model → dict like history.data["salary"]
    current_salary = {}
    if current:
        for field in [
            "gross_ctc_pm", "basic_pm", "hra_pm", "stat_bonus_pm",
            "sp_allowance_pm", "allowance1_pm", "allowance2_pm",
            "guaranteed_cash_pm", "ctc_pm",
            "pf_er_cont_pm", "pf_ee_cont_pm",
            "esic_er_cont_pm", "esic_ee_cont_pm",
            "profession_tax_pm", "net_salary_pm"
        ]:
            current_salary[field] = str(getattr(current, field, 0) or 0)

    components = {
        "gross_ctc_pm": "Gross CTC",
        "basic_pm": "Basic",
        "hra_pm": "HRA",
        "stat_bonus_pm": "Stat Bonus",
        "sp_allowance_pm": "Special Allowance",
        "allowance1_pm": "Allowance 1",
        "allowance2_pm": "Allowance 2",
        "guaranteed_cash_pm": "Guaranteed Cash",
        "ctc_pm": "CTC",
        "pf_er_cont_pm": "PF Employer",
        "pf_ee_cont_pm": "PF Employee",
        "esic_er_cont_pm": "ESIC Employer",
        "esic_ee_cont_pm": "ESIC Employee",
        "profession_tax_pm": "Profession Tax",
        "net_salary_pm": "Net Salary",
    }

    html = render_to_string("partials/salary_compare.html", {
        "employee": employee,
        "old": history.data["salary"],
        "current": current_salary,      # 👉 FIXED
        "components": components
    })

    return JsonResponse({"html": html})

# -------------------------
# Chart data endpoint for timeline
@login_required
# -------------------------
def salary_timeline_data(request, employee_id):
    # Return JSON data: labels = list of end_date, data = net_salary_pm values
    entries = SalaryHistory.objects.filter(employee_id=employee_id).order_by("end_date")
    labels = [e.end_date.isoformat() if e.end_date else "" for e in entries]
    data_points = [float((e.data or {}).get("salary", {}).get("net_salary_pm") or 0) for e in entries]
    return JsonResponse({"labels": labels, "data": data_points})


from django.core.files.storage import default_storage

@login_required
def upload_salary_increment(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        file_path = default_storage.save(f"temp/{excel_file.name}", excel_file)
        
        try:
            df = pd.read_excel(file_path)
            for _, row in df.iterrows():
                try:
                    employee = Employee.objects.get(pk=int(row["employee_id"]))
                    
                    new_increment = SalaryIncrement(
                        employee=employee,
                        pf_deducted=str(row["pf_deducted"]).lower() == 'yes',
                        gratuity_applicable=str(row["gratuity_applicable"]).lower() == 'yes',
                        esic_applicable=str(row["esic_applicable"]).lower() == 'yes',
                        
                        gross_ctc_pm=Decimal(row["gross_ctc_pm"]),
                        gross_ctc_pa=Decimal(row["gross_ctc_pm"]) * 12,
                        basic_pm=Decimal(row["basic_pm"]),
                        basic_pa=Decimal(row["basic_pm"]) * 12,
                        hra_pm=Decimal(row["hra_pm"]),
                        hra_pa=Decimal(row["hra_pm"]) * 12,
                        stat_bonus_pm=Decimal(row["stat_bonus_pm"]),
                        stat_bonus_pa=Decimal(row["stat_bonus_pm"]) * 12,
                        sp_allowance_pm=Decimal(row["special_allowance_pm"]),
                        sp_allowance_pa=Decimal(row["special_allowance_pm"]) * 12,
                        allowance1_pm=Decimal(row["allowance1_pm"]),
                        allowance1_pa=Decimal(row["allowance1_pm"]) * 12,
                        allowance2_pm=Decimal(row["allowance2_pm"]),
                        allowance2_pa=Decimal(row["allowance2_pm"]) * 12,
                        guaranteed_cash_pm=Decimal(row["guaranteed_cash_pm"]),
                        guaranteed_cash_pa=Decimal(row["guaranteed_cash_pm"]) * 12,
                        ctc_pm=Decimal(row["cost_to_company_pm"]),
                        ctc_pa=Decimal(row["cost_to_company_pm"]) * 12,
                        pf_er_cont_pm=Decimal(row["pf_er_cont_pm"]),
                        pf_er_cont_pa=Decimal(row["pf_er_cont_pm"]) * 12,
                        esic_er_cont_pm=Decimal(row["esic_er_cont_pm"]),
                        esic_er_cont_pa=Decimal(row["esic_er_cont_pm"]) * 12,
                        pf_ee_cont_pm=Decimal(row["pf_ee_cont_pm"]),
                        pf_ee_cont_pa=Decimal(row["pf_ee_cont_pm"]) * 12,
                        esic_ee_cont_pm=Decimal(row["esic_ee_cont_pm"]),
                        esic_ee_cont_pa=Decimal(row["esic_ee_cont_pm"]) * 12,
                        profession_tax_pm=Decimal(row["profession_tax_pm"]),
                        profession_tax_pa=Decimal(row["profession_tax_pm"]) * 12,
                        net_salary_pm=Decimal(row["net_salary_pm"]),
                        net_salary_pa=Decimal(row["net_salary_pm"]) * 12,
                    )
                    new_increment.save()
                except Exception as e:
                    messages.error(request, f"Error processing row {row}: {str(e)}")
        except Exception as e:
            messages.error(request, f"Error reading file: {str(e)}")
        finally:
            default_storage.delete(file_path)
        
        messages.success(request, "Salary increments uploaded successfully!")
        return redirect("upload_salary_increment")

    return render(request, "salary_increment/upload_increment.html")



# ── helpers ──────────────────────────────────────────────────────────────────


# ── views ────────────────────────────────────────────────────────────────────

@login_required
def get_payroll_settings(request):
    # If an employee_id is provided (global users creating salaries for other companies),
    # use that employee's company settings instead of the user's own company.
    employee_id = request.GET.get('employee_id')
    if employee_id and user_has_global_access(request.user):
        try:
            emp = Employee.objects.select_related('company').get(id=employee_id)
            company = emp.company
        except Employee.DoesNotExist:
            company = get_user_company(request.user)
    else:
        company = get_user_company(request.user)

    if company is None:
        return JsonResponse(
            {'error': 'No company linked to your account.'},
            status=400
        )

    settings, _ = PayrollSettings.objects.get_or_create(company=company)

    return JsonResponse({
        'pf_percentage':        float(settings.pf_percentage),
        'esic_percentage':      float(settings.esic_percentage),
        'gratuity_percentage':  float(settings.gratuity_percentage),
        'professional_tax':     float(settings.professional_tax),
        'bonus_percentage':     float(settings.bonus_percentage),
        'basic_percentage':     float(settings.basic_percentage),
        'hra_percentage':       float(settings.hra_percentage),
        'basic_cap':            float(settings.basic_cap),
        'pf_wage_ceiling':      float(settings.pf_wage_ceiling),
    })


@login_required
def settings_page(request):
    return redirect('company-settings-hub')
    company = get_user_company(request.user)
    if company is None:
        # You can redirect to an error page or show a message instead
        return render(request, 'settings/no_company.html', {
            'error': 'Your account is not linked to any company.'
        })

    payroll_settings, _ = PayrollSettings.objects.get_or_create(company=company)

    # ✅ FIX: LeaveSettings must also be scoped to company
    # Your LeaveSettings model needs a company FK — see note below
    leave_settings, _ = LeaveSettings.objects.get_or_create(company=company)

    context = {
        'company':          company,
        'payroll_settings': payroll_settings,
        'leave_settings':   leave_settings,
    }
    return render(request, 'settings/settings_page.html', context)


@login_required
@require_http_methods(["POST"])
def save_payroll_settings(request):
    company_id_param = request.POST.get('company_id')
    if user_has_global_access(request.user) and company_id_param:
        try:
            company = Company.objects.get(pk=company_id_param)
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Company not found.'}, status=400)
    else:
        company = get_user_company(request.user)
        if company is None:
            return JsonResponse({'success': False, 'error': 'No company linked to your account.'}, status=400)

    try:
        settings, _ = PayrollSettings.objects.get_or_create(company=company)

        # Capture which leave-affecting flags changed before overwriting them
        old_late_marks_affect_lwp = settings.late_marks_affect_lwp
        old_weekend_days = settings.weekend_days

        # Boolean
        settings.is_auto = request.POST.get("is_auto") == "on"

        # Integers
        settings.from_date             = int(request.POST.get("from_date")) if request.POST.get("from_date") else None
        settings.to_date               = int(request.POST.get("to_date"))   if request.POST.get("to_date")   else None
        settings.grace_period_minutes  = int(request.POST.get("grace_period_minutes", 15))
        settings.max_leave_balance     = int(request.POST.get("max_leave_balance", 30))
        settings.earned_leaves_per_year = int(request.POST.get("earned_leaves_per_year", 12))

        # Floats / Decimals
        settings.basic_percentage    = float(request.POST.get("basic_percentage",    50))
        settings.hra_percentage      = float(request.POST.get("hra_percentage",      60))
        settings.basic_cap           = float(request.POST.get("basic_cap",        21000))
        settings.pf_wage_ceiling     = float(request.POST.get("pf_wage_ceiling",   15000))
        settings.pf_percentage       = float(request.POST.get("pf_percentage",       12))
        settings.esic_percentage     = float(request.POST.get("esic_percentage",   3.67))
        settings.gratuity_percentage = float(request.POST.get("gratuity_percentage", 4.61))
        settings.bonus_percentage    = float(request.POST.get("bonus_percentage",   8.33))
        settings.professional_tax    = float(request.POST.get("professional_tax",    200))

        # Financial year + branch-specific holidays
        fy_month = request.POST.get("financial_year_start_month")
        if fy_month:
            old_fy = settings.financial_year_start_month
            new_fy = int(fy_month)
            settings.financial_year_start_month = new_fy

            if old_fy != new_fy:
                MonthlyEarnedLeaves.objects.filter(
                    payroll_settings=settings,
                    is_auto_generated=True
                ).delete()
                MonthlyEarnedLeaves.generate_for_payroll_settings(settings)

        branch_specific = request.POST.get("branch_specific_holidays")
        if branch_specific is not None:
            settings.branch_specific_holidays = branch_specific == "true"

        weekend_days = request.POST.get("weekend_days")
        if weekend_days in ('sat_sun', 'sun'):
            settings.weekend_days = weekend_days

        settings.late_marks_affect_lwp = request.POST.get("late_marks_affect_lwp") == "on"

        settings.save()
        MonthlyEarnedLeaves.sync_with_payroll_settings(settings)

        # If late-mark or weekend setting changed, auto-recalculate leave balances so
        # the report reflects the new rules immediately without a manual recalculate step.
        recalc_msg = ''
        if (settings.late_marks_affect_lwp != old_late_marks_affect_lwp
                or settings.weekend_days != old_weekend_days):
            try:
                recalc_count = generate_leave_balances_for_all_periods(company, settings)
                recalc_msg = f' Leave balances auto-recalculated ({recalc_count} record(s) updated).'
            except Exception:
                recalc_msg = ' (Leave balance recalculation failed — use Recalculate button manually.)'

        return JsonResponse({'success': True, 'message': f'Settings saved successfully!{recalc_msg}'})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

        
@login_required
@require_http_methods(["POST"])
def save_leave_settings(request):
    company_id_param = request.POST.get('company_id')
    if user_has_global_access(request.user) and company_id_param:
        try:
            company = Company.objects.get(pk=company_id_param)
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Company not found.'}, status=400)
    else:
        company = get_user_company(request.user)
        if company is None:
            return JsonResponse({'success': False, 'error': 'No company linked to your account.'}, status=400)

    try:
        settings, _ = LeaveSettings.objects.get_or_create(company=company)

        settings.carry_forward = request.POST.get("carry_forward") == "on"
        reset_month = request.POST.get("reset_month")
        settings.reset_month = int(reset_month) if reset_month else None

        settings.save()

        return JsonResponse({"success": True, "message": "Leave settings saved successfully!"})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required
@require_http_methods(["POST"])
def save_leave_credit_policy(request):
    company_id_param = request.POST.get('company_id')
    if user_has_global_access(request.user) and company_id_param:
        try:
            company = Company.objects.get(pk=company_id_param)
        except Company.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Company not found.'}, status=400)
    else:
        company = get_user_company(request.user)
        if company is None:
            return JsonResponse({'success': False, 'error': 'No company linked to your account.'}, status=400)

    try:
        policy, _ = LeaveCreditPolicy.objects.get_or_create(company=company)
        policy.credit_1_limit = int(request.POST.get("credit_1_limit", 15))
        policy.credit_2_limit = int(request.POST.get("credit_2_limit", 25))
        policy.credit_low = Decimal(request.POST.get("credit_low", 0))
        policy.credit_mid = Decimal(request.POST.get("credit_mid", 1))
        policy.credit_high = Decimal(request.POST.get("credit_high", 2))
        policy.save()
        return JsonResponse({"success": True, "message": "Leave Credit Policy saved successfully!"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@login_required
def company_settings_hub(request):
    if not user_has_global_access(request.user):
        messages.error(request, "You don't have permission to access the company settings hub.")
        return redirect('settings-page')

    companies = Company.objects.filter(status='active').order_by('name')

    # Which companies already have payroll settings saved
    companies_with_payroll = set(
        PayrollSettings.objects.filter(company__status='active').values_list('company_id', flat=True)
    )

    company_id = request.GET.get('company_id')
    selected_company = None
    payroll_settings = None
    leave_settings = None
    credit_policy = None

    if company_id:
        try:
            selected_company = Company.objects.get(pk=company_id, status='active')
        except Company.DoesNotExist:
            messages.error(request, "Company not found.")

    if selected_company is None and companies.exists():
        selected_company = companies.first()

    if selected_company:
        payroll_settings, _ = PayrollSettings.objects.get_or_create(company=selected_company)
        leave_settings, _ = LeaveSettings.objects.get_or_create(company=selected_company)
        credit_policy, _ = LeaveCreditPolicy.objects.get_or_create(company=selected_company)

    missing_count = companies.exclude(pk__in=companies_with_payroll).count()
    if selected_company:
        missing_count = companies.exclude(pk=selected_company.pk).exclude(pk__in=companies_with_payroll).count()

    return render(request, 'settings/company_settings_hub.html', {
        'companies': companies,
        'selected_company': selected_company,
        'payroll_settings': payroll_settings,
        'leave_settings': leave_settings,
        'credit_policy': credit_policy,
        'companies_with_payroll': companies_with_payroll,
        'missing_settings_count': missing_count,
        'total_other_companies': companies.exclude(pk=selected_company.pk).count() if selected_company else 0,
    })


@login_required
@group_required("Admin")
@require_http_methods(["POST"])
def broadcast_settings_to_all_companies(request):
    """Copy all settings from one company to all other active companies."""
    source_company_id = request.POST.get('source_company_id')
    if not source_company_id:
        return JsonResponse({'success': False, 'error': 'Source company ID is required.'}, status=400)

    try:
        source_company = Company.objects.get(pk=source_company_id, status='active')
    except Company.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Company not found.'}, status=400)

    try:
        source_payroll = PayrollSettings.objects.get(company=source_company)
    except PayrollSettings.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'No payroll settings found for "{source_company.name}". Please save settings first, then broadcast.'
        }, status=400)

    source_leave, _ = LeaveSettings.objects.get_or_create(company=source_company)
    source_policy, _ = LeaveCreditPolicy.objects.get_or_create(company=source_company)

    EXCLUDE = {'id', 'company', 'company_id', 'created_at', 'updated_at'}

    payroll_fields = {f.name: getattr(source_payroll, f.name) for f in PayrollSettings._meta.fields if f.name not in EXCLUDE}
    leave_fields   = {f.name: getattr(source_leave,   f.name) for f in LeaveSettings._meta.fields   if f.name not in EXCLUDE}
    policy_fields  = {f.name: getattr(source_policy,  f.name) for f in LeaveCreditPolicy._meta.fields if f.name not in EXCLUDE}

    target_companies = Company.objects.filter(status='active').exclude(pk=source_company_id)
    updated = 0
    errors = []

    for company in target_companies:
        try:
            with transaction.atomic():
                # PayrollSettings
                ps, _ = PayrollSettings.objects.get_or_create(company=company)
                old_fy = ps.financial_year_start_month
                for field, value in payroll_fields.items():
                    setattr(ps, field, value)
                ps.save()
                # Regenerate earned leaves
                if old_fy != ps.financial_year_start_month:
                    MonthlyEarnedLeaves.objects.filter(payroll_settings=ps, is_auto_generated=True).delete()
                MonthlyEarnedLeaves.generate_for_payroll_settings(ps)
                MonthlyEarnedLeaves.sync_with_payroll_settings(ps)

                # LeaveSettings
                ls, _ = LeaveSettings.objects.get_or_create(company=company)
                for field, value in leave_fields.items():
                    setattr(ls, field, value)
                ls.save()

                # LeaveCreditPolicy
                lcp, _ = LeaveCreditPolicy.objects.get_or_create(company=company)
                for field, value in policy_fields.items():
                    setattr(lcp, field, value)
                lcp.save()

                updated += 1
        except Exception as e:
            errors.append(f"{company.name}: {str(e)}")

    msg = f'Settings from "{source_company.name}" applied to {updated} company(ies).'
    if errors:
        msg += f' Errors: {"; ".join(errors)}'

    return JsonResponse({
        'success': updated > 0 or len(errors) == 0,
        'message': msg,
        'updated_count': updated,
        'errors': errors,
    })


@login_required
@group_required("Admin", "HR")
def advance_list(request):
    status_filter = request.GET.get('status', '').strip()
    search = request.GET.get('q', '').strip()

    qs = AdvanceMaster.objects.select_related('employee').prefetch_related('schedules').order_by('-created_at')

    if status_filter in ('active', 'completed'):
        qs = qs.filter(status=status_filter)

    if search:
        from django.db.models import Q as _Q
        qs = qs.filter(
            _Q(employee__first_name__icontains=search) |
            _Q(employee__last_name__icontains=search) |
            _Q(employee__employee_code__icontains=search)
        )

    all_qs = AdvanceMaster.objects.all()
    active_qs = all_qs.filter(status='active')
    stats = {
        'total': all_qs.count(),
        'active': active_qs.count(),
        'completed': all_qs.filter(status='completed').count(),
        'outstanding': sum(a.outstanding_amount for a in active_qs),
    }

    advances = list(qs)
    for adv in advances:
        paid = adv.advance_amount - adv.outstanding_amount
        adv.paid_amount = paid
        adv.progress = round((paid / adv.advance_amount) * 100) if adv.advance_amount else 0
        adv.schedules_paid = adv.schedules.filter(status='paid').count()

    return render(request, 'advances/advance_list.html', {
        'advances': advances,
        'stats': stats,
        'status_filter': status_filter,
        'search': search,
    })


@login_required
@group_required("Admin", "HR")
def advance_create(request):
    """Admin/HR creates an advance."""
    # if not request.user.is_staff:
    #     return HttpResponseBadRequest("Not allowed")

    if request.method == 'POST':
        form = AdvanceCreateForm(request.POST)
        if form.is_valid():
            # Use service create_advance so schedules auto-created
            data = form.cleaned_data
            employee = data['employee']
            amount = data['advance_amount']
            months = data['default_months']
            start_date = data.get('start_date') or None
            adv = create_advance(employee, amount, months, start_date=start_date)
            messages.success(request, f"Advance created for {employee} - ₹{amount}")
            return redirect('advances-list')
    else:
        form = AdvanceCreateForm(initial={'start_date': date.today().replace(day=1)})
    return render(request, 'advances/advance_create.html', {'form': form})



@login_required
@group_required("Admin", "HR")
def advance_detail(request, pk):
    """Show schedule, payments, actions (pay/skip)."""
    adv = get_object_or_404(AdvanceMaster, pk=pk)
    # security: if non-staff, ensure this belongs to user
    # if not request.user.is_staff:
    #     try:
    #         if request.user.employee != adv.employee:
    #             return HttpResponseBadRequest("Not allowed")
    #     except:
    #         return HttpResponseBadRequest("Not allowed")

    schedules = adv.schedules.order_by('due_month')
    payments = adv.payments.order_by('-date')
    payment_form = PaymentForm()
    skip_form = SkipMonthForm()
    adv.paid_amount = adv.advance_amount - adv.outstanding_amount
    adv.progress = round((adv.paid_amount / adv.advance_amount) * 100) if adv.advance_amount else 0
    adv.schedules_paid = schedules.filter(status='paid').count()
    adv.schedules_total = schedules.count()
    return render(request, 'advances/advance_detail.html', {
        'advance': adv,
        'schedules': schedules,
        'payments': payments,
        'payment_form': payment_form,
        'skip_form': skip_form
    })

@login_required
@group_required("Admin", "HR")
@require_POST
@transaction.atomic
def pay_advance(request, pk):
    """AJAX / form POST to apply payment."""
    adv = get_object_or_404(AdvanceMaster, pk=pk)
    # authorization check
    # if not request.user.is_staff:
    #     try:
    #         if request.user.employee != adv.employee:
    #             return JsonResponse({'error': 'Not allowed'}, status=403)
    #     except:
    #         return JsonResponse({'error': 'Not allowed'}, status=403)

    form = PaymentForm(request.POST)
    if not form.is_valid():
        return JsonResponse({'error': 'Invalid data', 'errors': form.errors}, status=400)
    amount = form.cleaned_data['amount']
    note = form.cleaned_data.get('note', '')

    # Use service to apply payment
    try:
        apply_payment(adv, amount, note=note)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    schedules = [
        {
            'due_month': s.due_month.strftime('%b %Y'),
            'due_month_iso': s.due_month.strftime('%Y-%m-%d'),
            'scheduled_amount': s.scheduled_amount,
            'paid_amount': s.paid_amount,
            'status': s.status,
        }
        for s in adv.schedules.order_by('due_month')
    ]
    return JsonResponse({
        'success': True,
        'outstanding': adv.outstanding_amount,
        'advance_status': adv.status,
        'schedules': schedules,
    })

@login_required
@group_required("Admin", "HR")
@require_POST
@transaction.atomic
def skip_advance_month(request, pk):
    """Mark a schedule month as skipped (AJAX). Expects due_month in POST (YYYY-MM-DD)."""
    adv = get_object_or_404(AdvanceMaster, pk=pk)
    # if not request.user.is_staff:
    #     try:
    #         if request.user.employee != adv.employee:
    #             return JsonResponse({'error': 'Not allowed'}, status=403)
    #     except:
    #         return JsonResponse({'error': 'Not allowed'}, status=403)

    form = SkipMonthForm(request.POST)
    if not form.is_valid():
        return JsonResponse({'error': 'Invalid data', 'errors': form.errors}, status=400)

    due_month = form.cleaned_data['due_month']
    try:
        skip_month(adv, due_month)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

    schedules = [
        {
            'due_month': s.due_month.strftime('%b %Y'),
            'due_month_iso': s.due_month.strftime('%Y-%m-%d'),
            'scheduled_amount': s.scheduled_amount,
            'paid_amount': s.paid_amount,
            'status': s.status,
        }
        for s in adv.schedules.order_by('due_month')
    ]
    return JsonResponse({
        'success': True,
        'outstanding': adv.outstanding_amount,
        'advance_status': adv.status,
        'schedules': schedules,
    })


@login_required
@group_required("Admin", "HR")
@require_POST
def revert_skip_view(request, pk):
    adv = get_object_or_404(AdvanceMaster, pk=pk)
    due_month = request.POST.get("due_month")

    try:
        due_month = date.fromisoformat(due_month)
        revert_skip(adv, due_month)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)

    updated = [
        {
            "due_month": s.due_month.strftime("%b %Y"),
            "due_month_iso": s.due_month.strftime("%Y-%m-%d"),
            "scheduled_amount": s.scheduled_amount,
            "paid_amount": s.paid_amount,
            "status": s.status,
        }
        for s in adv.schedules.order_by("due_month")
    ]
    return JsonResponse({"success": True, "schedules": updated, "advance_status": adv.status})




@login_required
def payroll_run_list(request):
    companies = Company.objects.filter(status="active").order_by("name")
    company_id = request.GET.get("company")
    status_filter = request.GET.get("status")

    runs = PayrollRun.objects.select_related("company").prefetch_related("records").order_by("-month")

    if company_id:
        runs = runs.filter(company_id=company_id)
    if status_filter:
        runs = runs.filter(status=status_filter)

    draft_count = runs.filter(status=PayrollRun.STATUS_DRAFT).count()
    finalized_count = runs.filter(status=PayrollRun.STATUS_FINALIZED).count()

    selected_company = None
    if company_id:
        selected_company = companies.filter(id=company_id).first()

    return render(request, "payroll/run_list.html", {
        "runs": runs,
        "companies": companies,
        "selected_company_id": company_id or "",
        "selected_company": selected_company,
        "status_filter": status_filter or "",
        "draft_count": draft_count,
        "finalized_count": finalized_count,
    })


def _compute_payroll_dates(ps, year, m):
    """Return (month_start, month_end) for the given year/month using company's payroll cycle."""
    if ps and ps.from_date and ps.to_date:
        from_day, to_day = ps.from_date, ps.to_date
        if from_day <= to_day:
            # Same-month cycle (e.g. 1 → 30)
            month_start = date(year, m, from_day)
            month_end = date(year, m, to_day)
        else:
            # Cross-month cycle (e.g. 27 → 26): month m is the ending month
            month_end = date(year, m, to_day)
            prev_m = m - 1 if m > 1 else 12
            prev_y = year if m > 1 else year - 1
            month_start = date(prev_y, prev_m, from_day)
    else:
        # Calendar month fallback
        month_start = date(year, m, 1)
        nxt = month_start.replace(day=28) + timedelta(days=4)
        month_end = nxt - timedelta(days=nxt.day)
    return month_start, month_end


@login_required
def payroll_run_create(request):
    companies = Company.objects.filter(status="active").order_by("name")
    if request.method == "POST":
        company_id = request.POST.get("company")
        month = request.POST.get("month")  # "YYYY-MM"
        if not company_id or not month:
            messages.error(request, "Company and month are required.")
            return render(request, "payroll/run_create.html", {"companies": companies})

        try:
            year, m = map(int, month.split("-"))
        except ValueError:
            messages.error(request, "Invalid month format.")
            return render(request, "payroll/run_create.html", {"companies": companies})

        if company_id == "all":
            created, skipped = [], []
            for company in companies:
                ps = PayrollSettings.objects.filter(company=company).first()
                month_start, month_end = _compute_payroll_dates(ps, year, m)
                if PayrollRun.objects.filter(company=company, start_date=month_start, end_date=month_end).exists():
                    skipped.append(company.name)
                    continue
                create_payroll_run(company, month_start, month_end)
                created.append(company.name)

            if created:
                messages.success(
                    request,
                    f"Payroll runs created for {len(created)} compan{'y' if len(created) == 1 else 'ies'}: "
                    f"{', '.join(created)}."
                )
            if skipped:
                messages.warning(
                    request,
                    f"Skipped {len(skipped)} compan{'y' if len(skipped) == 1 else 'ies'} that already had a run "
                    f"for this period: {', '.join(skipped)}."
                )
            if not created and not skipped:
                messages.error(request, "No active companies found.")
            return redirect("payroll-run-list")

        company = get_object_or_404(Company, id=company_id)
        ps = PayrollSettings.objects.filter(company=company).first()
        month_start, month_end = _compute_payroll_dates(ps, year, m)

        if PayrollRun.objects.filter(company=company, start_date=month_start, end_date=month_end).exists():
            messages.error(
                request,
                f"A payroll run already exists for {month_start:%d %b %Y} – {month_end:%d %b %Y}."
            )
            return render(request, "payroll/run_create.html", {"companies": companies})

        run = create_payroll_run(company, month_start, month_end)
        messages.success(request, f"Payroll run created: {month_start:%d %b %Y} – {month_end:%d %b %Y}")
        return redirect("payroll-run-detail", run_id=run.id)

    return render(request, "payroll/run_create.html", {"companies": companies})


@login_required
def payroll_period_preview(request):
    """AJAX: return payroll period dates for a given company + month."""
    company_id = request.GET.get("company_id")
    month = request.GET.get("month")  # "YYYY-MM"
    if not company_id or not month:
        return JsonResponse({"error": "Missing params"}, status=400)
    try:
        year, m = map(int, month.split("-"))
        company = Company.objects.get(id=company_id)
    except Exception:
        return JsonResponse({"error": "Invalid params"}, status=400)

    ps = PayrollSettings.objects.filter(company=company).first()
    start, end = _compute_payroll_dates(ps, year, m)

    existing = PayrollRun.objects.filter(company=company, start_date=start, end_date=end).first()
    return JsonResponse({
        "start": start.strftime("%d %b %Y"),
        "end": end.strftime("%d %b %Y"),
        "days": (end - start).days + 1,
        "existing_id": existing.id if existing else None,
    })

@login_required
def payroll_run_detail(request, run_id):
    run = get_object_or_404(PayrollRun, id=run_id)
    records = run.records.select_related("employee").all()
    settings = PayrollSettings.objects.filter(company=run.company).first()
    return render(request, "payroll/run_detail.html", {"run": run, "records": records, "settings": settings})

@login_required
@group_required("Admin", "HR")
@require_POST
def payroll_record_update(request, record_id):
    record = get_object_or_404(PayrollRecord, id=record_id)
    if record.payroll.status == PayrollRun.STATUS_FINALIZED:
        return JsonResponse({"success": False, "error": "Payroll finalized"}, status=400)
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    # apply edits to record fields
    editable = ["present_days", "leave_without_pay", "pf_employee", "professional_tax", "advance", "tds", "other_deductions", "esic_employee"]
    manual = {}
    for k in editable:
        if k in payload:
            try:
                val = Decimal(payload[k])
            except Exception:
                val = Decimal(0)
            setattr(record, k, val)
            manual[k] = float(val)

    # save then recalc authoritative
    record.save()
    recalc_and_save_record(record, manual_overrides=manual)

    return JsonResponse({
        "success": True,
        "record": {
            "id": record.id,
            "net_salary": float(record.net_salary),
            "total_deductions": float(record.total_deductions),
            "calculation_breakdown": record.calculation_breakdown
        }
    })

@login_required
@group_required("Admin", "HR")
@require_POST
def payroll_run_finalize(request, run_id):
    run = get_object_or_404(PayrollRun, id=run_id)
    if run.status == PayrollRun.STATUS_FINALIZED:
        return JsonResponse({"success": False, "error": "This payroll run is already finalized."}, status=400)
    for rec in run.records.all():
        recalc_and_save_record(rec, manual_overrides=rec.manual_override or {})
    run.status = PayrollRun.STATUS_FINALIZED
    run.save()
    return JsonResponse({"success": True})




from openpyxl import Workbook
from django.http import HttpResponse
from .models import PayrollRun, PayrollRecord

@login_required
def payroll_export_excel(request, run_id):
    run = PayrollRun.objects.get(id=run_id)
    records = PayrollRecord.objects.filter(payroll=run)   # ✅ FIXED FIELD NAME

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.month.strftime('%b %Y')}"

    headers = [
        "Employee Code", "Name", "Designation", "Branch",
        "Gross CTC", "Basic (PM)", "HRA (PM)", "Sp Allow (PM)", "Stat Bonus (PM)",
        "Allowance1", "Allowance2",
        "Total Days", "Present Days", "Leave Taken", "Leave Without Pay",  # LWP included
        "Basic Processed", "HRA Processed", "Sp Allow Processed", "Stat Bonus Processed",
        "Allowance1 Processed", "Allowance2 Processed", "Gross Processed",
        "PF (EE)", "Professional Tax", "Advance", "ESIC (EE)", "TDS", "Other Deductions",
        "Total Deductions", "Net Pay"
    ]

    ws.append(headers)

    for r in records:
        ws.append([
            r.employee_code, r.employee_name, r.designation, r.branch_name,
            float(r.gross_ctc), float(r.basic_pm), float(r.hra_pm), float(r.sp_allowance_pm), float(r.stat_bonus_pm),
            float(r.allowance1_pm), float(r.allowance2_pm),
            r.total_days, float(r.present_days), float(r.leave_taken), float(r.leave_without_pay),
            float(r.basic_processed), float(r.hra_processed), float(r.sp_allowance_processed), float(r.stat_bonus_processed),
            float(r.allowance1_processed), float(r.allowance2_processed), float(r.gross_processed),
            float(r.pf_employee), float(r.professional_tax), float(r.advance), float(r.esic_employee), float(r.tds), float(r.other_deductions),
            float(r.total_deductions), float(r.net_salary)
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Payroll_{run.month.strftime('%b_%Y')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response




from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse
from io import BytesIO

@login_required
def payroll_export_pdf(request, run_id):
    run = PayrollRun.objects.get(id=run_id)
    records = PayrollRecord.objects.filter(payroll=run)   # ✅ FIXED FIELD NAME

    template = get_template("payroll/payroll_pdf_template.html")
    html = template.render({"run": run, "records": records})

    pdf_file = BytesIO()
    pisa_status = pisa.CreatePDF(html, dest=pdf_file)

    if pisa_status.err:
        return HttpResponse("PDF generation failed", status=500)

    response = HttpResponse(pdf_file.getvalue(), content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="Payroll_{run.month.strftime("%b_%Y")}.pdf"'

    return response


# ---------------------------------------------------------------------------
# SALARY SLIP
# ---------------------------------------------------------------------------

def _amount_to_words(amount: int) -> str:
    """Return amount in Indian English words (e.g. 'Rupees Fifty Thousand Five Hundred Only')."""
    ones = [
        '', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
        'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
        'Seventeen', 'Eighteen', 'Nineteen',
    ]
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def _below_hundred(n):
        return ones[n] if n < 20 else tens[n // 10] + (' ' + ones[n % 10] if n % 10 else '')

    def _below_thousand(n):
        if n < 100:
            return _below_hundred(n)
        return ones[n // 100] + ' Hundred' + (' ' + _below_hundred(n % 100) if n % 100 else '')

    if amount == 0:
        return 'Rupees Zero Only'

    n = int(amount)
    parts = []
    for divisor, label in [(10_000_000, 'Crore'), (100_000, 'Lakh'), (1_000, 'Thousand')]:
        if n >= divisor:
            parts.append(_below_thousand(n // divisor) + ' ' + label)
            n %= divisor
    if n > 0:
        parts.append(_below_thousand(n))

    return 'Rupees ' + ' '.join(parts) + ' Only'


@login_required
def salary_slip_view(request, record_id):
    record = get_object_or_404(
        PayrollRecord.objects.select_related('payroll', 'payroll__company', 'employee'),
        id=record_id,
    )
    employee = record.employee
    company  = record.payroll.company

    context = {
        'record':       record,
        'run':          record.payroll,
        'employee':     employee,
        'company':      company,
        'net_in_words': _amount_to_words(int(record.net_salary)),
        'is_pdf':       request.GET.get('format') == 'pdf',
    }

    if context['is_pdf']:
        tpl  = get_template('payroll/salary_slip.html')
        html = tpl.render(context)
        buf  = BytesIO()
        if pisa.CreatePDF(html, dest=buf).err:
            return HttpResponse('PDF generation failed', status=500)
        fname = (
            f"SalarySlip_{employee.employee_code or employee.id}_"
            f"{record.payroll.month.strftime('%b_%Y')}.pdf"
        )
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="{fname}"'
        return resp

    return render(request, 'payroll/salary_slip.html', context)


@login_required
def download_empty_excel(request):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "attendance_file"

    # (Optional) Add headers only
    headers = [
        "Employee Code",
        "In Time",
        "Out Time",
        "AttendanceDate",
    ]

    ws.append(headers)  # comment this line if you want 100% empty

    # Prepare HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="attendance_file.xlsx"'

    wb.save(response)
    return response





from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)

            # ADMIN or HR → Admin Dashboard
            if user.groups.filter(name__in=["Admin", "HR","Manager"]).exists():
                return redirect("admin-dashboard")

            # EMPLOYEE → Employee Detail Page
            elif user.groups.filter(name="Employee").exists():
                try:
                    employee = Employee.objects.get(user=user)
                    return redirect("employee_detail", pk=employee.id)
                except Employee.DoesNotExist:
                    return redirect("admin-dashboard")

            # fallback for users with no group assigned
            return redirect("admin-dashboard")

        else:
            messages.error(request, "Invalid username or password")

    return render(request, "auth/login.html")


from django.contrib.auth import logout

def logout_view(request):
    logout(request)
    return redirect("login")



from django.contrib.auth.models import User, Group

@login_required
@group_required("Admin")
def create_user_view(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        role = request.POST.get("role")

        employee = Employee.objects.get(id=employee_id)

        if employee.user:
            messages.error(request, "User already exists for this employee.")
            return redirect("create-user")

        user = User.objects.create_user(
            username=employee.employee_code,
            email=employee.personal_email,
            password="Temp@123"
        )

        group = Group.objects.get(name=role)
        user.groups.add(group)

        employee.user = user
        employee.force_password_change = True   # 🔥 REQUIRED
        employee.save()

        messages.success(
            request,
            f"User created. Username: {user.username}, Password: Temp@123"
        )

        return redirect("create-user")

    employees = Employee.objects.filter(user__isnull=True)
    return render(request, "auth/create_user.html", {"employees": employees})



# views.py
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages

@login_required
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)

        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)

            employee = request.user.employee_profile
            employee.force_password_change = False
            employee.save()

            messages.success(request, "✅ Password updated successfully.")
            return redirect("admin-dashboard")

        else:
            # 🔴 IMPORTANT: show WHY it failed
            messages.error(request, "❌ Password update failed. Please fix the errors below.")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")

    else:
        form = PasswordChangeForm(request.user)

    return render(request, "auth/change_password.html", {"form": form})



# holiday calendar views


# website/views.py - COMPLETE HOLIDAY CALENDAR VIEWS

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from datetime import timedelta, date
import json

from .models import Holiday, HolidayType, MonthlyEarnedLeaves, PayrollSettings, Branch, Company
from .forms import HolidayForm, MonthlyEarnedLeavesForm

# ============================================================================
# HOLIDAY CALENDAR DASHBOARD VIEW
# ============================================================================

@login_required
@require_http_methods(["GET"])
def holiday_calendar_dashboard(request):
    """
    Main Holiday Calendar Dashboard View
    """

    # ----------------------------------------------------
    # 1️⃣ Year & Month handling
    # ----------------------------------------------------
    year = int(request.GET.get('year', timezone.now().year))
    month = int(request.GET.get('month', timezone.now().month))

    if not 1 <= month <= 12:
        month = timezone.now().month

    month_name = calendar.month_name[month]

    # ----------------------------------------------------
    # 2️⃣ Company & Payroll Settings resolution
    # ----------------------------------------------------
    _company = get_user_company(request.user)
    if _company is None and user_has_global_access(request.user):
        _cid = request.GET.get('company_id')
        if _cid:
            _company = Company.objects.filter(pk=_cid).first()
        else:
            _company = Company.objects.filter(payrollsettings__isnull=False).first()
    payroll_settings = PayrollSettings.objects.filter(company=_company).first() if _company else None

    # ----------------------------------------------------
    # 3️⃣ Auto-create Monthly Earned Leaves (CRITICAL FIX)
    # ----------------------------------------------------
    monthly_leaves = []

# In holiday_calendar_dashboard view, replace the monthly_leaves section:

    if payroll_settings:
        annual_leaves = payroll_settings.earned_leaves_per_year or 24
        monthly_default = (
            Decimal(str(annual_leaves)) / Decimal('12')
        ).quantize(Decimal('0.01'))
        
        fy_start = payroll_settings.financial_year_start_month or 4  # April default
        
        # Determine FY year range based on selected year
        # If FY starts April 2025 → months Apr2025 to Mar2026
        # Current year shown in dashboard determines which FY
        if month >= fy_start:
            fy_start_year = year
        else:
            fy_start_year = year - 1
        
        monthly_leaves = []
        for i in range(12):
            m = fy_start + i
            y = fy_start_year
            if m > 12:
                m -= 12
                y += 1
            
            leave, _ = MonthlyEarnedLeaves.objects.get_or_create(
                payroll_settings=payroll_settings,
                month=m,
                year=y,
                defaults={
                    'earned_leaves': monthly_default,
                    'is_auto_generated': True,
                }
            )
            monthly_leaves.append(leave)
        
        # Sort by year then month
        monthly_leaves = sorted(monthly_leaves, key=lambda x: (x.year, x.month))
    # ----------------------------------------------------
    # 4️⃣ Holidays for calendar (month-based)
    # ----------------------------------------------------
    all_holidays_month = Holiday.objects.filter(
        holiday_date__year=year,
        holiday_date__month=month
    ).order_by('holiday_date')

    # ----------------------------------------------------
    # 5️⃣ Holidays JSON for calendar UI
    # ----------------------------------------------------
    holidays_json = json.dumps([
        {
            'id': h.id,
            'holiday_date': h.holiday_date.isoformat(),
            'name': h.name,
            'status': h.status,
            'is_national': h.is_national,
        }
        for h in Holiday.objects.all()
    ])

    # ----------------------------------------------------
    # 6️⃣ Upcoming Holidays (next 7 days)
    # ----------------------------------------------------
    today = date.today()
    upcoming_holidays = Holiday.objects.filter(
        holiday_date__gte=today,
        holiday_date__lte=today + timedelta(days=7)
    ).order_by('holiday_date')[:10]

    # ----------------------------------------------------
    # 7️⃣ Statistics (year-based)
    # ----------------------------------------------------
    all_holidays_year = Holiday.objects.filter(holiday_date__year=year)

    total_holidays = all_holidays_year.count()
    national_holidays = all_holidays_year.filter(is_national=True).count()
    regional_holidays = all_holidays_year.filter(is_national=False).count()
    emergency_closures = all_holidays_year.filter(status='emergency').count()

    # ----------------------------------------------------
    # 8️⃣ Holiday Types (dropdowns)
    # ----------------------------------------------------
    holiday_types = HolidayType.objects.all()
    branches = Branch.objects.all()

    all_half_day_scenarios = HalfDayScenario.objects.all().order_by('scenario_date')
    
    half_day_json = json.dumps([
        {
            'id': s.id,
            'scenario_date': s.scenario_date.isoformat(),
            'description': s.description,    # ✅ correct field name
            'scenario_type': s.scenario_type,
            'is_approved': s.is_approved,
            'branch': s.branch.branch_name if s.branch else 'All',
            'credit_count': str(s.credit_count),
        }
        for s in all_half_day_scenarios
    ])

    # ----------------------------------------------------
    # 9️⃣ FINAL CONTEXT (⚠️ NOTHING REMOVED)
    # ----------------------------------------------------
    context = {
        'year': year,
        'month': month,
        'month_name': month_name,

        # Holidays
        'all_holidays': Holiday.objects.all().order_by('holiday_date'),
        'upcoming_holidays': upcoming_holidays,
        'holidays_json': holidays_json,

        # Stats
        'total_holidays': total_holidays,
        'national_holidays': national_holidays,
        'regional_holidays': regional_holidays,
        'emergency_closures': emergency_closures,

        # Earned Leaves
        'monthly_leaves': monthly_leaves,
        'payroll_settings': payroll_settings,

        # Dropdowns
        'holiday_types': holiday_types,

        'branches': branches,
        'all_half_day_scenarios': all_half_day_scenarios,
        'half_day_json': half_day_json,
        'total_half_days': all_half_day_scenarios.filter(
            scenario_date__year=year, is_approved=True
        ).count(),

    }

    return render(request, 'holiday_calendar/dashboard.html', context)







@login_required
@require_http_methods(["POST"])
def save_holiday_settings(request):
    try:
        company_id_param = request.POST.get('company_id')
        if user_has_global_access(request.user) and company_id_param:
            company = Company.objects.filter(pk=company_id_param).first()
        else:
            company = get_user_company(request.user)
        if not company:
            messages.error(request, 'No company found.')
            return redirect('holiday-calendar')
        settings, _ = PayrollSettings.objects.get_or_create(company=company)
        
        financial_year_start = request.POST.get('financial_year_start_month')
        branch_specific_holidays = request.POST.get('branch_specific_holidays') == 'true'
        
        if financial_year_start:
            fy_month = int(financial_year_start)
            if 1 <= fy_month <= 12:
                old_fy = settings.financial_year_start_month
                settings.financial_year_start_month = fy_month
                
                # If FY start month changed, regenerate earned leaves
                if old_fy != fy_month:
                    # Delete old auto-generated leaves for future years
                    MonthlyEarnedLeaves.objects.filter(
                        payroll_settings=settings,
                        is_auto_generated=True
                    ).delete()
                    # Regenerate for current FY
                    MonthlyEarnedLeaves.generate_for_payroll_settings(settings)
                    messages.success(
                        request,
                        f'✓ Financial year updated to start from month {fy_month}. '
                        f'Earned leaves regenerated.'
                    )
        
        settings.branch_specific_holidays = branch_specific_holidays
        settings.save()
        
        messages.success(request, '✓ Holiday settings saved successfully!')
        
    except Exception as e:
        messages.error(request, f'❌ Error saving settings: {str(e)}')
    
    return redirect('holiday-calendar')


# ============================================================================
# ADD HOLIDAY VIEW
# ============================================================================

@login_required
@require_http_methods(["GET", "POST"])
def add_holiday(request):
    """
    Add New Holiday View
    
    GET: Shows form (but in modal, so redirects)
    POST: Creates new Holiday object
    
    Form Fields:
    - holiday_date (DateField): Date of holiday
    - name (CharField): Name of holiday
    - holiday_type (ForeignKey): Type of holiday
    - status (CharField): Status (declared/optional/emergency)
    - is_national (BooleanField): Apply to all branches?
    - description (TextField): Optional description
    
    Redirects back to dashboard with success/error message
    """
    
    if request.method == 'POST':
        # Instantiate form with POST data
        form = HolidayForm(request.POST)
        
        if form.is_valid():
            holiday = form.save(commit=False)
            holiday.created_by = request.user
            # Auto-flag national holidays
            if holiday.holiday_type and holiday.holiday_type.type_category == "national":
                holiday.is_national = True
            holiday.save()

            # Religious/optional holidays scoped to specific employees only
            # (e.g. Eid for Muslim employees) — not part of the ModelForm
            # since it's driven by a searchable picker, not a plain field.
            if not holiday.applies_to_all_employees:
                employee_ids = request.POST.getlist('specific_employees')
                holiday.specific_employees.set(employee_ids)
            else:
                holiday.specific_employees.clear()

            # Auto-generate 'Holiday' attendance for anyone who doesn't
            # already have a record on this date — covers the case where
            # nobody uploads attendance for a day everyone knows is a holiday.
            backfilled = holiday.backfill_attendance()

            # Show success message
            success_msg = (
                f'✓ Holiday "{holiday.name}" ({holiday.holiday_date.strftime("%b %d, %Y")}) '
                f'added successfully!'
            )
            if backfilled:
                success_msg += f' Auto-generated attendance for {backfilled} employee(s).'
            messages.success(request, success_msg)

            return redirect('holiday-calendar')
        else:
            # Form has errors - show them
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field.title()}: {error}')
            
            return redirect('holiday-calendar')
    
    # GET request - show form (in modal, so just redirect)
    return redirect('holiday-calendar')

# ============================================================================
# EDIT HOLIDAY VIEW
# ============================================================================

@login_required
@require_http_methods(["GET", "POST"])
def edit_holiday(request, holiday_id):
    """
    Edit Existing Holiday View
    
    GET: Loads holiday data via AJAX (see api_get_holiday)
    POST: Updates existing Holiday object
    
    Parameters:
    - holiday_id: ID of holiday to edit
    
    Redirects back to dashboard with success/error message
    """
    
    # Get the holiday or 404
    holiday = get_object_or_404(Holiday, id=holiday_id)
    
    if request.method == 'POST':
        # Instantiate form with existing instance
        form = HolidayForm(request.POST, instance=holiday)
        
        if form.is_valid():
            # Save updates
            holiday = form.save()

            if not holiday.applies_to_all_employees:
                employee_ids = request.POST.getlist('specific_employees')
                holiday.specific_employees.set(employee_ids)
            else:
                holiday.specific_employees.clear()

            backfilled = holiday.backfill_attendance()

            # Show success message
            success_msg = f'✓ Holiday "{holiday.name}" updated successfully!'
            if backfilled:
                success_msg += f' Auto-generated attendance for {backfilled} employee(s).'
            messages.success(request, success_msg)

            return redirect('holiday-calendar')
        else:
            # Form has errors
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field.title()}: {error}')
    
    return redirect('holiday-calendar')


# ============================================================================
# DELETE HOLIDAY VIEW
# ============================================================================

@login_required
@require_http_methods(["POST"])
def delete_holiday(request, holiday_id):
    """
    Delete Holiday View
    
    POST only (with confirmation from frontend)
    
    Parameters:
    - holiday_id: ID of holiday to delete
    
    Deletes the holiday and redirects with message
    """
    
    # Get the holiday
    holiday = get_object_or_404(Holiday, id=holiday_id)
    holiday_name = holiday.name
    holiday_date = holiday.holiday_date.strftime("%b %d, %Y")
    
    # Delete it
    holiday.delete()
    
    # Show success message
    messages.success(
        request,
        f'✓ Holiday "{holiday_name}" ({holiday_date}) deleted successfully!'
    )
    
    return redirect('holiday-calendar')


# ============================================================================
# HOLIDAY LIST VIEW
# ============================================================================

@login_required
@require_http_methods(["GET"])
def holiday_list(request):
    """
    Holiday List View (List Tab)
    
    Shows all holidays in table format
    Can filter by type, status, branch
    
    Query Parameters:
    - type: Filter by holiday type
    - status: Filter by status
    - branch: Filter by branch
    
    Returns: Redirects to dashboard (list is shown in tab)
    """
    
    return redirect('holiday-calendar')


# ============================================================================
# EARNED LEAVES CONFIGURATION VIEW
# ============================================================================

@login_required
@require_http_methods(["GET"])
def earned_leaves_config(request):
    """
    Display Monthly Earned Leaves for a selected financial year.
    Auto-generation is handled via signals.
    """

    try:
        year = int(request.GET.get('year', timezone.now().year))

        _company = get_user_company(request.user)
        if _company is None and user_has_global_access(request.user):
            _cid = request.GET.get('company_id')
            _company = Company.objects.filter(pk=_cid).first() if _cid else Company.objects.filter(payrollsettings__isnull=False).first()
        payroll_settings = PayrollSettings.objects.filter(company=_company).first() if _company else None

        if not payroll_settings:
            messages.error(request, "❌ Payroll Settings not configured!")
            return render(request, 'holiday_calendar/dashboard.html', {
                'monthly_leaves': [],
                'year': year,
                'error': 'No payroll settings'
            })

        # ✅ JUST FETCH – do NOT create here
        monthly_leaves = MonthlyEarnedLeaves.objects.filter(
            payroll_settings=payroll_settings,
            year=year
        ).order_by('month')

        # Safety check (optional but good)
        if not monthly_leaves.exists():
            messages.warning(
                request,
                "⚠️ Earned leaves not found for this year. Please check Payroll Settings."
            )

        # Statistics
        total_earned_leaves = sum(
            float(leave.earned_leaves) for leave in monthly_leaves
        )
        auto_count = monthly_leaves.filter(is_auto_generated=True).count()
        manual_count = monthly_leaves.count() - auto_count

        context = {
            'monthly_leaves': monthly_leaves,
            'year': year,
            'payroll_settings': payroll_settings,
            'total_earned_leaves': round(total_earned_leaves, 2),
            'auto_generated_count': auto_count,
            'manual_count': manual_count,
        }

        return render(request, 'holiday_calendar/dashboard.html', context)

    except Exception as e:
        messages.error(request, f"❌ Error loading earned leaves: {str(e)}")
        return render(request, 'holiday_calendar/dashboard.html', {
            'monthly_leaves': [],
            'year': timezone.now().year,
        })


@login_required
@require_http_methods(["POST"])
def edit_earned_leave(request, leave_id):
    """
    ✅ Edit Monthly Earned Leave Value
    
    Updates the earned_leaves field for a specific month.
    
    Example:
    POST /holiday/earned-leave/1/edit/
    Data: {
        'earned_leaves': '3.5',
        'is_auto_generated': 'on'
    }
    """
    
    try:
        # Get the leave record
        leave = MonthlyEarnedLeaves.objects.get(id=leave_id)
        
        # Get form data
        earned_leaves_value = request.POST.get('earned_leaves', '').strip()
        is_auto_generated = request.POST.get('is_auto_generated', 'off') == 'on'
        
        # Validate input
        if not earned_leaves_value:
            messages.error(request, '❌ Earned leaves value is required!')
            return redirect('holiday-calendar')
        
        # Try to convert to Decimal
        try:
            earned_leaves_decimal = Decimal(earned_leaves_value)
        except:
            messages.error(
                request,
                f'❌ Invalid value "{earned_leaves_value}". Use decimal numbers like 2.5 or 3.75'
            )
            return redirect('holiday-calendar')
        
        # Validate range (0-31 days)
        if earned_leaves_decimal < 0 or earned_leaves_decimal > 31:
            messages.error(request, '❌ Earned leaves must be between 0 and 31 days!')
            return redirect('holiday-calendar')
        
        # Update the record
        leave.earned_leaves = earned_leaves_decimal
        leave.is_auto_generated = is_auto_generated
        leave.save()
        
        # Get month name
        month_names = {
            1: 'January', 2: 'February', 3: 'March', 4: 'April',
            5: 'May', 6: 'June', 7: 'July', 8: 'August',
            9: 'September', 10: 'October', 11: 'November', 12: 'December'
        }
        month_name = month_names.get(leave.month, 'Month')
        status_text = "Manual" if not is_auto_generated else "Auto"
        
        messages.success(
            request,
            f'✓ {month_name} {leave.year}: Updated to {earned_leaves_decimal} days ({status_text})'
        )
    
    except MonthlyEarnedLeaves.DoesNotExist:
        messages.error(request, '❌ Earned leave record not found!')
    except ValueError as e:
        messages.error(request, f'❌ Invalid input: {str(e)}')
    except Exception as e:
        messages.error(request, f'❌ Error updating earned leaves: {str(e)}')
    
    return redirect('holiday-calendar')


@login_required
@require_http_methods(["GET"])
def api_earned_leaves(request):
    """
    ✅ API Endpoint - Get Earned Leaves as JSON
    
    Used by JavaScript to fetch earned leaves data.
    
    Example:
    GET /api/earned-leaves/?year=2025
    
    Response:
    [
        {
            'id': 1,
            'month': 1,
            'month_name': 'January',
            'year': 2025,
            'earned_leaves': '3.50',
            'is_auto_generated': False
        },
        ...
    ]
    """
    
    try:
        year = request.GET.get('year', timezone.now().year)
        _company = get_user_company(request.user)
        if _company is None and user_has_global_access(request.user):
            _cid = request.GET.get('company_id')
            _company = Company.objects.filter(pk=_cid).first() if _cid else Company.objects.filter(payrollsettings__isnull=False).first()
        payroll_settings = PayrollSettings.objects.filter(company=_company).first() if _company else None

        if not payroll_settings:
            return JsonResponse({'error': 'No payroll settings'}, status=400)

        leaves = MonthlyEarnedLeaves.objects.filter(
            payroll_settings=payroll_settings,
            year=year
        ).order_by('month')
        
        month_names = {
            1: 'January', 2: 'February', 3: 'March', 4: 'April',
            5: 'May', 6: 'June', 7: 'July', 8: 'August',
            9: 'September', 10: 'October', 11: 'November', 12: 'December'
        }
        
        data = [
            {
                'id': leave.id,
                'month': leave.month,
                'month_name': month_names.get(leave.month, 'Unknown'),
                'year': leave.year,
                'earned_leaves': str(leave.earned_leaves),
                'is_auto_generated': leave.is_auto_generated,
            }
            for leave in leaves
        ]
        
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def add_half_day_scenario(request):
    if request.method == 'POST':
        scenario_date = request.POST.get('scenario_date')
        description = request.POST.get('description', '')
        scenario_type = request.POST.get('scenario_type', 'other')
        is_approved = request.POST.get('is_approved') == 'on'
        branch_id = request.POST.get('branch')
        credit_count = request.POST.get('credit_count', '0.50')

        if not scenario_date or not description:
            messages.error(request, '❌ Date and description are required.')
            return redirect('holiday-calendar')

        try:
            from datetime import date as date_cls
            parsed_date = date_cls.fromisoformat(scenario_date)

            # branch_id == 'all' or empty → applies to all branches (branch=None)
            if branch_id == 'all' or not branch_id:
                branch = None
                if HalfDayScenario.objects.filter(scenario_date=parsed_date, branch__isnull=True).exists():
                    messages.error(request, f'❌ An all-branches half-day scenario already exists for {parsed_date}.')
                    return redirect('holiday-calendar')
                branch_label = 'All Branches'
            else:
                branch = get_object_or_404(Branch, id=branch_id)
                if HalfDayScenario.objects.filter(scenario_date=parsed_date, branch=branch).exists():
                    messages.error(
                        request,
                        f'❌ A half-day scenario already exists for '
                        f'{branch.branch_name} on {parsed_date}.'
                    )
                    return redirect('holiday-calendar')
                branch_label = branch.branch_name

            HalfDayScenario.objects.create(
                scenario_date=parsed_date,
                description=description,
                scenario_type=scenario_type,
                is_approved=is_approved,
                branch=branch,
                credit_count=Decimal(credit_count),
                created_by=request.user,
            )

            messages.success(
                request,
                f'✓ Half-day scenario added for {branch_label} on {parsed_date.strftime("%b %d, %Y")}!'
            )
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')

    return redirect('holiday-calendar')


@login_required
@require_http_methods(["POST"])
def edit_half_day_scenario(request, scenario_id):
    scenario = get_object_or_404(HalfDayScenario, id=scenario_id)

    try:
        from datetime import date as date_cls
        scenario_date = request.POST.get('scenario_date')
        description = request.POST.get('description', '')
        scenario_type = request.POST.get('scenario_type', 'other')
        is_approved = request.POST.get('is_approved') == 'on'
        branch_id = request.POST.get('branch')
        credit_count = request.POST.get('credit_count', '0.50')

        scenario.scenario_date = date_cls.fromisoformat(scenario_date)
        scenario.description = description
        scenario.scenario_type = scenario_type
        scenario.is_approved = is_approved
        scenario.branch = None if (branch_id == 'all' or not branch_id) else get_object_or_404(Branch, id=branch_id)
        scenario.credit_count = Decimal(credit_count)

        # Set approved_by if being approved
        if is_approved and not scenario.approved_by:
            scenario.approved_by = request.user

        scenario.save()
        messages.success(request, '✓ Half-day scenario updated!')
    except Exception as e:
        messages.error(request, f'❌ Error: {str(e)}')

    return redirect('holiday-calendar')


@login_required
@require_http_methods(["POST"])
def delete_half_day_scenario(request, scenario_id):
    scenario = get_object_or_404(HalfDayScenario, id=scenario_id)
    branch_name = scenario.branch.branch_name if scenario.branch else 'All Branches'
    date_str = scenario.scenario_date.strftime("%b %d, %Y")
    scenario.delete()
    messages.success(request, f'✓ Half-day scenario for {branch_name} on {date_str} deleted!')
    return redirect('holiday-calendar')


@login_required
@require_http_methods(["GET"])
def api_get_half_day_scenario(request, scenario_id):
    try:
        s = HalfDayScenario.objects.get(id=scenario_id)
        return JsonResponse({
            'id': s.id,
            'scenario_date': s.scenario_date.isoformat(),
            'description': s.description,
            'scenario_type': s.scenario_type,
            'is_approved': s.is_approved,
            'branch': s.branch.id if s.branch else None,
            'credit_count': str(s.credit_count),
        })
    except HalfDayScenario.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)
# ============================================================================
# API ENDPOINTS - JSON RESPONSES FOR AJAX
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_get_holiday(request, holiday_id):
    """
    API: Get Single Holiday Data
    
    Used by AJAX to populate edit modal
    
    Returns: JSON object with holiday data
    
    Example Response:
    {
        'id': 1,
        'holiday_date': '2025-01-26',
        'name': 'Republic Day',
        'holiday_type': 1,
        'status': 'declared',
        'description': 'National holiday',
        'is_national': true
    }
    """
    
    try:
        # Get the holiday
        holiday = Holiday.objects.get(id=holiday_id)
        
        # Build response data
        data = {
            'id': holiday.id,
            'holiday_date': holiday.holiday_date.isoformat(),
            'name': holiday.name,
            'holiday_type': holiday.holiday_type.id,
            'status': holiday.status,
            'description': holiday.description or '',
            'is_national': holiday.is_national,
            'applies_to_all_employees': holiday.applies_to_all_employees,
            'specific_employees': [
                {
                    'id': emp.id,
                    'name': f'{emp.first_name} {emp.last_name}',
                    'code': emp.employee_code,
                }
                for emp in holiday.specific_employees.all()
            ],
        }

        return JsonResponse(data)
    
    except Holiday.DoesNotExist:
        # Holiday not found
        return JsonResponse({'error': 'Holiday not found'}, status=404)
    except Exception as e:
        # Other error
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_holidays_json(request):
    """
    API: Get All Holidays
    
    Returns JSON array of all holidays
    Used to populate holiday list table
    
    Query Parameters:
    - type: Filter by holiday type
    - status: Filter by status
    - year: Filter by year
    
    Example Response:
    [
        {
            'id': 1,
            'holiday_date': '2025-01-26',
            'name': 'Republic Day',
            'holiday_type': 'National',
            'status': 'declared',
            'is_national': true
        },
        ...
    ]
    """
    
    # Start with all holidays
    holidays = Holiday.objects.all().order_by('holiday_date')
    
    # Apply filters if provided
    holiday_type = request.GET.get('type')
    if holiday_type:
        holidays = holidays.filter(holiday_type__type_category=holiday_type)
    
    status = request.GET.get('status')
    if status:
        holidays = holidays.filter(status=status)
    
    year = request.GET.get('year')
    if year:
        holidays = holidays.filter(holiday_date__year=year)
    
    # Build response data
    data = [
        {
            'id': h.id,
            'holiday_date': h.holiday_date.isoformat(),
            'name': h.name,
            'holiday_type': h.holiday_type.name,
            'status': h.status,
            'is_national': h.is_national,
        }
        for h in holidays
    ]
    
    return JsonResponse(data, safe=False)


@login_required
@require_http_methods(["GET"])
def api_earned_leaves_json(request):
    """
    API: Get Earned Leaves Data
    
    Returns monthly earned leaves for calendar calculations
    
    Query Parameters:
    - year: Filter by year (default: current year)
    
    Example Response:
    [
        {
            'id': 1,
            'month': 1,
            'year': 2025,
            'earned_leaves': '2.0',
            'is_auto_generated': true
        },
        ...
    ]
    """
    
    # Get year from query params or use current
    year = request.GET.get('year', timezone.now().year)
    
    # Get earned leaves
    leaves = MonthlyEarnedLeaves.objects.filter(year=year).order_by('month')
    
    # Build response data
    data = [
        {
            'id': l.id,
            'month': l.month,
            'year': l.year,
            'earned_leaves': str(l.earned_leaves),
            'is_auto_generated': l.is_auto_generated,
        }
        for l in leaves
    ]
    
    return JsonResponse(data, safe=False)


# ============================================================================
# HOLIDAY TYPE MANAGEMENT
# ============================================================================

@login_required
@group_required("Admin", "HR")
@require_POST
def create_holiday_type(request):
    name = request.POST.get("name", "").strip()
    type_category = request.POST.get("type_category", "national")
    description = request.POST.get("description", "").strip()
    color_code = request.POST.get("color_code", "#2196F3").strip() or "#2196F3"

    if not name:
        messages.error(request, "Holiday type name is required.")
        return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")

    if HolidayType.objects.filter(name__iexact=name).exists():
        messages.error(request, f"A holiday type named '{name}' already exists.")
        return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")

    ht = HolidayType.objects.create(
        name=name,
        type_category=type_category,
        description=description,
        color_code=color_code,
    )
    messages.success(request, f"Holiday type '{ht.name}' created successfully.")
    return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")


@login_required
@group_required("Admin", "HR")
@require_POST
def edit_holiday_type(request, type_id):
    ht = get_object_or_404(HolidayType, id=type_id)
    name = request.POST.get("name", "").strip()
    type_category = request.POST.get("type_category", ht.type_category)
    description = request.POST.get("description", "").strip()
    color_code = request.POST.get("color_code", ht.color_code).strip() or "#2196F3"

    if not name:
        messages.error(request, "Holiday type name is required.")
        return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")

    if HolidayType.objects.filter(name__iexact=name).exclude(id=type_id).exists():
        messages.error(request, f"A holiday type named '{name}' already exists.")
        return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")

    ht.name = name
    ht.type_category = type_category
    ht.description = description
    ht.color_code = color_code
    ht.save()
    messages.success(request, f"Holiday type '{ht.name}' updated successfully.")
    return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")


@login_required
@group_required("Admin", "HR")
@require_POST
def delete_holiday_type(request, type_id):
    ht = get_object_or_404(HolidayType, id=type_id)
    if Holiday.objects.filter(holiday_type=ht).exists():
        messages.error(request, f"Cannot delete '{ht.name}': it is used by existing holidays.")
        return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")
    name = ht.name
    ht.delete()
    messages.success(request, f"Holiday type '{name}' deleted.")
    return redirect(f"{reverse('holiday-dashboard')}?tab=holiday-types")


@login_required
def api_get_holiday_type(request, type_id):
    ht = get_object_or_404(HolidayType, id=type_id)
    return JsonResponse({
        "id": ht.id,
        "name": ht.name,
        "type_category": ht.type_category,
        "description": ht.description,
        "color_code": ht.color_code,
    })


# ============================================================================
# UTILITY VIEWS (Optional but helpful)
# ============================================================================

@login_required
@require_http_methods(["GET"])
def api_holiday_types(request):
    """
    API: Get Holiday Types
    
    Returns all available holiday types
    Used to populate type filter and form dropdowns
    """
    
    types = HolidayType.objects.all()
    
    data = [
        {
            'id': t.id,
            'name': t.name,
            'type_category': t.type_category,
        }
        for t in types
    ]
    
    return JsonResponse(data, safe=False)


@login_required
@require_http_methods(["GET"])
def api_payroll_settings(request):
    """
    API: Get Payroll Settings
    
    Returns company-wide holiday and leave settings
    """
    
    try:
        _company = get_user_company(request.user)
        if _company is None and user_has_global_access(request.user):
            _cid = request.GET.get('company_id')
            _company = Company.objects.filter(pk=_cid).first() if _cid else Company.objects.filter(payrollsettings__isnull=False).first()
        settings = PayrollSettings.objects.filter(company=_company).first() if _company else None

        if not settings:
            return JsonResponse({'error': 'Settings not configured'}, status=404)

        data = {
            'id': settings.id,
            'earned_leaves_per_year': str(settings.earned_leaves_per_year),
            'financial_year_start_month': settings.financial_year_start_month,
        }
        
        return JsonResponse(data)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============================================================================
# HELPER FUNCTIONS (Not views, but useful)
# ============================================================================

def get_holidays_for_month(year, month):
    """
    Helper: Get all holidays for a specific month
    
    Parameters:
    - year: Year
    - month: Month (1-12)
    
    Returns: QuerySet of Holiday objects
    """
    
    return Holiday.objects.filter(
        holiday_date__year=year,
        holiday_date__month=month
    ).order_by('holiday_date')


def get_upcoming_holidays(days=7):
    """
    Helper: Get upcoming holidays
    
    Parameters:
    - days: Number of days to look ahead (default: 7)
    
    Returns: QuerySet of Holiday objects
    """
    
    today = date.today()
    end_date = today + timedelta(days=days)
    
    return Holiday.objects.filter(
        holiday_date__gte=today,
        holiday_date__lte=end_date
    ).order_by('holiday_date')


def calculate_working_days(start_date, end_date, payroll_settings=None):
    """
    Helper: Calculate working days (excluding holidays and weekends)

    Parameters:
    - start_date: Start date
    - end_date: End date
    - payroll_settings: PayrollSettings instance (used for weekend_days config)

    Returns: Number of working days
    """

    working_days = 0
    current = start_date

    sunday_only = payroll_settings is not None and getattr(payroll_settings, 'weekend_days', 'sat_sun') == 'sun'

    # Get all holidays in date range
    holidays = Holiday.objects.filter(
        holiday_date__gte=start_date,
        holiday_date__lte=end_date
    )
    holiday_dates = set(h.holiday_date for h in holidays)

    # Count working days
    while current <= end_date:
        # Skip weekends based on setting (Python weekday: 5=Saturday, 6=Sunday)
        is_weekend = (current.weekday() == 6) if sunday_only else (current.weekday() >= 5)
        if not is_weekend:
            # Skip holidays
            if current not in holiday_dates:
                working_days += 1
        
        current += timedelta(days=1)


# ─── Employee Excel Import ────────────────────────────────────────────────────

import re
import pandas as pd
from decimal import Decimal
from datetime import date as date_cls, timedelta

from django.http import JsonResponse
from django.db import transaction
from django.contrib.auth.models import User

# ─── Helpers ──────────────────────────────────────────────────────────────────

def normalize_blood_group(val):
    """
    Convert Excel variants → model choices.
    'B (+ve)', 'A (+ ve)', 'O (- ve)', 'AB (+ ve)' → 'B+', 'A+', 'O-', 'AB+'
    Already-correct formats like 'B+', 'AB-' are passed through unchanged.
    """
    if not val or str(val).strip().lower() in ("", "nan", "na", "n/a"):
        return ""
    v = str(val).strip().upper()
    m = re.match(r'^(AB|A|B|O)\s*[\(\[]?\s*([+\-])\s*V?E?\s*[\)\]]?', v)
    if m:
        return m.group(1) + m.group(2)
    # already in correct format (A+, B-, AB+, O-)
    m2 = re.match(r'^(AB|A|B|O)[+\-]$', v)
    if m2:
        return m2.group(0)
    return ""   # unrecognised → blank (field is nullable)


def normalize_relation(val):
    """
    Fix typos/variants in emergency contact relation so they match model choices.
    Model choices: Spouse, Father, Mother, Brother, Sister, Son, Daughter, Other
    """
    if not val or str(val).strip().lower() in ("", "nan", "0"):
        return ""
    mapping = {
        "spouse": "Spouse",
        "wife": "Spouse",
        "husband": "Spouse",
        "father": "Father",
        "fathar": "Father",
        "dad": "Father",
        "mother": "Mother",
        "mather": "Mother",        # typo in actual data
        "mom": "Mother",
        "brother": "Brother",
        "bro": "Brother",
        "sister": "Sister",
        "sis": "Sister",
        "son": "Son",
        "daughter": "Daughter",
        "other": "Other",
    }
    return mapping.get(str(val).strip().lower(), "Other")


def parse_excel_date(val):
    """
    Robustly parse a date value that may be:
    - Already a datetime / date object (pandas already parsed it)
    - A string in YYYY-MM-DD, DD-MM-YYYY, DD/MM/YYYY etc.
    - An Excel serial number integer stored as string e.g. "31664"
    Returns a date object or None.
    """
    if val is None:
        return None
    # guard against pandas NaT (NaTType.date() raises an error)
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    # already a date/datetime from pandas
    if hasattr(val, "date"):
        try:
            return val.date()
        except Exception:
            return None
    if isinstance(val, date_cls):
        return val

    s = str(val).strip()
    if s in ("", "nan", "NaT", "0", "None"):
        return None

    # try standard pandas parsing first (handles most string formats)
    try:
        parsed = pd.to_datetime(s, dayfirst=False, errors="coerce")
        if parsed is not None and not pd.isna(parsed):
            return parsed.date()
    except Exception:
        pass

    # try as Excel serial number
    try:
        serial = int(float(s))
        if 1000 < serial < 60000:          # sanity range for real dates
            return (date_cls(1899, 12, 30) + timedelta(days=serial))
    except Exception:
        pass

    return None


def safe_str(val):
    """Return clean string or empty string for NaN/None/0."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "nat", "none") else s


# ─── Template download ────────────────────────────────────────────────────────

@login_required
@group_required("Admin", "HR")
def download_employee_import_template(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = [
        "Employee Code*", "Salutation*", "First Name*", "Middle Name", "Last Name*",
        "Father Name*", "Gender* (Male/Female)", "Blood Group*", "Date of Birth* (YYYY-MM-DD)",
        "Place of Birth*", "Personal Email*", "Personal Mobile*",
        "Present Address*", "Permanent Address*", "Date of Marriage (YYYY-MM-DD)",
        "Company Name*", "Branch Name*", "Designation*", "Department*",
        "Date of Joining* (YYYY-MM-DD)", "Date of Confirmation (YYYY-MM-DD)",
        "Location*", "On Payroll Of",
        "Shift Start Time (HH:MM)", "Shift End Time (HH:MM)",
        "PAN No*", "Aadhar No*", "Voter ID", "Passport", "UAN No", "PF No", "ESIC No",
        "Name As Per Bank*", "Salary Account Number*", "IFSC Code*",
        "Emergency Contact Name 1*", "Emergency Contact Relation 1*", "Emergency Contact Mobile 1*",
        "Emergency Contact Name 2", "Emergency Contact Relation 2", "Emergency Contact Mobile 2",
        "Status (Active/Pending/Left)",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employee_import_template.xlsx"'
    wb.save(response)
    return response


# ─── Main import view ─────────────────────────────────────────────────────────

@login_required
@group_required("Admin", "HR")
def import_employees_excel(request):
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "Not authenticated. Please log in."}, status=401)
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST required"}, status=405)

    try:
        return _do_import_employees(request)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[IMPORT FATAL ERROR]\n{tb}")
        return JsonResponse({
            "success": False,
            "error": f"Server error: {str(e)}",
        })


def _do_import_employees(request):
    from .models import Employee, Company, Branch   # adjust import path if needed

    print("[IMPORT] Request received")

    uploaded_file = request.FILES.get("employee_file")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "No file uploaded."})

    print(f"[IMPORT] Reading file: {uploaded_file.name}, size: {uploaded_file.size} bytes")

    # ── 1. Read Excel ─────────────────────────────────────────────────────────
    # Do NOT use dtype=str — let pandas parse dates natively.
    # We convert to strings ourselves field-by-field below.
    try:
        df = pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception as e:
        print(f"[IMPORT] File read error: {e}")
        return JsonResponse({"success": False, "error": f"Cannot read file: {e}"})

    print(f"[IMPORT] Rows: {len(df)}, Columns: {list(df.columns)}")
    df.columns = df.columns.str.strip()

    # ── 2. Column → field mapping ─────────────────────────────────────────────
    # "_skip" prefix = read but ignore (column exists in template but not in model)
    col_map = {
        "Employee Code*":                          "employee_code",
        "Salutation*":                             "salutation",
        "First Name*":                             "first_name",
        "Middle Name":                             "middle_name",
        "Last Name*":                              "last_name",
        "Father Name*":                            "father_name",
        "Gender* (Male/Female)":                   "gender",
        "Blood Group*":                            "blood_group",
        "Date of Birth* (YYYY-MM-DD)":             "date_of_birth",
        "Place of Birth*":                         "place_of_birth",
        "Personal Email*":                         "personal_email",
        "Personal Mobile*":                        "personal_mobile",
        "Present Address*":                        "present_address",
        "Permanent Address*":                      "permanent_address",
        "Date of Marriage (YYYY-MM-DD)":           "date_of_marriage",
        "Company Name*":                           "_company_name",
        "Branch Name*":                            "_branch_name",
        "Designation*":                            "designation",
        "Department*":                             "department",
        "Date of Joining* (YYYY-MM-DD)":           "date_of_joining",
        "Date of Confirmation (YYYY-MM-DD)":       "date_of_confirmation",
        "Location*":                               "location",
        "On Payroll Of":                           "_skip_on_payroll",   # not in model
        "Shift Start Time (HH:MM)":                "shift_start_time",
        "Shift End Time (HH:MM)":                  "shift_end_time",
        "PAN No*":                                 "pan_no",
        "Aadhar No*":                              "aadhar_no",
        "Voter ID":                                "voter_id",
        "Passport":                                "passport",
        "UAN No":                                  "uan_no",
        "PF No":                                   "pf_no",
        "ESIC No":                                 "esic_no",
        "Name As Per Bank*":                       "name_as_per_bank",
        "Salary Account Number*":                  "salary_account_number",
        "IFSC Code*":                              "ifsc_code",
        "Emergency Contact Name 1*":               "emergency_contact_name1",
        "Emergency Contact Relation 1*":           "emergency_contact_relation1",
        "Emergency Contact Mobile 1*":             "emergency_contact_mobile1",
        "Emergency Contact Name 2":                "emergency_contact_name2",
        "Emergency Contact Relation 2":            "emergency_contact_relation2",
        "Emergency Contact Mobile 2":              "emergency_contact_mobile2",
        "Status (Active/Pending/Left)":            "status",
    }

    date_fields  = {"date_of_birth", "date_of_joining", "date_of_confirmation", "date_of_marriage"}
    time_fields  = {"shift_start_time", "shift_end_time"}

    # Fields that are truly optional — missing value is fine
    optional_fields = {
        "middle_name", "father_name", "blood_group", "place_of_birth",
        "personal_email", "personal_mobile", "present_address", "permanent_address",
        "date_of_marriage", "date_of_confirmation", "location",
        "shift_start_time", "shift_end_time", "designation", "department",
        "voter_id", "passport", "uan_no", "pf_no", "esic_no",
        "name_as_per_bank", "salary_account_number", "ifsc_code",
        "emergency_contact_name1", "emergency_contact_relation1", "emergency_contact_mobile1",
        "emergency_contact_name2", "emergency_contact_relation2", "emergency_contact_mobile2",
        "status", "salutation",
    }

    created_count = 0
    skipped_count = 0
    errors        = []

    # ── 3. Process rows ───────────────────────────────────────────────────────
    for row_idx, row in df.iterrows():
        row_num  = row_idx + 2
        row_data = {}

        for excel_col, field in col_map.items():
            if excel_col not in df.columns:
                continue

            raw = row.get(excel_col)

            if field.startswith("_skip"):
                continue

            if field in date_fields:
                row_data[field] = parse_excel_date(raw)
                continue

            if field in time_fields:
                if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
                    try:
                        t = pd.to_datetime(str(raw), format="%H:%M", errors="coerce")
                        if t is not None and not pd.isna(t):
                            row_data[field] = t.time()
                        else:
                            t2 = pd.to_datetime(str(raw), errors="coerce")
                            row_data[field] = t2.time() if t2 and not pd.isna(t2) else None
                    except Exception:
                        row_data[field] = None
                else:
                    row_data[field] = None
                continue

            if field in ("_company_name", "_branch_name"):
                row_data[field] = safe_str(raw)
                continue

            if field == "blood_group":
                row_data[field] = normalize_blood_group(raw)
                continue

            if field in ("emergency_contact_relation1", "emergency_contact_relation2"):
                row_data[field] = normalize_relation(raw)
                continue

            if field == "gender":
                val = safe_str(raw)
                row_data[field] = "Male" if val.lower() == "male" else ("Female" if val.lower() == "female" else "")
                continue

            if field == "status":
                val = safe_str(raw)
                if val.lower() in ("left", "resigned"):
                    row_data[field] = "Left"
                elif val.lower() == "pending":
                    row_data[field] = "Pending"
                else:
                    row_data[field] = "Active"
                continue

            val = safe_str(raw)
            row_data[field] = "" if val == "0" else val

        # Store employee_code as None when blank to avoid unique constraint clash
        emp_code = str(row_data.get("employee_code", "")).strip() or None
        row_data["employee_code"] = emp_code

        # ── Resolve Company & Branch ──────────────────────────────────────────
        company_name = row_data.pop("_company_name", "")
        branch_name  = row_data.pop("_branch_name", "")

        company = None
        if company_name:
            company = Company.objects.filter(name__iexact=company_name).first()
            if not company:
                # try partial match
                company = Company.objects.filter(name__icontains=company_name.split()[0]).first()

        branch = None
        if branch_name:
            branch = Branch.objects.filter(branch_name__iexact=branch_name).first()

        # ── Save employee ─────────────────────────────────────────────────────
        try:
            status_val = row_data.pop("status", "Active") or "Active"

            emp = Employee(
                company=company,
                branch=branch,
                status=status_val,
                **{k: v for k, v in row_data.items() if not k.startswith("_")},
            )
            emp.save()

            # auto-create linked User if employee_code exists
            if emp_code:
                username = emp_code.lower()
                user, _ = User.objects.get_or_create(username=username)
                if _:
                    user.set_unusable_password()
                    user.save()
                emp.user = user
                emp.force_password_change = True
                emp.save(update_fields=["user", "force_password_change"])

            created_count += 1
            print(f"[IMPORT] Row {row_num}: created employee '{emp_code}'")

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print(f"[IMPORT] Row {row_num} save error: {tb}")
            errors.append({"row": row_num, "errors": [str(e)]})
            skipped_count += 1

    print(f"[IMPORT] Done — created: {created_count}, skipped: {skipped_count}, errors: {len(errors)}")

    return JsonResponse({
        "success":  True,
        "created":  created_count,
        "skipped":  skipped_count,
        "errors":   errors,
    })

    
# ─── Salary Excel Import ──────────────────────────────────────────────────────

@login_required
@group_required("Admin", "HR")
def download_salary_import_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary"

    headers = [
        "Employee Code*",
        "Gross CTC Monthly*", "Basic Monthly*", "HRA Monthly*",
        "Stat Bonus Monthly", "Allowance 1 Monthly", "Allowance 2 Monthly",
        "Special Allowance Monthly", "Guaranteed Cash Monthly",
        "PF Employer Monthly", "PF Employee Monthly",
        "ESIC Employer Monthly", "ESIC Employee Monthly",
        "Gratuity Monthly", "Profession Tax Monthly",
        "CTC Monthly", "Net Salary Monthly*",
        "Include PF (yes/no)", "Include ESIC (yes/no)", "Include Gratuity (yes/no)",
        "Effective Date (YYYY-MM-DD)",
    ]

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 20)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="salary_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
@group_required("Admin", "HR")
def import_salary_excel(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST required"}, status=405)

    uploaded_file = request.FILES.get("salary_file")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "No file uploaded."})

    try:
        df = pd.read_excel(uploaded_file, dtype=str)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Cannot read file: {e}"})

    df.columns = df.columns.str.strip()

    def safe_decimal(val):
        try:
            cleaned = str(val).strip().replace(",", "")
            if cleaned in ("", "nan", "NaT", "None"):
                return Decimal("0.00")
            return Decimal(cleaned)
        except Exception:
            return Decimal("0.00")

    created_count = 0
    skipped_count = 0
    errors = []

    from django.db import transaction

    for row_idx, row in df.iterrows():
        row_num = row_idx + 2
        row_errors = []

        emp_code = str(row.get("Employee Code*", "")).strip()
        if not emp_code or emp_code in ("nan", ""):
            errors.append({"row": row_num, "errors": ["Employee Code: required"]})
            skipped_count += 1
            continue

        try:
            employee = Employee.objects.get(employee_code=emp_code)
        except Employee.DoesNotExist:
            errors.append({"row": row_num, "errors": [f"Employee '{emp_code}' not found"]})
            skipped_count += 1
            continue

        def get(col):
            return str(row.get(col, "")).strip()

        pf_deducted = get("Include PF (yes/no)").lower() == "yes"
        esic_applicable = get("Include ESIC (yes/no)").lower() == "yes"
        gratuity_applicable = get("Include Gratuity (yes/no)").lower() == "yes"

        effective_date = None
        eff_raw = get("Effective Date (YYYY-MM-DD)")
        if eff_raw and eff_raw not in ("nan", ""):
            try:
                effective_date = pd.to_datetime(eff_raw).date()
            except Exception:
                row_errors.append("Effective Date: invalid date format")

        if row_errors:
            errors.append({"row": row_num, "errors": row_errors})
            skipped_count += 1
            continue

        pm_fields = {
            "gross_ctc_pm": safe_decimal(get("Gross CTC Monthly*")),
            "basic_pm": safe_decimal(get("Basic Monthly*")),
            "hra_pm": safe_decimal(get("HRA Monthly*")),
            "stat_bonus_pm": safe_decimal(get("Stat Bonus Monthly")),
            "allowance1_pm": safe_decimal(get("Allowance 1 Monthly")),
            "allowance2_pm": safe_decimal(get("Allowance 2 Monthly")),
            "sp_allowance_pm": safe_decimal(get("Special Allowance Monthly")),
            "guaranteed_cash_pm": safe_decimal(get("Guaranteed Cash Monthly")),
            "pf_er_cont_pm": safe_decimal(get("PF Employer Monthly")),
            "pf_ee_cont_pm": safe_decimal(get("PF Employee Monthly")),
            "esic_er_cont_pm": safe_decimal(get("ESIC Employer Monthly")),
            "esic_ee_cont_pm": safe_decimal(get("ESIC Employee Monthly")),
            "gratuity_pm": safe_decimal(get("Gratuity Monthly")),
            "profession_tax_pm": safe_decimal(get("Profession Tax Monthly")),
            "ctc_pm": safe_decimal(get("CTC Monthly")),
            "net_salary_pm": safe_decimal(get("Net Salary Monthly*")),
        }

        pa_fields = {k.replace("_pm", "_pa"): v * 12 for k, v in pm_fields.items()}

        try:
            with transaction.atomic():
                SalaryMaster.objects.filter(employee=employee, is_active=True).update(is_active=False)
                sm = SalaryMaster(
                    employee=employee,
                    pf_deducted=pf_deducted,
                    esic_applicable=esic_applicable,
                    gratuity_applicable=gratuity_applicable,
                    effective_date=effective_date,
                    is_active=True,
                )
                for k, v in {**pm_fields, **pa_fields}.items():
                    setattr(sm, k, v)
                sm.save()
            created_count += 1
        except Exception as e:
            errors.append({"row": row_num, "errors": [str(e)]})
            skipped_count += 1

    return JsonResponse({
        "success": True,
        "created": created_count,
        "skipped": skipped_count,
        "errors": errors,
    })

    return working_days